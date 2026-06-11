import os
import re
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog, Canvas
import pandas as pd
from datetime import datetime
import sys
import json
import sqlite3 # Importar la librería sqlite3 directamente
import webbrowser
from collections import defaultdict
import threading
import traceback
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
import subprocess
import requests
import sys
import ctypes
import psutil

class LoginAuditorDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Acceso a Calidad")
        self.geometry("350x250")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.result = None

        # Centrar la ventana en la pantalla
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - 175
        y = parent.winfo_y() + (parent.winfo_height() // 2) - 125
        self.geometry(f"+{x}+{y}")

        ttk.Label(self, text="🔒 Área Restringida", font=("Helvetica", 14, "bold"), foreground="#2C3E50").pack(pady=(15, 5))

        ttk.Label(self, text="Nombre del Auditor:", font=("Helvetica", 10, "bold")).pack(pady=(10, 2))
        self.user_var = tk.StringVar()
        self.user_entry = ttk.Entry(self, textvariable=self.user_var, font=("Helvetica", 11), width=25)
        self.user_entry.pack()
        self.user_entry.focus()

        ttk.Label(self, text="Contraseña:", font=("Helvetica", 10, "bold")).pack(pady=(10, 2))
        self.pass_var = tk.StringVar()
        self.pass_entry = ttk.Entry(self, textvariable=self.pass_var, font=("Helvetica", 11), width=25, show="*")
        self.pass_entry.pack()
        
        # Permitir entrar presionando la tecla Enter
        self.pass_entry.bind("<Return>", lambda e: self.check_login())

        ttk.Button(self, text="Ingresar", command=self.check_login, style='success.TButton').pack(pady=20)

    def check_login(self):
        user = self.user_var.get().strip()
        password = self.pass_var.get()

        if not user:
            messagebox.showwarning("Error", "Por favor ingresa tu nombre.", parent=self)
            return
        if password != "Calidad2024":
            messagebox.showerror("Acceso Denegado", "Contraseña incorrecta.", parent=self)
            return

        self.result = user # Guardamos el nombre del auditor
        self.destroy()

# =======================================================================================
# ======================= MOTOR DE INTEGRACIÓN CLOUD (FEISHU / LARK) ====================
# =======================================================================================
class FeishuIntegrator:
    def __init__(self, app_id, app_secret, app_token, table_id):
        self.app_id = app_id
        self.app_secret = app_secret
        self.app_token = app_token
        self.table_id = table_id
        # NOTA: Si tu empresa usa la versión global es larksuite.com. Si es la versión china es feishu.cn
        self.base_url = "https://open.feishu.cn/open-apis" # (Versión China)

    def get_tenant_access_token(self):
        url = f"{self.base_url}/auth/v3/tenant_access_token/internal"
        payload = {"app_id": self.app_id, "app_secret": self.app_secret}
        headers = {"Content-Type": "application/json; charset=utf-8"}
        
        response = requests.post(url, json=payload, headers=headers)
        if response.status_code == 200 and response.json().get("code") == 0:
            return response.json().get("tenant_access_token")
        else:
            raise Exception(f"Fallo de Autenticación Feishu: {response.text}")

    def create_bitable_record(self, campos_dict):
        token = self.get_tenant_access_token()
        url = f"{self.base_url}/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records"
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json; charset=utf-8"
        }
        
        # Estructura requerida por la API de Feishu
        payload = {"fields": campos_dict}
        
        response = requests.post(url, json=payload, headers=headers)
        if response.status_code == 200 and response.json().get("code") == 0:
            return True, "Registro creado exitosamente en Bitable."
        else:
            raise Exception(f"Error al escribir en Bitable: {response.text}")

try:
    import winsound
except ImportError:
    print("Librería 'winsound' no encontrada. No se reproducirán sonidos (solo disponible en Windows).")
    winsound = None

__version__ = "1.2.33" # IMPORTANTE: Esta es la versión de tu script local

# Reemplaza 'tu-usuario' y 'tu-repositorio' con los tuyos
URL_VERSION = "https://raw.githubusercontent.com/ZombPool/P-11-Sistema-verificaci-n-de-datos/main/version.txt"
URL_SCRIPT = "https://github.com/ZombPool/P-11-Sistema-verificaci-n-de-datos/releases/download/1.2.33/interfaz.exe"
# --- Dependencias Requeridas 
# Intenta importar xlrd para archivos .xls (Geometría antigua)
try:
    import xlrd
except ImportError:
    messagebox.showwarning("Dependencia Faltante", 
                           "La librería 'xlrd' no está instalada. Es necesaria para leer archivos .xls.\n\n"
                           "Para instalarla, ejecuta: pip install xlrd")

# Intenta importar openpyxl para archivos .xlsx (Reportes y resultados nuevos)
try:
    import openpyxl
except ImportError:
    messagebox.showwarning("Dependencia Faltante", 
                           "La librería 'openpyxl' no está instalada. Es necesaria para leer y escribir archivos .xlsx.\n\n"
                           "Para instalarla, ejecuta: pip install openpyxl")

# Para un mejor aspecto visual, usaremos ttkbootstrap.
try:
    import ttkbootstrap as ttk
    from ttkbootstrap.constants import *
except ImportError:
    messagebox.showerror("Librería Faltante", 
                         "Por favor, instala la librería 'ttkbootstrap' para la interfaz gráfica:\n\npip install ttkbootstrap")
    sys.exit()

def is_osk_running():
    """Verifica si el teclado en pantalla (osk.exe) ya está en ejecución."""
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] == "osk.exe":
            return True
    return False

def open_keyboard(event=None):
    """Abre el teclado en pantalla pidiendo elevación de administrador."""
    if not is_osk_running():
        try:
            # Este método le pide a Windows que ejecute 'osk.exe' usando el verbo 'runas',
            # que significa "Ejecutar como Administrador".
            # Esto mostrará la ventana de confirmación (UAC) si es necesario.
            ctypes.windll.shell32.ShellExecuteW(
                None,       # hwnd
                "runas",    # lpOperation
                "osk.exe",  # lpFile
                None,       # lpParameters
                None,       # lpDirectory
                1           # nShowCmd
            )
        except Exception as e:
            # Esto puede ocurrir si el usuario hace clic en "No" en la ventana de permisos.
            # El código de error para "operación cancelada por el usuario" es 1223.
            if hasattr(e, 'winerror') and e.winerror == 1223:
                print("El usuario canceló la solicitud de permisos para abrir el teclado.")
            else:
                messagebox.showerror("Error de Teclado", f"No se pudo iniciar el teclado virtual: {e}")

# --- CLASES DE ANÁLISIS INTEGRADAS ---

class AnalisisILRL:
    def __init__(self, master=None):
        self.master = master

    def abrir_archivo(self, ruta_archivo):
        try:
            if sys.platform == 'win32':
                os.startfile(ruta_archivo)
            elif sys.platform == 'darwin':
                subprocess.run(['open', ruta_archivo], check=True)
            else:
                subprocess.run(['xdg-open', ruta_archivo], check=True)
        except Exception as e:
            print(f"Advertencia: No se pudo abrir el archivo automáticamente: {e}")

    def extraer_clave(self, archivo):
        base = os.path.splitext(os.path.basename(archivo))[0]
        # Patrón para nombres cortos (ej. JMO-2512000030001)
        patron_nuevo = r'J(?:R)?MO-(\d{9})(\d{4})'
        m_nuevo = re.match(patron_nuevo, base, re.IGNORECASE)
        if m_nuevo:
            return f"{m_nuevo.group(1)}-{m_nuevo.group(2)}"
        # Patrón para nombres largos (ej. JMO-251200003-SC-LC-0001)
        patron_anterior = r'J(?:R)?MO-(\d{9})-(?:LC|SC|SCLC|LCSC|SC-LC|LC-SC)(?:-[KF])?-(\d{4})'
        m_anterior = re.match(patron_anterior, base, re.IGNORECASE)
        if m_anterior:
            return f"{m_anterior.group(1)}-{m_anterior.group(2)}"
        return None

    def leer_resultado_y_fecha(self, ruta):
        try:
            df = pd.read_excel(ruta, header=None)
            inicio_filas_datos = 12
            col7_vals = df.iloc[inicio_filas_datos:, 7].dropna().astype(str).str.upper()
            col8_vals = df.iloc[inicio_filas_datos:, 8].dropna().astype(str).str.upper()
            count_col7 = col7_vals.isin(['PASS', 'FAIL']).sum()
            count_col8 = col8_vals.isin(['PASS', 'FAIL']).sum()
            col_resultado = 8 if count_col8 >= count_col7 else 7
            col_fecha = 10 if col_resultado == 8 else 9
            resultados_raw = df.iloc[inicio_filas_datos:, col_resultado].dropna().astype(str).str.upper().tolist()
            fechas_raw = df.iloc[inicio_filas_datos:, col_fecha].dropna().tolist()
            if not resultados_raw:
                return None, None, None
            resultado_final = 'APROBADO' if all(r == 'PASS' for r in resultados_raw) else 'RECHAZADO'
            fechas_datetime = [pd.to_datetime(f) for f in fechas_raw if f]
            ultima_fecha = max(fechas_datetime).strftime("%d/%m/%Y %H:%M") if fechas_datetime else ''
            detalle_str = ", ".join([f"C{i+1} {r}" for i, r in enumerate(resultados_raw)])
            return resultado_final, ultima_fecha, detalle_str
        except Exception:
            return None, None, None

    # --- MODIFICACIÓN PRINCIPAL: Acepta lista de carpetas ---
    def analizar_carpetas_ilrl(self, lista_carpetas, total_esperado, update_progress_callback=None):
        archivos_excel = []
        
        # 1. Recolectar archivos de TODAS las carpetas encontradas
        for carpeta in lista_carpetas:
            if os.path.isdir(carpeta):
                files = [os.path.join(dirpath, f) for dirpath, _, filenames in os.walk(carpeta) for f in filenames if f.endswith('.xlsx')]
                archivos_excel.extend(files)

        if not archivos_excel:
            return {}, [], {}, ["No se encontraron archivos .xlsx en las rutas de la O.T."]

        agrupados = defaultdict(list)
        errores_lectura = []
        
        # 2. Procesar todos los archivos encontrados (mismo lógica de antes)
        total_archivos = len(archivos_excel)
        for i, ruta_completa in enumerate(archivos_excel):
            archivo_nombre = os.path.basename(ruta_completa)
            clave = self.extraer_clave(archivo_nombre)
            if not clave:
                errores_lectura.append(f"'{archivo_nombre}': No se pudo extraer clave. Ignorado.")
                continue
            resultado, fecha, detalle = self.leer_resultado_y_fecha(ruta_completa)
            if resultado is None:
                errores_lectura.append(f"'{archivo_nombre}': Error al leer. Ignorado.")
                continue
            
            # Guardamos la ruta para referencia
            agrupados[clave].append({'archivo': archivo_nombre, 'ruta': ruta_completa, 'resultado': resultado, 'fecha': fecha, 'detalle': detalle})
            
            if update_progress_callback:
                update_progress_callback((i + 1) / total_archivos * 100)
        
        resultados_finales = {}
        rechazados = []
        
        # 3. Consolidar: Si hay duplicados (ej. en carpeta 1 y carpeta 2), toma el más reciente
        for clave, items in agrupados.items():
            items_filtrados = [i for i in items if i['fecha']]
            if not items_filtrados:
                mas_reciente = items[-1]
            else:
                mas_reciente = max(items_filtrados, key=lambda x: datetime.strptime(x['fecha'], "%d/%m/%Y %H:%M"))
            
            resultados_finales[clave] = {'resultado': mas_reciente['resultado'], 'fecha': mas_reciente['fecha'], 'archivos': [i['archivo'] for i in items]}
            
            if mas_reciente['resultado'] == 'RECHAZADO':
                rechazados.append([clave, 'RECHAZADO', mas_reciente['detalle'], mas_reciente['archivo'], mas_reciente['fecha']])
        
        return resultados_finales, rechazados, agrupados, errores_lectura

    def generar_reporte_excel_ilrl(self, resultados, total_esperado, rechazados, ruta_destino_base, agrupados):
        carpeta_analisis = os.path.join(ruta_destino_base, "ANALISIS DE O.T")
        os.makedirs(carpeta_analisis, exist_ok=True)
        nombre_archivo_reporte = "Reporte_Analisis_ILRL.xlsx"
        nueva_ruta_destino = os.path.join(carpeta_analisis, nombre_archivo_reporte)
        
        if os.path.exists(nueva_ruta_destino):
            try:
                os.remove(nueva_ruta_destino)
            except Exception as e:
                return None, f"No se pudo eliminar el archivo anterior. Asegúrate de que no esté abierto.\nError: {e}"
        
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Resultados OT"
        ws1.append(['Cable', 'Resultado Final', 'Última Fecha de Medición', 'Archivos Analizados', 'Rutas Completas'])
        
        def sort_key_numeric_cable_id(cable_str):
            parts = cable_str.split('-')
            return int(parts[1]) if len(parts) >= 2 and parts[1].isdigit() else cable_str
        
        for cable in sorted(resultados.keys(), key=sort_key_numeric_cable_id):
            info = resultados[cable]
            ws1.append([cable, info['resultado'], info['fecha'], ", ".join(info['archivos']), ", ".join([i['ruta'] for i in agrupados[cable]])])
        
        if ws1.max_row > 1:
            tabla1 = Table(displayName="TablaResultados", ref=f"A1:E{ws1.max_row}")
            tabla1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws1.add_table(tabla1)
        
        ws2 = wb.create_sheet("Cantidades de cables")
        orden_trabajo_prefijo = ""
        if resultados:
            first_key = next(iter(resultados.keys()))
            match = re.match(r'(\d{9})-\d{4}', first_key)
            if match:
                orden_trabajo_prefijo = match.group(1) + "-"
        
        esperados_formato_completo_set = {f"{orden_trabajo_prefijo}{str(i).zfill(4)}" for i in range(1, total_esperado + 1)} if orden_trabajo_prefijo else set()
        claves_encontradas_set = set(resultados.keys())
        faltantes = sorted(list(esperados_formato_completo_set - claves_encontradas_set))
        
        ws2.append(['Total Esperado', total_esperado])
        ws2.append(['Encontrados', len(claves_encontradas_set)])
        ws2.append(['Faltantes', len(faltantes)])
        ws2.append([])
        ws2.append(['Listado de Faltantes (sugerido)'])
        for f in faltantes:
            ws2.append([f])
        
        ws3 = wb.create_sheet("Cables rechazados")
        ws3.append(['Cable', 'Estado', 'Detalle', 'Archivo Fuente', 'Fecha de Medición'])
        for r in rechazados:
            ws3.append(r)
            
        for ws in [ws1, ws2, ws3]:
            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col if cell.value) if any(c.value for c in col) else 0
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
        
        try:
            wb.save(nueva_ruta_destino)
            return nueva_ruta_destino, None
        except PermissionError:
            return None, f"No se pudo guardar el archivo. Asegúrate de que no esté abierto:\n{nueva_ruta_destino}"
        except Exception as e:
            return None, f"Error inesperado al guardar el archivo: {e}"

    # --- MÉTODO PROCESAR ACTUALIZADO ---
    def procesar_ilrl(self, ot_number, config, total_esperado, update_progress_callback=None):
        # 1. Identificar rutas
        rutas_posibles = []
        if config.get('ruta_base_ilrl'):
            rutas_posibles.append(os.path.join(config['ruta_base_ilrl'], ot_number))
        if config.get('ruta_base_ilrl_2'):
            rutas_posibles.append(os.path.join(config['ruta_base_ilrl_2'], ot_number))
            
        rutas_validas = [r for r in rutas_posibles if os.path.isdir(r)]
        
        if not rutas_validas:
             return None, f"No se encontró la carpeta de la OT '{ot_number}' en ninguna de las rutas configuradas."

        # 2. Analizar múltiples carpetas
        resultados, rechazados, agrupados, errores_lectura = self.analizar_carpetas_ilrl(rutas_validas, total_esperado, update_progress_callback)
        
        if not resultados and not errores_lectura:
            return None, "No se encontraron datos válidos para el análisis IL/RL."
        
        # Usamos la primera ruta válida como base para guardar el reporte
        ruta_reporte, error_guardado = self.generar_reporte_excel_ilrl(resultados, total_esperado, rechazados, rutas_validas[0], agrupados)
        
        if error_guardado:
            return None, error_guardado
        return ruta_reporte, "\n".join(errores_lectura) if errores_lectura else None

class AnalisisGEO:
    def __init__(self, master=None):
        self.master = master

    def abrir_archivo(self, ruta_archivo):
        try:
            if sys.platform == 'win32':
                os.startfile(ruta_archivo)
            elif sys.platform == 'darwin':
                subprocess.run(['open', ruta_archivo], check=True)
            else:
                subprocess.run(['xdg-open', ruta_archivo], check=True)
        except Exception as e:
            raise RuntimeError(f"No se pudo abrir el archivo automáticamente: {e}")

    # ... (Mantén los métodos 'ajustar_columnas' y 'aplicar_estilos' igual que antes) ...
    def ajustar_columnas(self, ws):
        column_widths = {'A': 22, 'B': 15, 'C': 40, 'D': 25}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    def aplicar_estilos(self, ws):
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                if cell.column_letter == 'B':
                    cell.alignment = center_alignment
                    if cell.value == "ACEPTADO":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    elif cell.value == "RECHAZADO":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical='center')

    # --- LÓGICA MULTI-ARCHIVO ---
    def analizar_archivos_geo_multi(self, lista_archivos, total_esperado, update_progress_callback=None, order_id_to_filter=None, mode="Duplex"):
        if not lista_archivos:
            return {}, [], 0, 0, ["No se proporcionaron archivos para analizar."]

        errores_lectura = []
        puntas_requeridas = ['1', '2'] if mode == "Simplex" else ['1', '2', '3', '4']
        
        # Estructura maestra: { serial_base: { punta_id: { resultado: ..., es_retrabajo: ..., timestamp: ... } } }
        datos_globales = {} 
        seriales_encontrados_set = set()

        total_files = len(lista_archivos)
        
        # 1. Leer y Fusionar todos los archivos
        for i, file_path in enumerate(lista_archivos):
            try:
                df = pd.read_excel(file_path, header=None, skiprows=12)
                
                # Función auxiliar regex (misma que en verif. individual)
                def parse_serial(s):
                    if not isinstance(s, str): return None, None
                    s = s.strip().upper()
                    match = re.search(r'(J(?:R)?MO-?\d{13}|\d{13})(-?([1-4R][1-4]?))?', s)
                    if match:
                         base = re.sub(r'[^0-9]', '', match.group(1))
                         punta = match.group(3) if match.group(3) else "N/A"
                         return base, punta
                    # Fallback simple
                    match_simple = re.search(r'(\d{13})', s)
                    if match_simple: return match_simple.group(1), "N/A"
                    return None, None

                for _, row in df.iterrows():
                    if len(row) < 7: continue
                    clave_base, punta = parse_serial(str(row[0]))
                    
                    if clave_base and punta != "N/A":
                        # Filtrar por O.T. si se requiere (evitar leer basura de otros cables si el archivo está mezclado)
                        if order_id_to_filter:
                            ot_clean = re.sub(r'[^0-9]', '', order_id_to_filter)
                            if not clave_base.startswith(ot_clean):
                                continue

                        seriales_encontrados_set.add(clave_base)
                        
                        resultado = str(row[6]).strip().upper() if pd.notna(row[6]) else "SIN_DATO"
                        # Intentar leer fecha
                        ts = None
                        if pd.notna(row[3]) and pd.notna(row[4]):
                             ts = pd.to_datetime(f"{row[3]} {row[4]}", errors='coerce')
                        
                        punta_limpia = punta.replace('R', '')
                        es_retrabajo = 'R' in punta
                        
                        if clave_base not in datos_globales:
                            datos_globales[clave_base] = {}
                        
                        # --- LÓGICA DE PRIORIDAD (Fusión) ---
                        # Si la punta no existe, o si es un Retrabajo y lo que teníamos NO era retrabajo -> Guardar/Sobrescribir
                        if punta_limpia not in datos_globales[clave_base]:
                            datos_globales[clave_base][punta_limpia] = {'res': resultado, 'rework': es_retrabajo, 'ts': ts}
                        else:
                            existente = datos_globales[clave_base][punta_limpia]
                            if es_retrabajo and not existente['rework']:
                                datos_globales[clave_base][punta_limpia] = {'res': resultado, 'rework': es_retrabajo, 'ts': ts}
                            # Si ambos son R o ambos normales, quedarse con el más reciente (si hay fecha)
                            elif es_retrabajo == existente['rework'] and ts and existente['ts'] and ts > existente['ts']:
                                datos_globales[clave_base][punta_limpia] = {'res': resultado, 'rework': es_retrabajo, 'ts': ts}

            except Exception as e:
                errores_lectura.append(f"Error leyendo {os.path.basename(file_path)}: {e}")
            
            if update_progress_callback:
                update_progress_callback((i + 1) / total_files * 80)

        # 2. Generar Reporte Final
        reporte_data = []
        
        # Calcular faltantes generales
        orden_trabajo_prefijo = ""
        if order_id_to_filter:
             # Formato JMO-XXXX... -> XXXX...-
             match = re.search(r'(\d{9})', order_id_to_filter)
             if match: orden_trabajo_prefijo = match.group(1) + "-"
        
        series_esperadas_set = {f"{orden_trabajo_prefijo}{str(i).zfill(4)}" for i in range(1, total_esperado + 1)} if orden_trabajo_prefijo else set()
        
        # Ojo: seriales_encontrados_set tiene el formato numérico (2512...). Necesitamos asegurarnos de comparar peras con peras.
        # series_esperadas_set también es "2512...-0001".
        # Vamos a formatear las encontradas para match (esto asume que clave_base ya viene como 2512000010001)
        claves_encontradas_formateadas = set()
        for cb in seriales_encontrados_set:
            # Insertar guion si es necesario para matchear formato esperado? 
            # El formato esperado es OT-Consecutivo (ej. 251200003-0001). 
            # El clave_base es 13 dígitos (ej. 2512000030001).
            if len(cb) == 13:
                cf = cb[:9] + "-" + cb[9:]
                claves_encontradas_formateadas.add(cf)
            else:
                claves_encontradas_formateadas.add(cb)

        faltantes_claves = sorted(list(series_esperadas_set - claves_encontradas_formateadas))

        # 3. Evaluar cada cable encontrado
        processed_count = 0
        total_items = len(datos_globales)

        for clave_base, puntas_data in datos_globales.items():
            # Formato visual OT-Consec
            clave_visual = clave_base[:9] + "-" + clave_base[9:] if len(clave_base) == 13 else clave_base
            
            estado_final = "ACEPTADO"
            detalles_puntas = []
            max_ts = pd.NaT

            for p_num in puntas_requeridas:
                if p_num in puntas_data:
                    data = puntas_data[p_num]
                    res_punta = "PASS" if data['res'] == 'PASS' else "FAIL"
                    tipo = "(R)" if data['rework'] else ""
                    detalles_puntas.append(f"P{p_num}{tipo}: {res_punta}")
                    
                    if data['res'] != 'PASS': estado_final = "RECHAZADO"
                    if data['ts'] and (pd.isna(max_ts) or data['ts'] > max_ts): max_ts = data['ts']
                else:
                    detalles_puntas.append(f"P{p_num}: FALTANTE")
                    estado_final = "RECHAZADO"

            reporte_data.append({
                'Número de Serie (OT-Cable)': clave_visual,
                'Estado Final': estado_final,
                'Detalle Puntas': ', '.join(detalles_puntas),
                'Última Medición': max_ts
            })
            
            processed_count += 1
            if update_progress_callback and total_items > 0:
                 update_progress_callback(80 + (processed_count / total_items) * 20)

        return reporte_data, faltantes_claves, len(datos_globales), len(faltantes_claves), errores_lectura

    def generar_reporte_excel_geo(self, reporte_data, faltantes_claves, total_esperado, total_encontrados, total_faltantes, ruta_destino_base, order_id=""):
        carpeta_analisis = os.path.join(ruta_destino_base, "ANALISIS DE O.T")
        os.makedirs(carpeta_analisis, exist_ok=True)
        nombre_archivo_reporte = f"AnalisisGEO_{order_id}.xlsx" if order_id and order_id != "Desconocida" else "AnalisisGEO_Reporte.xlsx"
        nueva_ruta_destino = os.path.join(carpeta_analisis, nombre_archivo_reporte)
        
        if os.path.exists(nueva_ruta_destino):
            try:
                os.remove(nueva_ruta_destino)
            except Exception as e:
                return None, f"No se pudo eliminar el archivo anterior.\nError: {e}"
        
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Resultados GEO"
        df_reporte = pd.DataFrame(reporte_data)
        if not df_reporte.empty:
            # Ordenar por consecutivo
            df_reporte['Orden_Cable'] = df_reporte['Número de Serie (OT-Cable)'].apply(lambda x: int(x.split('-')[1]) if '-' in x and x.split('-')[1].isdigit() else 0)
            df_reporte.sort_values(by='Orden_Cable', inplace=True)
            df_reporte.drop(columns=['Orden_Cable'], inplace=True)
            
            ws1.append(df_reporte.columns.tolist())
            for row in df_reporte.to_numpy().tolist():
                ws1.append(row)
            self.aplicar_estilos(ws1)
            self.ajustar_columnas(ws1)
            if ws1.max_row > 1:
                tabla1 = Table(displayName="TablaResultadosGEO", ref=f"A1:{get_column_letter(ws1.max_column)}{ws1.max_row}")
                tabla1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                ws1.add_table(tabla1)
        else:
            ws1.append(['Número de Serie (OT-Cable)', 'Estado Final', 'Detalle Puntas', 'Última Medición'])

        ws2 = wb.create_sheet("Resumen GEO")
        ws2.append(['Métrica', 'Valor'])
        ws2.append(['Total esperado', total_esperado])
        ws2.append(['Total encontrados', total_encontrados])
        ws2.append(['Total faltantes', total_faltantes])
        ws2.append([])
        ws2.append(['Cables Faltantes'])
        for f in faltantes_claves:
            ws2.append([f])
        
        for col in ws2.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value) if any(c.value for c in col) else 0
            ws2.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
            
        try:
            wb.save(nueva_ruta_destino)
            return nueva_ruta_destino, None
        except PermissionError:
            return None, f"No se pudo guardar el archivo.\n{nueva_ruta_destino}"
        except Exception as e:
            return None, f"Error inesperado al guardar: {e}"

    # --- MÉTODO PROCESAR ACTUALIZADO ---
    def procesar_geo(self, ot_number, config, total_esperado, update_progress_callback=None, mode="Duplex"):
        # 1. Buscar archivos en TODAS las rutas configuradas
        rutas_base = []
        if config.get('ruta_base_geo'): rutas_base.append(config['ruta_base_geo'])
        if config.get('ruta_base_geo_2'): rutas_base.append(config['ruta_base_geo_2'])
        
        archivos_candidatos = []
        for ruta_base in rutas_base:
            if os.path.isdir(ruta_base):
                encontrados = [os.path.join(ruta_base, f) for f in os.listdir(ruta_base) 
                               if f.lower().endswith(('.xlsx', '.xls')) 
                               and not f.startswith('~$') 
                               and ot_number in f] # Filtro simple por nombre
                # Ojo: si hay varios archivos de la misma OT en una carpeta (ej. versiones viejas),
                # aquí los añadimos todos. La función de análisis ya se encarga de priorizar por fecha/retrabajo.
                archivos_candidatos.extend(encontrados)
        
        if not archivos_candidatos:
            return None, f"No se encontraron archivos para la OT '{ot_number}' en las rutas configuradas."

        # 2. Analizar y Fusionar
        reporte_data, faltantes, total_encontrados, total_faltantes, errores = self.analizar_archivos_geo_multi(
            archivos_candidatos, total_esperado, update_progress_callback, order_id_to_filter=ot_number, mode=mode
        )

        if not reporte_data and errores:
             return None, f"Errores de lectura: {errores[0]}"
        if not reporte_data:
             return None, "No se encontraron datos válidos."

        # 3. Generar Excel
        # Usamos la primera carpeta configurada como destino principal
        ruta_destino_base = config['ruta_base_geo'] if os.path.isdir(config['ruta_base_geo']) else os.path.dirname(archivos_candidatos[0])
        
        ruta_reporte, error_guardado = self.generar_reporte_excel_geo(
            reporte_data, faltantes, total_esperado, total_encontrados, total_faltantes, ruta_destino_base, order_id=ot_number
        )
        
        if error_guardado: return None, error_guardado
        return ruta_reporte, "\n".join(errores) if errores else None


class App(ttk.Window):
    def __init__(self):
        super().__init__(themename="litera")
        self.title("FibraTrace - Sistema de Trazabilidad")
        self.geometry("1200x800")

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        self.cable_mode = tk.StringVar(value="Duplex")

        # --- Configuración Centralizada ---
        self.config_file = "config.json"
        self.password = "admin123" # Contraseña para funciones de administrador
        self.config = self.load_config()
        self.init_database()
        
        self.create_sidebar()
        self.create_main_content_area()

        self.pages = {}
        self.create_pages()
        self.show_page("Dashboard")

        self.bind_class("TEntry", "<FocusIn>", open_keyboard)

    # ... (dentro de la clase App) ...

    def check_for_updates(self):
        """Verifica si hay una nueva versión disponible en GitHub."""
        try:
            response = requests.get(URL_VERSION, timeout=5)
            if response.status_code == 200:
                remote_version = response.text.strip()
                
                # Comparamos las versiones (una simple comparación de strings funciona para "1.0.1" > "1.0.0")
                if remote_version > __version__:
                    if messagebox.askyesno("Actualización Disponible", 
                                          f"Hay una nueva versión ({remote_version}) disponible.\n"
                                          f"Tu versión actual es {__version__}.\n\n"
                                          "¿Deseas descargarla y reiniciar la aplicación ahora?"):
                        self.apply_update()
                else:
                    messagebox.showinfo("Sin Actualizaciones", "Ya estás utilizando la versión más reciente.")
            else:
                messagebox.showerror("Error de Verificación", f"No se pudo verificar la versión (Código: {response.status_code}).")
        except requests.RequestException as e:
            messagebox.showerror("Error de Conexión", f"No se pudo conectar a GitHub para buscar actualizaciones.\n\nError: {e}")

    def apply_update(self):
        """Descarga el nuevo ejecutable y ejecuta el actualizador."""
        try:
            # Mostramos un mensaje de que la descarga está en proceso
            self.page_title_label.config(text="Descargando actualización...")
            self.update_idletasks()

            response = requests.get(URL_SCRIPT, timeout=60) # Aumentamos a 60s por si la conexión es lenta
            if response.status_code == 200:
                # --- CORRECCIÓN 1: El archivo temporal ahora es un .exe ---
                nuevo_script_path = "interfaz_new.exe"
                with open(nuevo_script_path, "wb") as f:
                    f.write(response.content)

                # Obtenemos el nombre del ejecutable actual (ej. "interfaz.exe")
                script_actual = os.path.basename(sys.argv[0])

                # --- CORRECCIÓN 2: Ejecutamos updater.exe directamente ---
                # Esto asume que updater.exe está en la misma carpeta que interfaz.exe
                subprocess.Popen(['updater.exe', script_actual, nuevo_script_path])

                # Cerramos la aplicación actual para que el updater pueda trabajar
                self.destroy()

            else:
                messagebox.showerror("Error de Descarga", f"No se pudo descargar la actualización (Código: {response.status_code}).")
                self.page_title_label.config(text=self.pages["Dashboard"].__class__.__name__) # Restaura el título a la página actual

        except requests.RequestException as e:
            messagebox.showerror("Error de Conexión", f"Ocurrió un error al descargar el archivo.\n\nError: {e}")
            self.page_title_label.config(text=self.pages["Dashboard"].__class__.__name__)
        except Exception as e:
            messagebox.showerror("Error Inesperado", f"Ocurrió un error al aplicar la actualización:\n\n{e}")
            self.page_title_label.config(text=self.pages["Dashboard"].__class__.__name__)

    # Dentro de la clase App
    def load_config(self):
        """Carga la configuración desde config.json o crea un archivo por defecto."""
        default_config = {
            "linea_actual": "JWS1-1",  # <--- NUEVO: Identificador de la PC
            "ruta_base_ilrl": "C:\\Ruta\\Por\\Defecto\\ILRL_SC_LC",
            "ruta_base_ilrl_2": "", 
            "ruta_base_geo": "C:\\Ruta\\Por\\Defecto\\Geometria_SC_LC",
            "ruta_base_geo_2": "",
            "ruta_base_ilrl_mpo": "C:\\Ruta\\Por\\Defecto\\ILRL_MPO",
            "ruta_base_geo_mpo": "C:\\Ruta\\Por\\Defecto\\Geometria_MPO",
            "ruta_base_polaridad_mpo": "C:\\Ruta\\Por\\Defecto\\Polaridad_MPO",
            "ruta_base_geo_fanout_lc": "C:\\Ruta\\Por\\Defecto\\Geometria_Fanout_LC",
            
            # --- NUEVAS RUTAS PARA UNIBOOT ---
            "ruta_base_ilrl_uniboot": "",
            "ruta_base_geo_uniboot": "",
            # ---------------------------------
            
            "check_mpo_ilrl": True,
            "check_mpo_geo": True,
            "check_mpo_polaridad": True,
            
            "db_path": os.path.join(os.path.expanduser('~'), 'Documents', 'FibraTraceData', 'verifications.db'),
            "db_path_jws1_1": "",
            "db_path_jws1_2": "",
            "db_path_jws1_3": ""
        }
        if os.path.exists(self.config_file):
            with open(self.config_file, 'r') as f:
                config = json.load(f)
                for key, value in default_config.items():
                    config.setdefault(key, value)
                return config
        else:
            self.save_config(default_config)
            return default_config

    def save_config(self, config_data):
        """Guarda la configuración actual en config.json."""
        with open(self.config_file, 'w') as f:
            json.dump(config_data, f, indent=4)
        self.config = config_data # Actualizar la configuración en memoria

    def init_database(self):
        """Inicializa la base de datos y crea/actualiza las tablas necesarias."""
        try:
            db_dir = os.path.dirname(self.config['db_path'])
            os.makedirs(db_dir, exist_ok=True)
            conn = sqlite3.connect(self.config['db_path'])
            cursor = conn.cursor()

            # --- Tabla de Verificaciones ---
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS cable_verifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, entry_date TEXT NOT NULL,
                    serial_number TEXT NOT NULL, ot_number TEXT NOT NULL,
                    overall_status TEXT NOT NULL, ilrl_status TEXT,
                    ilrl_details TEXT, geo_status TEXT, geo_details TEXT
                )
            """)
            # Añadir columnas de MPO si no existen
            cursor.execute("PRAGMA table_info(cable_verifications)")
            columns = [info[1] for info in cursor.fetchall()]
            if 'polaridad_status' not in columns:
                cursor.execute("ALTER TABLE cable_verifications ADD COLUMN polaridad_status TEXT")
            if 'polaridad_details' not in columns:
                cursor.execute("ALTER TABLE cable_verifications ADD COLUMN polaridad_details TEXT")
            if 'digital_seal' not in columns:
                cursor.execute("ALTER TABLE cable_verifications ADD COLUMN digital_seal TEXT")

            # --- Tabla de Configuraciones de OT para MPO ---
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS ot_configurations (
                    ot_number TEXT PRIMARY KEY, drawing_number TEXT, link TEXT,
                    num_conectores_a INTEGER, fibers_per_connector_a INTEGER,
                    num_conectores_b INTEGER, fibers_per_connector_b INTEGER,
                    ilrl_ot_header TEXT, ilrl_serie_header TEXT,
                    ilrl_fecha_header TEXT, ilrl_hora_header TEXT,
                    ilrl_estado_header TEXT, ilrl_conector_header TEXT,
                    last_modified TEXT
                )
            """)
            
            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("Error de Base de Datos", f"No se pudo inicializar la base de datos: {e}")

    # En la clase App, reemplaza este método completo:
    def create_sidebar(self):
        # Frame principal de la barra lateral
        self.sidebar_frame = ttk.Frame(self, style='secondary.TFrame', padding=(0, 10))
        self.sidebar_frame.grid(row=0, column=0, sticky="ns")

        # Título y separador
        ttk.Label(self.sidebar_frame, text="FibraTrace", font=("Helvetica", 18, "bold"), style='inverse-secondary.TLabel').pack(pady=(10, 5))
        ttk.Separator(self.sidebar_frame).pack(fill='x', pady=10, padx=15)

        # Botón Dashboard (Siempre visible en la parte superior)
        ttk.Button(self.sidebar_frame, text="🏠 Dashboard", command=lambda: self.show_page("Dashboard"), style='primary.TButton').pack(fill='x', pady=5, padx=15)

        # =====================================================================
        # FUNCIÓN AUXILIAR: Crea menús estilo "Acordeón"
        # =====================================================================
        def crear_seccion(titulo, botones_info, incluir_switch=False, color_estilo='info.TButton'):
            # Botón principal (Cabecera)
            head_btn = ttk.Button(self.sidebar_frame, text=f"▶ {titulo}", style=color_estilo)
            head_btn.pack(fill='x', pady=(10, 0), padx=15)
            
            # Contenedor de las sub-opciones (inicia oculto)
            sub_frame = ttk.Frame(self.sidebar_frame, style='secondary.TFrame')
            
            # Lógica para mostrar/ocultar
            def toggle(frame=sub_frame, btn=head_btn):
                if frame.winfo_ismapped():
                    frame.pack_forget()
                    btn.config(text=btn.cget('text').replace('▼', '▶'))
                else:
                    frame.pack(fill='x', pady=5, padx=25, after=btn) # after=btn lo coloca justo debajo
                    btn.config(text=btn.cget('text').replace('▶', '▼'))

            head_btn.config(command=toggle)
            
            # (Opcional) Agregar el switch de Simplex/Duplex dentro del acordeón
            if incluir_switch:
                switch_frame = ttk.Frame(sub_frame, style='secondary.TFrame')
                switch_frame.pack(fill='x', pady=(5, 10))
                ttk.Label(switch_frame, text="Simplex", style='inverse-secondary.TLabel', font=("Helvetica", 8)).pack(side='left')
                self.mode_switch = ttk.Checkbutton(switch_frame, bootstyle="primary-round-toggle", variable=self.cable_mode, onvalue="Duplex", offvalue="Simplex")
                self.mode_switch.pack(side='left', padx=5)
                ttk.Label(switch_frame, text="Duplex", style='inverse-secondary.TLabel', font=("Helvetica", 8)).pack(side='left')

            # Crear los botones internos con estilo "Outline" para diferenciarlos
            for btn_texto, icono, comando in botones_info:
                btn = ttk.Button(sub_frame, text=f"{icono} {btn_texto}", command=comando, style='primary.outline.TButton')
                btn.pack(fill='x', pady=3)
                
            return sub_frame, head_btn

        # =====================================================================
        # CREANDO LAS SECCIONES DE PRODUCCIÓN
        # =====================================================================
        
        crear_seccion("SC & LC", [
            ("Verificación Individual", "🔍", lambda: self.show_page("Verificacion_LC_SC")),
            ("Revisar Lote", "📦", lambda: self.show_page("RevisarLote_LC_SC"))
        ], incluir_switch=True)

        crear_seccion("MPO", [
            ("Verificación Individual", "🔍", lambda: self.show_page("Verificacion_MPO")),
            ("Análisis de O.T.", "📊", lambda: self.show_page("Reportes_MPO")),
            ("Registro en Almacén", "📦", lambda: self.show_page("RegistroWHMPO"))
        ])

        crear_seccion("FANOUT", [
            ("Verificación Individual", "🔍", lambda: self.show_page("VerificacionFanout")),
            ("Liberación de Lote", "📊", lambda: self.show_page("ReportesFanout")),
            ("Registro en Almacén", "📦", lambda: self.show_page("RegistroWHFanout"))
        ])

        crear_seccion("UNIBOOT", [
            ("Verificación Individual", "🔍", lambda: self.show_page("VerificacionUniboot")),
        ])

        crear_seccion("Herramientas del auditor", [
            ("Buscar Sello Digital", "🔎", lambda: self.show_page("BuscadorSellos")),
            ("Auditoría O.T. (LC/SC)", "📊", lambda: self.show_page("Auditoria_LC_SC"))
        ])
        
        # =====================================================================
        # ZONA DE ADMINISTRACIÓN (Al fondo de la pantalla)
        # =====================================================================
        
        # Espaciador invisible que empuja todo lo de abajo hacia el final de la pantalla
        ttk.Frame(self.sidebar_frame, style='secondary.TFrame').pack(fill='both', expand=True)
        ttk.Separator(self.sidebar_frame).pack(fill='x', pady=5, padx=15)

        # Sección de herramientas con un color gris más sobrio (secondary)
        sub_tools, btn_tools = crear_seccion("ADMINISTRACIÓN", [
            ("Configurar Rutas", "⚙️", self.open_settings_window),
            ("Ver Registros DB", "📋", lambda: self.show_page("Registros")),
            ("Diagnóstico DB", "🗄️", self.show_db_diagnostics),
            ("Actualizaciones", "🔄", self.check_for_updates)
        ], color_estilo='secondary.TButton')
        
        # Reposicionamos el botón base para que se ancle al fondo
        btn_tools.pack(side='bottom', fill='x', pady=(0, 20), padx=15)
    def create_main_content_area(self):
        self.main_frame = ttk.Frame(self, padding=20)
        self.main_frame.grid(row=0, column=1, sticky="nsew")
        self.main_frame.grid_rowconfigure(1, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)
        
        header_frame = ttk.Frame(self.main_frame)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        self.page_title_label = ttk.Label(header_frame, text="Dashboard", font=("Helvetica", 24, "bold"))
        self.page_title_label.pack(side='left')

    def create_pages(self):
        self.pages["Dashboard"] = DashboardPage(self.main_frame, self) # <--- Pasamos self (app)
        self.pages["Verificacion_LC_SC"] = VerificacionLC_SC_Page(self.main_frame, self)
        #self.pages["Reportes_LC_SC"] = Reportes_LC_SC_Page(self.main_frame, self)
        self.pages["RevisarLote_LC_SC"] = RevisarLote_LC_SC_Page(self.main_frame, self)
        self.pages["Registros"] = RecordsPage(self.main_frame, self)
        self.pages["Verificacion_MPO"] = VerificacionMPO_Page(self.main_frame, self)
        self.pages["Reportes_MPO"] = AnalisisMPOPage(self.main_frame, self)
        #self.pages["RegistroWH"] = RegistroWH_Page(self.main_frame, self)
        self.pages["RegistroWHMPO"] = RegistroWHMPO_Page(self.main_frame, self)
        self.pages["VerificacionFanout"] = VerificacionFanout_Page(self.main_frame, self)
        self.pages["ReportesFanout"] = AnalisisFanoutPage(self.main_frame, self)
        self.pages["RegistroWHFanout"] = RegistroWHFanout_Page(self.main_frame, self)
        self.pages["VerificacionUniboot"] = VerificacionUniboot_Page(self.main_frame, self)
        self.pages["BuscadorSellos"] = BuscadorSellos_Page(self.main_frame, self)
        self.pages["Auditoria_LC_SC"] = Auditoria_LC_SC_Page(self.main_frame, self)

    def show_page(self, page_name):
        # --- NUEVO: CANDADO DE SEGURIDAD PARA HERRAMIENTAS DE CALIDAD ---
        paginas_protegidas = ["Auditoria_LC_SC", "BuscadorSellos"] 
        
        if page_name in paginas_protegidas:
            # Si el auditor no se ha logueado en esta sesión, le pedimos la contraseña
            if not getattr(self, 'auditor_name', None):
                dialog = LoginAuditorDialog(self)
                self.wait_window(dialog)
                
                # Si cierra la ventana o falla, abortamos y no lo dejamos entrar
                if not dialog.result:
                    return 
                
                # Si es exitoso, guardamos su nombre en la memoria del programa
                self.auditor_name = dialog.result
        title_map = {
            "Verificacion_LC_SC": "Verificación Individual (LC/SC)",
            "RevisarLote_LC_SC": "Revisar Lote de Producción (LC/SC)",
            "Reportes_LC_SC": "Análisis de O.T. (LC/SC)",
            "Verificacion_MPO": "Verificación Individual (MPO)",
            "Reportes_MPO": "Análisis de O.T. (MPO)",
            "VerificacionUniboot": "Verificación Individual (Uniboot)",
            "BuscadorSellos": "Buscador de Sellos Digitales",
            "Auditoria_LC_SC": "Auditoría Automática (LC/SC)"
        }
        display_title = title_map.get(page_name, page_name.replace("_", " "))
        self.page_title_label.config(text=display_title)
        
        for page in self.pages.values():
            page.grid_forget()
        
        page = self.pages[page_name]
        page.grid(row=1, column=0, sticky="nsew")
        if isinstance(page, RecordsPage):
            page.load_records()

    def request_password(self, callback):
        password_ingresada = simpledialog.askstring("Contraseña Requerida", 
                                                    "Ingrese la contraseña de administrador:", 
                                                    show='*', parent=self)
        if password_ingresada == self.password:
            callback()
        else:
            messagebox.showerror("Acceso Denegado", "Contraseña incorrecta.")
        
    def open_settings_window(self):
        self.request_password(lambda: SettingsWindow(self))

    def show_db_diagnostics(self):
        db_path = self.config['db_path']
        size = os.path.getsize(db_path) if os.path.exists(db_path) else 0
        messagebox.showinfo("Diagnóstico de Base de Datos",
                            f"La base de datos se está guardando en:\n\n{db_path}\n\n"
                            f"Tamaño del archivo: {size} bytes")

    def guardar_ot_configuration(self, ot_data):
        """Guarda o actualiza la configuración de una OT en la base de datos."""
        try:
            conn = sqlite3.connect(self.config['db_path'])
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO ot_configurations (
                    ot_number, drawing_number, link, num_conectores_a,
                    fibers_per_connector_a, num_conectores_b, fibers_per_connector_b,
                    ilrl_ot_header, ilrl_serie_header, ilrl_fecha_header,
                    ilrl_hora_header, ilrl_estado_header, ilrl_conector_header, last_modified
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(ot_number) DO UPDATE SET
                    drawing_number=excluded.drawing_number, link=excluded.link,
                    num_conectores_a=excluded.num_conectores_a, fibers_per_connector_a=excluded.fibers_per_connector_a,
                    num_conectores_b=excluded.num_conectores_b, fibers_per_connector_b=excluded.fibers_per_connector_b,
                    ilrl_ot_header=excluded.ilrl_ot_header, ilrl_serie_header=excluded.ilrl_serie_header,
                    ilrl_fecha_header=excluded.ilrl_fecha_header, ilrl_hora_header=excluded.ilrl_hora_header,
                    ilrl_estado_header=excluded.ilrl_estado_header, ilrl_conector_header=excluded.ilrl_conector_header,
                    last_modified=excluded.last_modified
            """, (
                ot_data['ot_number'], ot_data['drawing_number'], ot_data['link'],
                ot_data['num_conectores_a'], ot_data['fibers_per_connector_a'],
                ot_data['num_conectores_b'], ot_data['fibers_per_connector_b'],
                ot_data.get('ilrl_ot_header', 'Work number'), ot_data.get('ilrl_serie_header', 'Serial number'),
                ot_data.get('ilrl_fecha_header', 'Date'), ot_data.get('ilrl_hora_header', 'Time'),
                ot_data.get('ilrl_estado_header', 'Alarm Status'), ot_data.get('ilrl_conector_header', 'connector label'),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            messagebox.showerror("Error de Base de Datos", f"No se pudo guardar la configuración: {e}")
            return False

# --- PÁGINAS DE LA APLICACIÓN ---

class DashboardPage(ttk.Frame):
    def __init__(self, parent, app_instance):
        super().__init__(parent, padding=10)
        self.app = app_instance
        self.create_widgets()

    def create_widgets(self):
        # --- AQUÍ ESTÁ LA CORRECCIÓN DEL PADDING ---
        container = ttk.Frame(self, style='TFrame', padding=10)
        container.pack(expand=True, fill='both')

        ttk.Label(container, text="Bienvenido al Sistema de Trazabilidad FibraTrace.", font=("Helvetica", 16)).pack(pady=20)
        ttk.Label(container, text="Selecciona una opción del menú de la izquierda para comenzar.", font=("Helvetica", 12)).pack(pady=10)

        # =========================================================================
        # --- SECCIÓN: CONTROL DE VALIDACIONES (BYPASS TEMPORAL) ---
        # =========================================================================
        bypass_frame = ttk.LabelFrame(container, text="⚙️ Control de Validaciones en Línea (Bypass)", padding=15)
        bypass_frame.pack(fill='x', pady=15, padx=10)

        ttk.Label(bypass_frame, text="Desactive los interruptores para omitir temporalmente la medición y forzar un 'APROBADO' directo.", 
                  font=("Helvetica", 10, "italic"), foreground="#555555").grid(row=0, column=0, columnspan=3, pady=(0, 10), sticky='w')

        lineas = ["JWS1-1", "JWS1-2", "JWS1-3"]
        
        for idx, linea in enumerate(lineas):
            linea_key = linea.lower().replace("-", "_")
            frame_linea = ttk.Frame(bypass_frame)
            frame_linea.grid(row=1, column=idx, padx=30, pady=5)
            
            ttk.Label(frame_linea, text=f"Línea {linea}", font=("Helvetica", 11, "bold")).pack(anchor='center', pady=(0, 5))
            
            # Variables de estado enlazadas a config
            setattr(self, f"var_ilrl_{linea_key}", tk.BooleanVar(value=self.app.config.get(f'val_ilrl_{linea_key}', True)))
            setattr(self, f"var_geo_{linea_key}", tk.BooleanVar(value=self.app.config.get(f'val_geo_{linea_key}', True)))
            
            var_ilrl = getattr(self, f"var_ilrl_{linea_key}")
            var_geo = getattr(self, f"var_geo_{linea_key}")
            
            def guardar_cambios_bypass(k=linea_key, v_i=var_ilrl, v_g=var_geo):
                self.app.config[f'val_ilrl_{k}'] = v_i.get()
                self.app.config[f'val_geo_{k}'] = v_g.get()
                self.app.save_config(self.app.config)
            
            chk_ilrl = ttk.Checkbutton(frame_linea, text="IL/RL (Activo)", variable=var_ilrl, bootstyle="success-round-toggle", command=guardar_cambios_bypass)
            chk_ilrl.pack(anchor='w', pady=5)
            
            chk_geo = ttk.Checkbutton(frame_linea, text="Geometría (Activo)", variable=var_geo, bootstyle="success-round-toggle", command=guardar_cambios_bypass)
            chk_geo.pack(anchor='w', pady=5)
        # =========================================================================

        version_label = ttk.Label(self, text=f"Versión {__version__}", font=("Helvetica", 10), style='secondary.TLabel')
        version_label.pack(side='bottom', pady=10, anchor='se')

class VerificacionLC_SC_Page(ttk.Frame):
    def __init__(self, parent, app_instance):
        super().__init__(parent, padding=10)
        self.app = app_instance 
        self.last_ilrl_result = None 
        self.last_geo_result = None
        self.create_widgets()

    def create_widgets(self):
        container = ttk.Frame(self, style='TFrame', padding=20)
        container.pack(expand=True, fill='both')
        
        input_frame = ttk.Frame(container)
        input_frame.pack(fill='x', pady=10)
        
        ttk.Label(input_frame, text="Número de OT:", font=("Helvetica", 11, "bold")).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.ot_entry = ttk.Entry(input_frame, width=30, font=("Helvetica", 11))
        self.ot_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)
        
        ttk.Label(input_frame, text="Número de Serie (13 dígitos):", font=("Helvetica", 11, "bold")).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.serie_entry = ttk.Entry(input_frame, width=30, font=("Helvetica", 11))
        self.serie_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)
        self.serie_entry.bind("<KeyRelease>", self.verificar_cable_automatico)
        
        # --- NUEVO SELECTOR SANA EN VERIFICACIÓN ---
        ttk.Label(input_frame, text="Formato Geometría:", font=("Helvetica", 11, "bold")).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.sana_var = tk.StringVar(value="SANA 1.0")
        sana_cb = ttk.Combobox(input_frame, textvariable=self.sana_var, values=["SANA 1.0", "SANA 2.0"], state="readonly", width=28, font=("Helvetica", 11))
        sana_cb.grid(row=2, column=1, sticky='ew', padx=5, pady=5)
        # -------------------------------------------
        
        input_frame.columnconfigure(1, weight=1)

        verify_button = ttk.Button(container, text="Verificar Cable (Manual)", command=self.verificar_cable, style='success.TButton', padding=10)
        verify_button.pack(pady=20)

        self.result_text = tk.Text(container, height=15, width=80, wrap="word", font=("Courier New", 10), state=tk.DISABLED, relief="flat", bg="#f0f0f0")
        self.result_text.pack(fill='both', expand=True, pady=10)

        self.result_text.tag_configure("header", font=("Helvetica", 14, "bold"), foreground="#0056b3")
        self.result_text.tag_configure("bold", font=("Courier New", 10, "bold"))
        self.result_text.tag_configure("info", font=("Courier New", 10, "italic"), foreground="grey")
        self.result_text.tag_configure("APROBADO", foreground="#28a745")
        self.result_text.tag_configure("RECHAZADO", foreground="#dc3545")
        self.result_text.tag_configure("NO ENCONTRADO", foreground="#fd7e14")
        self.result_text.tag_configure("ERROR", foreground="#dc3545", font=("Courier New", 10, "bold"))
        self.result_text.tag_configure("ilrl_link", foreground="#0056b3", underline=True)
        self.result_text.tag_bind("ilrl_link", "<Button-1>", lambda e: self.show_details_window("ilrl"))
        self.result_text.tag_bind("ilrl_link", "<Enter>", lambda e: self.result_text.config(cursor="hand2"))
        self.result_text.tag_bind("ilrl_link", "<Leave>", lambda e: self.result_text.config(cursor=""))
        self.result_text.tag_configure("geo_link", foreground="#0056b3", underline=True)
        self.result_text.tag_bind("geo_link", "<Button-1>", lambda e: self.show_details_window("geo"))
        self.result_text.tag_bind("geo_link", "<Enter>", lambda e: self.result_text.config(cursor="hand2"))
        self.result_text.tag_bind("geo_link", "<Leave>", lambda e: self.result_text.config(cursor=""))

        self.result_text.tag_configure("ilrl_file_link", foreground="#4682B4", underline=True)
        self.result_text.tag_bind("ilrl_file_link", "<Button-1>", lambda e: self.open_file_location("ilrl"))
        self.result_text.tag_bind("ilrl_file_link", "<Enter>", lambda e: self.result_text.config(cursor="hand2"))
        self.result_text.tag_bind("ilrl_file_link", "<Leave>", lambda e: self.result_text.config(cursor=""))
        
        self.result_text.tag_configure("geo_file_link", foreground="#4682B4", underline=True)
        self.result_text.tag_bind("geo_file_link", "<Button-1>", lambda e: self.open_file_location("geo"))
        self.result_text.tag_bind("geo_file_link", "<Enter>", lambda e: self.result_text.config(cursor="hand2"))
        self.result_text.tag_bind("geo_file_link", "<Leave>", lambda e: self.result_text.config(cursor=""))

        self.result_text.tag_configure("ERROR", foreground="#dc3545", font=("Courier New", 10, "bold"))
        self.result_text.tag_configure("final_status_large", font=("Courier New", 14, "bold"))
        self.result_text.tag_configure("ilrl_link", foreground="#0056b3", underline=True)

        # --- NUEVA SECCIÓN DE SCRAP ---
        scrap_frame = ttk.LabelFrame(container, text="Registro de Scrap (Cables Dañados/Irreparables)", padding=15)
        scrap_frame.pack(fill='x', pady=(10, 0))

        ttk.Label(scrap_frame, text="N.S. a Scrapear (13 dígitos):", font=("Helvetica", 11, "bold")).grid(row=0, column=0, padx=5, pady=5)
        self.scrap_entry = ttk.Entry(scrap_frame, width=25, font=("Helvetica", 11))
        self.scrap_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Button(scrap_frame, text="🗑️ Mandar a Scrap", command=self.mandar_a_scrap, style='danger.TButton').grid(row=0, column=2, padx=15, pady=5)
        # ------------------------------

        self.show_waiting_message()

    def show_waiting_message(self):
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        self.result_text.insert(tk.END, "Esperando un número de serie valido (13 digitos)", "info")
        self.result_text.config(state=tk.DISABLED)

    def _log_verification(self, log_data):
        try:
            conn = sqlite3.connect(self.app.config['db_path'])
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO cable_verifications (
                    entry_date, serial_number, ot_number, overall_status,
                    ilrl_status, ilrl_details, geo_status, geo_details, digital_seal
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                log_data['serial_number'],
                log_data['ot_number'],
                log_data['overall_status'],
                log_data['ilrl_status'],
                json.dumps(log_data['ilrl_details']),
                log_data['geo_status'],
                json.dumps(log_data['geo_details']),
                log_data.get('digital_seal', 'N/A') # <--- Guardado en BD
            ))
            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("Error de Base de Datos", f"No se pudo registrar la verificación: {e}")

    def verificar_cable_automatico(self, event=None):
        serie_raw = self.serie_entry.get().strip()
        ot_numero = self.ot_entry.get().strip()
        
        # --- MODIFICACIÓN: Regex ajustado para aceptar JMO o JRMO ---
        if re.match(r'J(R)?MO\d{13}', serie_raw, re.IGNORECASE):
            # Si escanea un código completo, extraemos números y rellenamos OT si falta
            numeros_serie = re.sub(r'[^0-9]', '', serie_raw)
            if not ot_numero:
                 self.ot_entry.insert(0, f"JMO-{numeros_serie[:9]}")
            self.verificar_cable()
            
        elif len(serie_raw) == 13 and serie_raw.isdigit():
            # Caso manual de 13 dígitos
            self.verificar_cable()
        else:
            self.show_waiting_message()
    
    def mandar_a_scrap(self):
        serie_raw = self.scrap_entry.get().strip()
        
        # 1. Limpieza y validación de los 13 dígitos
        serie_numerica = re.sub(r'[^0-9]', '', serie_raw)

        if len(serie_numerica) != 13:
            messagebox.showerror("Formato Inválido", "El número de serie debe contener exactamente 13 dígitos para enviarse a Scrap.", parent=self)
            return

        # 2. Reconstrucción automática (Detectar OT a partir del N.S.)
        prefijo_serie = "JRMO-" if "JRMO" in serie_raw.upper() else "JMO-"
        serie_completa = f"{prefijo_serie}{serie_numerica}"
        ot_completa = f"JMO-{serie_numerica[:9]}" # Extraemos los primeros 9 dígitos para la O.T.

        # 3. Ventana de Advertencia (Poka-Yoke)
        confirmacion = messagebox.askyesno(
            "⚠️ Advertencia Crítica de Scrap",
            f"Estás a punto de catalogar el cable:\n\n{serie_completa}\n\ncomo SCRAP (Dañado/Irreparable).\n\n¿Estás completamente seguro de realizar esta acción?",
            parent=self,
            icon='warning'
        )

        if not confirmacion:
            return

        # 4. Registrar en la Base de Datos simulando el formato estándar (COMO PENDIENTE)
        log_data = {
            'serial_number': serie_completa,
            'ot_number': ot_completa,
            'overall_status': 'SCRAP PENDIENTE',
            'ilrl_status': 'N/A (SCRAP PENDIENTE)',
            'ilrl_details': {'status': 'SCRAP PENDIENTE', 'details': 'Pendiente de confirmación por Calidad.', 'raw_data': []},
            'geo_status': 'N/A (SCRAP PENDIENTE)',
            'geo_details': {'status': 'SCRAP PENDIENTE', 'details': 'Pendiente de confirmación por Calidad.', 'raw_data': []},
            'digital_seal': serie_completa
        }

        self._log_verification(log_data)

        # 5. Mostrar confirmación visual en la pantalla principal
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        self.result_text.insert(tk.END, f"Registro de SCRAP Exitoso\n", "header")
        self.result_text.insert(tk.END, "-"*70 + "\n\n")
        self.result_text.insert(tk.END, f"El cable {serie_completa} (perteneciente a la {ot_completa}) ha sido marcado como dañado y guardado en la Base de Datos.\n\n")
        
        self.result_text.insert(tk.END, "ESTADO FINAL: ", ("bold", "final_status_large"))
        self.result_text.insert(tk.END, "SCRAP\n", ("ERROR", "final_status_large")) # Usa el color rojo de ERROR
        self.result_text.insert(tk.END, f"SELLO DIGITAL:  {serie_completa}\n", "header")
        self.result_text.config(state=tk.DISABLED)

        # Limpiamos el recuadro para el siguiente cable
        self.scrap_entry.delete(0, tk.END)
        
        # Alerta sonora distinta (más grave y larga) para Scrap
        if winsound:
            try:
                winsound.Beep(300, 800)
            except: pass
            
        messagebox.showinfo("Scrap Registrado", f"El cable {serie_completa} se ha enviado a Scrap correctamente.", parent=self)

    # En la clase VerificacionLC_SC_Page, reemplaza este método:

    def verificar_cable(self, event=None):
        ot_numero = self.ot_entry.get().strip().upper()
        serie_raw = self.serie_entry.get().strip()
        
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)

        if not ot_numero or not serie_raw:
            self.result_text.insert(tk.END, "ERROR: Por favor ingrese OT y Número de Serie.", "ERROR")
            self.result_text.config(state=tk.DISABLED)
            return

        # 1. Normalización numérica
        serie_numerica = re.sub(r'[^0-9]', '', serie_raw)
        
        if len(serie_numerica) != 13:
            messagebox.showerror("Formato Inválido", "El número de serie debe contener 13 dígitos.")
            self.result_text.config(state=tk.DISABLED)
            return

        # 2. Construcción del Serial con Prefijo Correcto (JMO o JRMO)
        prefijo_serie = "JRMO-" if "JRMO" in serie_raw.upper() else "JMO-"
        serie_cable = f"{prefijo_serie}{serie_numerica}"
        
        # 3. Validación de coincidencia OT
        ot_parte_input = re.sub(r'[^0-9]', '', ot_numero)
        serie_ot_parte = serie_numerica[:9]

        if ot_parte_input != serie_ot_parte:
            messagebox.showerror("Error de Coincidencia", "La OT del número de serie no corresponde a la OT trabajada.")
            self.result_text.config(state=tk.DISABLED)
            return

        # =========================================================================
        # 4. NUEVA VALIDACIÓN POKA-YOKE: COMPROBAR SI EL CABLE ESTÁ EN SCRAP
        # =========================================================================
        try:
            conn = sqlite3.connect(self.app.config['db_path'], timeout=10)
            cursor = conn.cursor()
            # Buscamos el último registro de este cable en la Base de Datos
            cursor.execute("""
                SELECT overall_status FROM cable_verifications 
                WHERE serial_number LIKE ? OR digital_seal LIKE ? 
                ORDER BY id DESC LIMIT 1
            """, (f"%{serie_numerica}%", f"%{serie_numerica}%"))
            registro_previo = cursor.fetchone()
            conn.close()

            if registro_previo and registro_previo[0] == 'SCRAP':
                self.result_text.insert(tk.END, f"⚠️ ALERTA: EL CABLE {serie_cable} ESTÁ EN SCRAP ⚠️\n", "ERROR")
                self.result_text.insert(tk.END, "-"*60 + "\n\n")
                self.result_text.insert(tk.END, "Este número de serie fue marcado previamente como DAÑADO/IRREPARABLE.\nNo se puede volver a verificar ni procesar.\n\n", "bold")
                
                self.result_text.insert(tk.END, "ESTADO FINAL: ", ("bold", "final_status_large"))
                self.result_text.insert(tk.END, "SCRAP\n", ("ERROR", "final_status_large"))
                self.result_text.config(state=tk.DISABLED)
                
                if winsound:
                    try:
                        winsound.Beep(300, 800) # Sonido grave de advertencia
                    except: pass
                    
                messagebox.showerror("Cable Bloqueado", f"El cable {serie_cable} ya ha sido reportado como SCRAP en la base de datos.", parent=self)
                return # Abortamos la verificación aquí mismo
        except Exception as e:
            print(f"Error comprobando estado de scrap en BD: {e}")
        # =========================================================================

        current_mode = self.app.cable_mode.get()
        self.result_text.insert(tk.END, f"Verificando cable {serie_cable} en OT {ot_numero} (Modo: {current_mode})...\n", "header")
        self.result_text.insert(tk.END, "-"*60 + "\n\n")

        # =========================================================================
        # --- FILTRO DE BYPASS SEGÚN DASHBOARD ---
        # =========================================================================
        linea_pc = self.app.config.get('linea_actual', 'JWS1-1').lower().replace("-", "_")
        val_ilrl_activa = self.app.config.get(f'val_ilrl_{linea_pc}', True)
        val_geo_activa = self.app.config.get(f'val_geo_{linea_pc}', True)

        if val_ilrl_activa:
            self.last_ilrl_result = self.buscar_y_procesar_ilrl(ot_numero, serie_cable, current_mode)
        else:
            self.last_ilrl_result = {'status': 'APROBADO', 'details': 'BYPASS ACTIVO: Omitido desde Dashboard.', 'raw_data': []}

        if val_geo_activa:
            self.last_geo_result = self.buscar_y_procesar_geo(ot_numero, serie_cable, current_mode)
        else:
            self.last_geo_result = {'status': 'APROBADO', 'details': 'BYPASS ACTIVO: Omitido desde Dashboard.', 'raw_data': []}
        # =========================================================================
        
        self.mostrar_resultado("IL/RL", self.last_ilrl_result)
        self.mostrar_resultado("Geometría", self.last_geo_result)

        # --- Semáforo Final ---
        final_status = "NO ENCONTRADO"
        if self.last_ilrl_result['status'] not in ['NO ENCONTRADO', 'ERROR'] or self.last_geo_result['status'] not in ['NO ENCONTRADO', 'ERROR']:
            if self.last_ilrl_result['status'] == 'APROBADO' and self.last_geo_result['status'] == 'APROBADO':
                final_status = 'APROBADO'
            else:
                final_status = 'RECHAZADO'
        
        # --- CREACIÓN DEL SELLO DIGITAL ---
        sello_digital = serie_cable
        
        self.result_text.insert(tk.END, "\n" + "-"*60 + "\n")
        self.result_text.insert(tk.END, "ESTADO FINAL: ", ("bold", "final_status_large"))
        self.result_text.insert(tk.END, f"{final_status}\n", (final_status, "final_status_large"))
        
        self.result_text.insert(tk.END, f"SELLO DIGITAL:  {sello_digital}\n", "header")
        self.result_text.insert(tk.END, "-"*60 + "\n")
        
        if winsound:
            try:
                if final_status == "APROBADO":
                    winsound.Beep(1200, 200)
                elif final_status == "RECHAZADO":
                    winsound.Beep(400, 500)
            except Exception as e:
                print(f"No se pudo reproducir el sonido: {e}")
        
        self.result_text.config(state=tk.DISABLED)
        
        log_data = {
            'serial_number': serie_cable,
            'ot_number': ot_numero,
            'overall_status': final_status,
            'ilrl_status': self.last_ilrl_result['status'],
            'ilrl_details': self.last_ilrl_result,
            'geo_status': self.last_geo_result['status'],
            'geo_details': self.last_geo_result,
            'digital_seal': sello_digital 
        }
        self._log_verification(log_data)

    def mostrar_resultado(self, tipo, resultado):
        link_tag = "ilrl_link" if tipo == "IL/RL" else "geo_link"
        file_link_tag = "ilrl_file_link" if tipo == "IL/RL" else "geo_file_link" # <-- Nuevo tag

        self.result_text.insert(tk.END, f"Análisis {tipo}:\n", "bold")
        self.result_text.insert(tk.END, f"  Estado: ")
        
        # Aplicar el tag de ESTADO (para detalles)
        self.result_text.insert(tk.END, f"{resultado['status']}", (resultado['status'], link_tag))
        
        # --- INICIO DE LA MODIFICACIÓN ---
        details = resultado['details']
        if "Archivo:" in details:
            details_text, file_text = details.rsplit("Archivo:", 1)
            file_text = "Archivo:" + file_text
            
            self.result_text.insert(tk.END, f"\n  Detalles: {details_text.strip()}")
            # Aplicar el tag de ARCHIVO (para explorador)
            self.result_text.insert(tk.END, f"\n  {file_text.strip()}", (file_link_tag)) 
            self.result_text.insert(tk.END, "\n\n")
        else:
            self.result_text.insert(tk.END, f"\n  Detalles: {details}\n\n")
        # --- FIN DE LA MODIFICACIÓN ---

    def buscar_y_procesar_ilrl(self, ot, serie, mode):
        # Lista de rutas a buscar (Primaria y Secundaria)
        rutas_base = [self.app.config['ruta_base_ilrl']]
        if self.app.config.get('ruta_base_ilrl_2'):
            rutas_base.append(self.app.config['ruta_base_ilrl_2'])

        serie_terminacion = serie[-4:]
        candidatos = []

        # Buscar en todas las rutas configuradas
        for ruta_base in rutas_base:
            if not os.path.isdir(ruta_base): continue
            
            ruta_ot = os.path.join(ruta_base, ot)
            if os.path.isdir(ruta_ot):
                # Buscar archivos que coincidan con la terminación
                for root, _, files in os.walk(ruta_ot):
                    for f in files:
                        if f.endswith('.xlsx') and not f.startswith('~$') and serie_terminacion in f:
                            full_path = os.path.join(root, f)
                            candidatos.append(full_path)

        if not candidatos:
            return {'status': 'NO ENCONTRADO', 'details': f'Ningún archivo con terminación "{serie_terminacion}" en las rutas configuradas.', 'raw_data': []}
        
        # Si hay duplicados o archivos en ambas rutas, tomamos el más reciente
        archivo_a_procesar = max(candidatos, key=os.path.getmtime)
        
        return self.procesar_archivo_ilrl(archivo_a_procesar, mode)

    def buscar_y_procesar_geo(self, ot, serie, mode):
        # 1. Identificar rutas configuradas
        rutas_base = [self.app.config['ruta_base_geo']]
        if self.app.config.get('ruta_base_geo_2'):
            rutas_base.append(self.app.config['ruta_base_geo_2'])

        archivos_candidatos = []

        # 2. Buscar archivos de la OT en TODAS las rutas
        for ruta_base in rutas_base:
            if not os.path.isdir(ruta_base): continue
            
            # Buscamos TODOS los archivos que contengan el número de OT en el nombre
            encontrados = [os.path.join(ruta_base, f) for f in os.listdir(ruta_base) 
                           if f.lower().endswith(('.xlsx', '.xls')) 
                           and not f.startswith('~$') 
                           and ot in f]
            
            if encontrados:
                # --- NUEVO: En lugar de tomar solo el más reciente, tomamos TODOS ---
                archivos_candidatos.extend(encontrados)

        if not archivos_candidatos:
            return {'status': 'NO ENCONTRADO', 'details': f'Ningún archivo para la OT "{ot}" en las rutas configuradas.', 'raw_data': []}

        # 3. Procesar TODOS los archivos encontrados y fusionar los datos
        return self.procesar_multiples_archivos_geo(archivos_candidatos, serie, mode)

    # En la clase VerificacionLC_SC_Page, reemplaza estos dos métodos:

    def procesar_archivo_ilrl(self, ruta, mode): # Ahora recibe 'mode'
        try:
            expected_count = 2 if mode == "Simplex" else 4 # Usa el 'mode' recibido
            
            df = pd.read_excel(ruta, header=None)
            rows = df.iloc[12:]
            # ... (el resto del código de esta función no cambia)
            col7_vals = rows[7].dropna().astype(str).str.upper()
            col8_vals = rows[8].dropna().astype(str).str.upper()
            count_col7 = col7_vals.isin(['PASS', 'FAIL']).sum()
            count_col8 = col8_vals.isin(['PASS', 'FAIL']).sum()
            result_col = 8 if count_col8 >= count_col7 else 7
            results = rows[result_col].dropna().astype(str).str.upper().tolist()
            valid_results = [r for r in results if r in ['PASS', 'FAIL']]
            if not valid_results: 
                return {'status': 'RECHAZADO', 'details': 'No se encontraron mediciones.', 'raw_data': []}
            all_pass = all(r == 'PASS' for r in valid_results)
            count_ok = len(valid_results) == expected_count
            status = 'APROBADO' if all_pass and count_ok else 'RECHAZADO'
            details = f"{len(valid_results)}/{expected_count} mediciones encontradas. "
            if not count_ok:
                details += f"Se esperaban {expected_count} para modo {mode}. "
            if not all_pass:
                details += f"{valid_results.count('FAIL')} mediciones con FALLA. "
            details += f"Archivo: {os.path.basename(ruta)}"
            raw_data = [{'linea': i + 1, 'resultado': res} for i, res in enumerate(valid_results)]
            return {'status': status, 'details': details, 'raw_data': raw_data, 'file_path': ruta}
        except Exception as e:
            return {'status': 'ERROR', 'details': f'Fallo al procesar {os.path.basename(ruta)}: {e}', 'raw_data': []}

    # En la clase VerificacionLC_SC_Page, reemplaza este método:

    def procesar_archivo_geo(self, ruta, serie_objetivo, mode):
        try:
            puntas_requeridas = ['1', '2'] if mode == "Simplex" else ['1', '2', '3', '4']

            df = pd.read_excel(ruta, header=None, skiprows=12)
            
            # Función auxiliar interna actualizada para JRMO
            def normalize_serial_geo(s):
                if not isinstance(s, str): return "", ""
                s_upper = s.strip().upper()
                # --- MODIFICACIÓN DE REGEX: Acepta JMO o JRMO ---
                match = re.search(r'(J(?:R)?MO-?\d{13}|\d{13})(-?([1-4R][1-4]?))?', s_upper)
                if match:
                    # Extraemos SOLO los números del serial base para comparar
                    base_serial_raw = match.group(1)
                    base_serial_numeric = re.sub(r'[^0-9]', '', base_serial_raw)
                    
                    punta = match.group(3) if match.group(3) else "N/A"
                    return base_serial_numeric, punta
                
                # Fallback: intentar extraer 13 digitos si el regex falla
                fallback_match = re.search(r'(\d{13})', s_upper)
                if fallback_match:
                     return fallback_match.group(1), "N/A"
                     
                return "", "N/A"

            # --- MODIFICACIÓN: Usamos solo la parte numérica del input para buscar ---
            serie_objetivo_norm = re.sub(r'[^0-9]', '', serie_objetivo)
            # -----------------------------------------------------------------------

            punta_results = []
            
            for _, row in df.iterrows():
                # raw_serial ahora será solo números (los 13 dígitos) gracias a la función de arriba
                raw_serial_numeric, punta = normalize_serial_geo(str(row[0]))
                
                # Comparamos números con números
                if serie_objetivo_norm == raw_serial_numeric and punta != "N/A":
                    result = str(row[6]).upper() if len(row) > 6 and pd.notna(row[6]) else "SIN_DATO"
                    punta_results.append({'punta': punta, 'resultado': result, 'fuente': str(row[0])})
            
            if not punta_results:
                return {'status': 'NO ENCONTRADO', 'details': 'No hay mediciones para este serial.', 'raw_data': []}

            found_puntas = {p['punta'].replace('R',''): p['resultado'] for p in punta_results}
            status = 'APROBADO'
            missing_puntas = []
            
            for p_req in puntas_requeridas:
                if p_req not in found_puntas:
                    status = 'RECHAZADO'
                    missing_puntas.append(p_req)
                elif found_puntas[p_req] != 'PASS':
                    status = 'RECHAZADO'
            
            pass_count = sum(1 for p in found_puntas.values() if p == 'PASS')
            details = f"{pass_count}/{len(puntas_requeridas)} puntas requeridas OK."
            
            if missing_puntas:
                details += f" Faltan: {', '.join(missing_puntas)}."
            details += f" Archivo: {os.path.basename(ruta)}"
            
            return {'status': status, 'details': details, 'raw_data': punta_results, 'file_path': ruta}

        except Exception as e:
            return {'status': 'ERROR', 'details': f'Fallo al procesar {os.path.basename(ruta)}: {e}', 'raw_data': []}

    def show_details_window(self, analysis_type):
        if analysis_type == "ilrl":
            data = self.last_ilrl_result
            title = "Detalles de Análisis IL/RL"
        elif analysis_type == "geo":
            data = self.last_geo_result
            title = "Detalles de Análisis de Geometría"
        else:
            return
        if not data or not data.get('raw_data'):
            messagebox.showinfo(title, "No hay datos detallados para mostrar.")
            return
        
        # --- LÍNEA AÑADIDA: Inyectamos el N/S desde el campo de texto ---
        data['serial_number'] = self.serie_entry.get().strip()
        # -----------------------------------------------------------
        
        DetailsWindow(self, title, data, analysis_type)

    # En la clase VerificacionLC_SC_Page, reemplaza este método:

    # En la clase VerificacionLC_SC_Page, reemplaza este método:

    def open_file_location(self, analysis_type):
        """Abre la carpeta que contiene el archivo de reporte."""
        if analysis_type == 'ilrl':
            data = self.last_ilrl_result
        elif analysis_type == 'geo':
            data = self.last_geo_result
        else:
            return

        if not data or not data.get('file_path'):
            messagebox.showinfo("Error", "No se encontró la ruta del archivo.", parent=self)
            return
            
        file_path = data['file_path']
        if not os.path.exists(file_path):
            messagebox.showerror("Error", f"La ruta del archivo no existe:\n{file_path}", parent=self)
            return
            
        try:
            # --- INICIO DE LA CORRECCIÓN ---
            # 1. Obtener el DIRECTORIO que contiene el archivo
            folder_path = os.path.dirname(file_path)
            
            # 2. Asegurarse de que la ruta sea absoluta
            path_absoluto = os.path.abspath(folder_path)

            if not os.path.isdir(path_absoluto):
                messagebox.showerror("Error", f"La ruta de la carpeta no existe:\n{path_absoluto}", parent=self)
                return

            # 3. Usar os.startfile() - es la forma más robusta en Windows
            os.startfile(path_absoluto)
            # --- FIN DE LA CORRECCIÓN ---
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la carpeta del archivo:\n{e}", parent=self)
    
    def procesar_multiples_archivos_geo(self, lista_archivos, serie_objetivo, mode):
        try:
            puntas_requeridas = ['1', '2'] if mode == "Simplex" else ['1', '2', '3', '4']
            sana_version = self.sana_var.get() # <--- Capturamos versión SANA

            serie_objetivo_norm = re.sub(r'[^0-9]', '', serie_objetivo)
            ot_completa = f"JMO-{serie_objetivo_norm[:9]}" 
            secuencial_buscado = serie_objetivo_norm[-4:] 
            
            todas_mediciones = [] 
            puntas_encontradas_map = {} 
            archivos_usados = set()

            for ruta in lista_archivos:
                try:
                    df = pd.read_excel(ruta, header=None, skiprows=12)
                    
                    for _, row in df.iterrows():
                        if len(row) < 9: continue

                        # --- NUEVA BIFURCACIÓN INTELIGENTE (Auto-detección) ---
                        s_upper = str(row[0]).strip().upper()
                        
                        # 1er Intento: Formato SANA 1.0
                        match_sana1 = re.search(r'(J(?:R)?MO-?\d{13}|\d{13})(-?([1-4R][1-4]?))?', s_upper)
                        
                        if match_sana1 and match_sana1.group(3):
                            base_serial_numeric = re.sub(r'[^0-9]', '', match_sana1.group(1))
                            punta_original = match_sana1.group(3)
                            fuente_texto = s_upper
                            resultado_raw = str(row[6]).upper() if len(row) > 6 and pd.notna(row[6]) else "FAIL"
                            
                        # 2do Intento: Formato SANA 2.0
                        elif ot_completa in s_upper:
                            sec_raw = re.sub(r'[^0-9]', '', str(row[1]).strip())
                            if not sec_raw: continue
                            base_serial_numeric = serie_objetivo_norm[:9] + sec_raw.zfill(4)
                            punta_original = str(row[2]).strip().upper()
                            fuente_texto = f"{s_upper} {sec_raw}"
                            resultado_raw = str(row[8]).upper() if len(row) > 8 and pd.notna(row[8]) else "FAIL"
                            
                        else:
                            continue # Si no es ningún formato, pasamos a la siguiente fila
                        # --------------------------------------------------------

                        if serie_objetivo_norm == base_serial_numeric and punta_original != "N/A":
                            resultado = "PASS" if "PASS" in resultado_raw else "FAIL"
                            
                            punta_limpia = punta_original.replace('R', '').replace('-', '')
                            es_retrabajo = 'R' in punta_original

                            datos_medicion = {
                                'punta_original': punta_original,
                                'punta_limpia': punta_limpia,
                                'resultado': resultado,
                                'fuente': fuente_texto,
                                'archivo': os.path.basename(ruta),
                                'es_retrabajo': es_retrabajo
                            }
                            
                            todas_mediciones.append(datos_medicion)
                            archivos_usados.add(os.path.basename(ruta))

                            if punta_limpia not in puntas_encontradas_map:
                                puntas_encontradas_map[punta_limpia] = datos_medicion
                            else:
                                existente = puntas_encontradas_map[punta_limpia]
                                if es_retrabajo and not existente['es_retrabajo']:
                                    puntas_encontradas_map[punta_limpia] = datos_medicion
                                elif es_retrabajo == existente['es_retrabajo']:
                                    puntas_encontradas_map[punta_limpia] = datos_medicion

                except Exception as e:
                    print(f"Error leyendo archivo {ruta}: {e}")

            # --- EVALUACIÓN FINAL ---
            if not puntas_encontradas_map:
                return {'status': 'NO ENCONTRADO', 'details': 'No hay mediciones para este serial en los archivos revisados.', 'raw_data': []}

            status = 'APROBADO'
            missing_puntas = []
            
            for p_req in puntas_requeridas:
                if p_req not in puntas_encontradas_map:
                    status = 'RECHAZADO'
                    missing_puntas.append(p_req)
                elif puntas_encontradas_map[p_req]['resultado'] != 'PASS':
                    status = 'RECHAZADO'
            
            pass_count = sum(1 for p in puntas_requeridas if p in puntas_encontradas_map and puntas_encontradas_map[p]['resultado'] == 'PASS')
            
            details = f"{pass_count}/{len(puntas_requeridas)} puntas OK."
            if missing_puntas: details += f" Faltan: {', '.join(missing_puntas)}."
            details += f" Fuentes: {', '.join(list(archivos_usados))}"
            
            raw_data_formatted = [{'punta': m['punta_original'], 'resultado': m['resultado'], 'fuente': f"{m['fuente']} ({m['archivo']})"} for m in todas_mediciones]
            ruta_principal = lista_archivos[0] if lista_archivos else ""

            return {'status': status, 'details': details, 'raw_data': raw_data_formatted, 'file_path': ruta_principal}

        except Exception as e:
            return {'status': 'ERROR', 'details': f'Fallo al procesar múltiples archivos: {e}', 'raw_data': []}

class RegistroWHMPO_Page(ttk.Frame):
    def __init__(self, parent, app_instance):
        super().__init__(parent, padding=20)
        self.app = app_instance

        # --- Creamos los widgets para esta página ---
        container = ttk.Frame(self, style='TFrame')
        container.pack(expand=True, fill='both', pady=20)

        # Un texto descriptivo
        ttk.Label(
            container,
            text="Módulo de Registro en Almacén (MPO)",
            font=("Helvetica", 16, "bold")
        ).pack(pady=10)

        ttk.Label(
            container,
            text="Presiona el botón para abrir el archivo de registro de series para empaque de productos MPO.",
            font=("Helvetica", 11),
            wraplength=500
        ).pack(pady=10)

        # El botón que abrirá el archivo de Excel
        open_button = ttk.Button(
            container,
            text="Abrir Registro WH (MPO)", # Texto del botón actualizado
            command=self.abrir_registro_mpo,
            style='success.TButton',
            padding=15
        )
        open_button.pack(pady=30)

    def abrir_registro_mpo(self):
        """
        Busca y abre el archivo de registro para MPO.
        Asumiremos un nombre de archivo como 'MPORegistroWH.xlsm'.
        """
        # Puedes cambiar 'MPORegistroWH.xlsm' por el nombre real de tu archivo de MPO
        file_name = "MPORegistroWH.xlsm"
        try:
            base_path = os.path.dirname(sys.argv[0])
            file_path = os.path.join(base_path, file_name)

            if os.path.exists(file_path):
                os.startfile(file_path)
            else:
                messagebox.showerror(
                    "Archivo no Encontrado",
                    f"No se pudo encontrar el archivo '{file_name}'.\n\nAsegúrate de que esté en la misma carpeta que el programa."
                )
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al intentar abrir el archivo:\n\n{e}")

# --- PÁGINA PARA ANÁLISIS DE O.T. (MPO) ---
class AnalisisMPOPage(ttk.Frame):
    def __init__(self, parent, app_instance):
        super().__init__(parent, padding=20)
        self.app = app_instance
        self.create_widgets()

    def create_widgets(self):
        container = ttk.Frame(self, style='TFrame')
        container.pack(expand=True, fill='both')

        # --- Frame de Configuración ---
        config_frame = ttk.LabelFrame(container, text="Parámetros del Análisis MPO", padding=15)
        config_frame.pack(fill='x', pady=(0, 20))
        config_frame.columnconfigure(1, weight=1)

        # ILRL File
        ttk.Label(config_frame, text="Archivo IL/RL MPO:", font="-weight bold").grid(row=0, column=0, sticky='w', pady=5, padx=5)
        self.ilrl_file_var = tk.StringVar()
        ttk.Entry(config_frame, textvariable=self.ilrl_file_var, state="readonly").grid(row=0, column=1, sticky='ew', pady=5, padx=5)
        ttk.Button(config_frame, text="Seleccionar Archivo...", command=self.select_ilrl_file, style='outline.TButton').grid(row=0, column=2, sticky='w', pady=5, padx=5)

        # GEO File
        ttk.Label(config_frame, text="Archivo Geometría MPO:", font="-weight bold").grid(row=1, column=0, sticky='w', pady=5, padx=5)
        self.geo_file_var = tk.StringVar()
        ttk.Entry(config_frame, textvariable=self.geo_file_var, state="readonly").grid(row=1, column=1, sticky='ew', pady=5, padx=5)
        ttk.Button(config_frame, text="Seleccionar Archivo...", command=self.select_geo_file, style='outline.TButton').grid(row=1, column=2, sticky='w', pady=5, padx=5)

        # Polarity Folder
        ttk.Label(config_frame, text="Carpeta Polaridad MPO:", font="-weight bold").grid(row=2, column=0, sticky='w', pady=5, padx=5)
        self.polaridad_folder_var = tk.StringVar()
        ttk.Entry(config_frame, textvariable=self.polaridad_folder_var, state="readonly").grid(row=2, column=1, sticky='ew', pady=5, padx=5)
        ttk.Button(config_frame, text="Seleccionar Carpeta...", command=self.select_polaridad_folder, style='outline.TButton').grid(row=2, column=2, sticky='w', pady=5, padx=5)
        
        ttk.Separator(config_frame, orient='horizontal').grid(row=3, column=0, columnspan=3, sticky='ew', pady=10)

        # OT and Quantity
        ttk.Label(config_frame, text="Número de O.T. (JMO-):", font="-weight bold").grid(row=4, column=0, sticky='w', pady=5, padx=5)
        self.ot_var = tk.StringVar()
        ttk.Entry(config_frame, textvariable=self.ot_var).grid(row=4, column=1, sticky='w', pady=5, padx=5)

        ttk.Label(config_frame, text="Total de cables esperados:", font="-weight bold").grid(row=5, column=0, sticky='w', pady=5, padx=5)
        self.total_cables_var = tk.StringVar()
        ttk.Entry(config_frame, textvariable=self.total_cables_var).grid(row=5, column=1, sticky='w', pady=5, padx=5)

        # --- Frame de Acciones ---
        action_frame = ttk.Frame(container)
        action_frame.pack(fill='x', pady=10)
        
        ttk.Button(action_frame, text="Generar Reporte IL/RL", command=lambda: self.run_analysis("ilrl"), style='info.TButton', padding=10).pack(side='left', padx=10, expand=True)
        ttk.Button(action_frame, text="Generar Reporte Geometría", command=lambda: self.run_analysis("geo"), style='info.TButton', padding=10).pack(side='left', padx=10, expand=True)
        ttk.Button(action_frame, text="Generar Reporte Polaridad", command=lambda: self.run_analysis("polaridad"), style='info.TButton', padding=10).pack(side='left', padx=10, expand=True)

        # --- Frame de Progreso y Resultados ---
        result_frame = ttk.LabelFrame(container, text="Estado del Análisis", padding=15)
        result_frame.pack(fill='both', expand=True, pady=(20, 0))

        self.progress_bar = ttk.Progressbar(result_frame, mode='determinate')
        self.progress_bar.pack(fill='x', pady=10)
        
        self.result_label = ttk.Label(result_frame, text="Listo para iniciar el análisis MPO.", wraplength=700)
        self.result_label.pack(fill='x', pady=10)

    def _select_file(self, var, title):
        file_selected = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_selected:
            var.set(file_selected)
            match = re.search(r'(JMO-\d{9})', os.path.basename(file_selected), re.IGNORECASE)
            if match and not self.ot_var.get():
                self.ot_var.set(match.group(1).upper())

    def select_ilrl_file(self):
        self._select_file(self.ilrl_file_var, "Seleccionar archivo de IL/RL MPO")

    def select_geo_file(self):
        self._select_file(self.geo_file_var, "Seleccionar archivo de Geometría MPO")
        
    def select_polaridad_folder(self):
        folder_selected = filedialog.askdirectory(title="Seleccionar carpeta de Polaridad (ej. JMO-250800005)")
        if folder_selected:
            self.polaridad_folder_var.set(folder_selected)
            match = re.search(r'JMO-(\d{9})', os.path.basename(folder_selected), re.IGNORECASE)
            if match and not self.ot_var.get():
                # Corrección para consistencia: sugerir sin prefijo
                self.ot_var.set(match.group(1))

    def run_analysis(self, analysis_type):
        ot_number = self.ot_var.get().strip().upper()
        if not ot_number.startswith("JMO-"):
            ot_number = f"JMO-{ot_number}"
            
        total_str = self.total_cables_var.get()

        # Validaciones comunes
        if not ot_number:
            messagebox.showerror("Error", "Por favor, ingresa el número de O.T.", parent=self)
            return
        if not total_str.isdigit() or int(total_str) <= 0:
            messagebox.showerror("Error", "Por favor, ingresa un número válido para el total de cables.", parent=self)
            return
        total = int(total_str)
        
        self.progress_bar["value"] = 0
        self.result_label.config(text=f"Iniciando análisis de {analysis_type.upper()} para MPO...")
        
        # Validaciones y arranque de hilos específicos
        if analysis_type == "ilrl":
            file_path = self.ilrl_file_var.get()
            if not file_path:
                messagebox.showerror("Error", "Por favor, selecciona el archivo de IL/RL para MPO.", parent=self)
                return
            thread = threading.Thread(target=self._run_ilrl_mpo_thread, args=(file_path, ot_number, total), daemon=True)
            thread.start()
        elif analysis_type == "geo":
            file_path = self.geo_file_var.get()
            if not file_path:
                messagebox.showerror("Error", "Por favor, selecciona el archivo de Geometría para MPO.", parent=self)
                return
            thread = threading.Thread(target=self._run_geo_mpo_thread, args=(file_path, ot_number, total), daemon=True)
            thread.start()
        elif analysis_type == "polaridad":
            folder_path = self.polaridad_folder_var.get()
            if not folder_path:
                messagebox.showerror("Error", "Por favor, selecciona la carpeta de Polaridad para MPO.", parent=self)
                return
            thread = threading.Thread(target=self._run_polaridad_mpo_thread, args=(folder_path, ot_number, total), daemon=True)
            thread.start()

    def _cargar_ot_configuration(self, ot_number):
        try:
            conn = sqlite3.connect(self.app.config['db_path'])
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM ot_configurations WHERE ot_number = ?", (ot_number,))
            row = cursor.fetchone()
            conn.close()
            return dict(row) if row else None
        except Exception as e:
            messagebox.showerror("Error de Base de Datos", f"No se pudo cargar la configuración de la OT: {e}", parent=self)
            return None

    def _run_ilrl_mpo_thread(self, file_path, ot_number, total_cables):
        try:
            self.update_progress(5)
            ot_config = self._cargar_ot_configuration(ot_number)
            if not ot_config:
                self.after(0, self.show_result, None, f"No se encontró configuración para la OT {ot_number}. Por favor, configúrela primero.", "IL/RL MPO")
                return

            self.update_progress(15)
            df = pd.read_excel(file_path, sheet_name="Results")
            h = {k: ot_config.get(v, d) for k, v, d in [('serie', 'ilrl_serie_header', 'Serial number'), ('estado', 'ilrl_estado_header', 'Alarm Status'), ('conector', 'ilrl_conector_header', 'connector label'), ('fecha', 'ilrl_fecha_header', 'Date'), ('hora', 'ilrl_hora_header', 'Time')]}
            
            if any(header not in df.columns for header in h.values()):
                missing = [k for k, v in h.items() if v not in df.columns]
                self.after(0, self.show_result, None, f"El archivo no contiene los encabezados esperados: {', '.join(missing)}.", "IL/RL MPO")
                return
            
            self.update_progress(30)
            df[h['serie']] = df[h['serie']].astype(str)
            df['timestamp'] = pd.to_datetime(df[h['fecha']].astype(str) + ' ' + df[h['hora']].astype(str), errors='coerce')
            all_series = df[h['serie']].dropna().unique()

            reporte_data = []
            rechazados = []
            
            num_conectores_a = ot_config.get('num_conectores_a', 1)
            fibras_por_conector_a = ot_config.get('fibers_per_connector_a', 12)
            num_conectores_b = ot_config.get('num_conectores_b', 1)
            fibras_por_conector_b = ot_config.get('fibers_per_connector_b', 12)
            total_fibras_esperadas = (num_conectores_a * fibras_por_conector_a) + (num_conectores_b * fibras_por_conector_b)

            for i, serie in enumerate(all_series):
                df_cable = df[df[h['serie']] == serie].copy()
                
                estado_final = "APROBADO"
                detalles_list = []

                # --- LÓGICA DE SELECCIÓN DE ÚLTIMAS MEDICIONES ---
                df_cable.sort_values(by='timestamp', ascending=False, inplace=True)

                # Filtrar las últimas N mediciones para cada lado
                df_lado_a = df_cable[df_cable[h['conector']] == 'A'].head(fibras_por_conector_a * num_conectores_a)
                df_lado_b = df_cable[df_cable[h['conector']] == 'B'].head(fibras_por_conector_b * num_conectores_b)
                
                # Combinar los dataframes filtrados para el análisis final
                df_final_cable = pd.concat([df_lado_a, df_lado_b])
                
                # Normalizar la columna de estado en el DF final
                df_final_cable[h['estado']] = df_final_cable[h['estado']].str.strip().str.upper()

                # 1. Validar fallas en el conjunto final
                fails = len(df_final_cable[df_final_cable[h['estado']] != 'PASS'])
                if fails > 0:
                    estado_final = "RECHAZADO"
                    detalles_list.append(f"{fails} medicion(es) con FALLA")
                else:
                    detalles_list.append(f"{len(df_final_cable)} medicion(es) con PASS")
                
                # 2. Validar Lado A
                fibras_a_reales = len(df_lado_a)
                fibras_a_esperadas = num_conectores_a * fibras_por_conector_a
                detalles_list.append(f"Fibras Lado A: {fibras_a_reales}/{fibras_a_esperadas}")
                if fibras_a_reales != fibras_a_esperadas:
                    estado_final = "RECHAZADO"
                
                # 3. Validar Lado B
                fibras_b_reales = len(df_lado_b)
                fibras_b_esperadas = num_conectores_b * fibras_por_conector_b
                detalles_list.append(f"Fibras Lado B: {fibras_b_reales}/{fibras_b_esperadas}")
                if fibras_b_reales != fibras_b_esperadas:
                    estado_final = "RECHAZADO"
                
                # 4. Validar total de fibras
                total_mediciones = len(df_final_cable)
                detalles_list.append(f"Total Fibras: {total_mediciones}/{total_fibras_esperadas}")
                if total_mediciones != total_fibras_esperadas:
                    estado_final = "RECHAZADO"
                
                ultima_medicion = df_final_cable['timestamp'].max()
                ultima_medicion_str = ultima_medicion.strftime('%d/%m/%Y %H:%M:%S') if pd.notna(ultima_medicion) else "N/A"
                
                detalle_str = ". ".join(detalles_list)
                
                reporte_data.append({
                    'Número de Serie': serie, 
                    'Estado Final': estado_final, 
                    'Detalle': detalle_str,
                    'Última Medición': ultima_medicion_str
                })
                if estado_final == "RECHAZADO":
                    rechazados.append([serie, estado_final, detalle_str, ultima_medicion_str])
                
                self.update_progress(30 + int((i + 1) / len(all_series) * 65))

            ruta_reporte, error = self._generar_reporte_excel_mpo(
                reporte_data, total_cables, rechazados, file_path, ot_number, "ILRL_MPO"
            )
            self.after(0, self.show_result, ruta_reporte, error, "IL/RL MPO")
            
        except Exception as e:
            self.after(0, self.show_result, None, f"Error crítico en el análisis IL/RL MPO:\n{traceback.format_exc()}", "IL/RL MPO")

    def _run_polaridad_mpo_thread(self, folder_path, ot_number, total_cables):
        try:
            self.update_progress(10)
            
            mediciones = {} # {serie: {'fecha': dt, 'estado': 'APROBADO/RECHAZADO', 'archivo': ...}}
            
            for estado_carpeta, estado_resultado in [('PASS', 'APROBADO'), ('FAIL', 'RECHAZADO')]:
                subcarpeta = os.path.join(folder_path, estado_carpeta)
                if not os.path.isdir(subcarpeta):
                    continue
                
                for filename in os.listdir(subcarpeta):
                    if filename.lower().endswith('.xlsx') and not filename.startswith('~$'):
                        match_serie = re.search(r'(JMO\d{13})', filename, re.IGNORECASE)
                        match_fecha = re.search(r'(\d{4}-\d{2}-\d{2} \d{2}-\d{2}-\d{2})', filename)
                        
                        if match_serie and match_fecha:
                            serie = match_serie.group(1).upper()
                            fecha_str = match_fecha.group(1)
                            fecha_dt = datetime.strptime(fecha_str, '%Y-%m-%d %H-%M-%S')

                            if serie not in mediciones or fecha_dt > mediciones[serie]['fecha']:
                                mediciones[serie] = {
                                    'fecha': fecha_dt,
                                    'estado': estado_resultado,
                                    'archivo': filename
                                }
            
            self.update_progress(60)

            reporte_data = []
            rechazados = []
            for serie, data in mediciones.items():
                reporte_data.append({
                    'Número de Serie': serie,
                    'Estado Final': data['estado'],
                    'Último Archivo': data['archivo'],
                    'Fecha Medición': data['fecha'].strftime('%Y-%m-%d %H:%M:%S')
                })
                if data['estado'] == 'RECHAZADO':
                    rechazados.append([serie, 'RECHAZADO', data['archivo']])

            self.update_progress(90)
            
            ruta_reporte, error = self._generar_reporte_excel_mpo(
                reporte_data, total_cables, rechazados, folder_path, ot_number, "Polaridad_MPO"
            )
            self.after(0, self.show_result, ruta_reporte, error, "Polaridad MPO")

        except Exception as e:
            self.after(0, self.show_result, None, f"Error crítico en el análisis de Polaridad MPO:\n{traceback.format_exc()}", "Polaridad MPO")

    def update_progress(self, value):
        self.progress_bar["value"] = value

    def show_result(self, ruta_generada, errores, tipo_analisis):
        self.progress_bar["value"] = 100
        if ruta_generada:
            self.result_label.config(text=f"Reporte de {tipo_analisis} generado con éxito.")
            if messagebox.askyesno("Éxito", f"Reporte de {tipo_analisis} generado con éxito.\n\n"
                                          f"Guardado en:\n{ruta_generada}\n\n"
                                          "¿Deseas abrir el archivo ahora?", parent=self):
                self.abrir_archivo(ruta_generada)
        else:
            self.result_label.config(text=f"El análisis de {tipo_analisis} falló.")
            messagebox.showerror("Análisis Fallido", f"No se pudo generar el reporte de {tipo_analisis}.\n\n"
                                                   f"Detalles:\n{errores}", parent=self)
        if errores and ruta_generada:
             messagebox.showwarning("Advertencia", f"El reporte se generó, pero se encontraron los siguientes problemas:\n\n{errores}", parent=self)

    def abrir_archivo(self, ruta_archivo):
        try:
            if sys.platform == 'win32':
                os.startfile(ruta_archivo)
            elif sys.platform == 'darwin':
                subprocess.run(['open', ruta_archivo], check=True)
            else:
                subprocess.run(['xdg-open', ruta_archivo], check=True)
        except Exception as e:
            print(f"Advertencia: No se pudo abrir el archivo automáticamente: {e}")

# --- PÁGINA PARA ANÁLISIS DE O.T. (MPO) ---
class AnalisisMPOPage(ttk.Frame):
    def __init__(self, parent, app_instance):
        super().__init__(parent, padding=20)
        self.app = app_instance
        self.create_widgets()

    def create_widgets(self):
        container = ttk.Frame(self, style='TFrame')
        container.pack(expand=True, fill='both')

        config_frame = ttk.LabelFrame(container, text="Parámetros del Análisis MPO", padding=15)
        config_frame.pack(fill='x', pady=(0, 20))
        config_frame.columnconfigure(1, weight=1)

        # ILRL File
        ttk.Label(config_frame, text="Archivo IL/RL MPO:", font="-weight bold").grid(row=0, column=0, sticky='w', pady=5, padx=5)
        self.ilrl_file_var = tk.StringVar()
        ttk.Entry(config_frame, textvariable=self.ilrl_file_var, state="readonly").grid(row=0, column=1, sticky='ew', pady=5, padx=5)
        ttk.Button(config_frame, text="Seleccionar Archivo...", command=self.select_ilrl_file, style='outline.TButton').grid(row=0, column=2, sticky='w', pady=5, padx=5)

        # GEO File
        ttk.Label(config_frame, text="Archivo Geometría MPO:", font="-weight bold").grid(row=1, column=0, sticky='w', pady=5, padx=5)
        self.geo_file_var = tk.StringVar()
        ttk.Entry(config_frame, textvariable=self.geo_file_var, state="readonly").grid(row=1, column=1, sticky='ew', pady=5, padx=5)
        ttk.Button(config_frame, text="Seleccionar Archivo...", command=self.select_geo_file, style='outline.TButton').grid(row=1, column=2, sticky='w', pady=5, padx=5)

        # Polarity Folder
        ttk.Label(config_frame, text="Carpeta Polaridad MPO:", font="-weight bold").grid(row=2, column=0, sticky='w', pady=5, padx=5)
        self.polaridad_folder_var = tk.StringVar()
        ttk.Entry(config_frame, textvariable=self.polaridad_folder_var, state="readonly").grid(row=2, column=1, sticky='ew', pady=5, padx=5)
        ttk.Button(config_frame, text="Seleccionar Carpeta...", command=self.select_polaridad_folder, style='outline.TButton').grid(row=2, column=2, sticky='w', pady=5, padx=5)
        
        ttk.Separator(config_frame, orient='horizontal').grid(row=3, column=0, columnspan=3, sticky='ew', pady=10)

        # OT and Quantity
        ttk.Label(config_frame, text="Número de O.T. (sin JMO-):", font="-weight bold").grid(row=4, column=0, sticky='w', pady=5, padx=5)
        self.ot_var = tk.StringVar()
        ttk.Entry(config_frame, textvariable=self.ot_var).grid(row=4, column=1, sticky='w', pady=5, padx=5)

        ttk.Label(config_frame, text="Total de cables esperados:", font="-weight bold").grid(row=5, column=0, sticky='w', pady=5, padx=5)
        self.total_cables_var = tk.StringVar()
        ttk.Entry(config_frame, textvariable=self.total_cables_var).grid(row=5, column=1, sticky='w', pady=5, padx=5)

        # --- Frame de Acciones ---
        action_frame = ttk.Frame(container)
        action_frame.pack(fill='x', pady=10)
        
        ttk.Button(action_frame, text="Generar Reporte IL/RL", command=lambda: self.run_analysis("ilrl"), style='info.TButton', padding=10).pack(side='left', padx=10, expand=True)
        ttk.Button(action_frame, text="Generar Reporte Geometría", command=lambda: self.run_analysis("geo"), style='info.TButton', padding=10).pack(side='left', padx=10, expand=True)
        ttk.Button(action_frame, text="Generar Reporte Polaridad", command=lambda: self.run_analysis("polaridad"), style='info.TButton', padding=10).pack(side='left', padx=10, expand=True)

        # --- Frame de Progreso y Resultados ---
        result_frame = ttk.LabelFrame(container, text="Estado del Análisis", padding=15)
        result_frame.pack(fill='both', expand=True, pady=(20, 0))

        self.progress_bar = ttk.Progressbar(result_frame, mode='determinate')
        self.progress_bar.pack(fill='x', pady=10)
        
        self.result_label = ttk.Label(result_frame, text="Listo para iniciar el análisis MPO.", wraplength=700)
        self.result_label.pack(fill='x', pady=10)

    def _select_file(self, var, title):
        file_selected = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_selected:
            var.set(file_selected)
            # --- CORRECCIÓN: Acepta JMO o JRMO ---
            match = re.search(r'(J(?:R)?MO-\d{9})', os.path.basename(file_selected), re.IGNORECASE)
            if match and not self.ot_var.get():
                self.ot_var.set(match.group(1).upper())

    def select_ilrl_file(self):
        self._select_file(self.ilrl_file_var, "Seleccionar archivo de IL/RL MPO")

    def select_geo_file(self):
        self._select_file(self.geo_file_var, "Seleccionar archivo de Geometría MPO")
        
    def select_polaridad_folder(self):
        folder_selected = filedialog.askdirectory(title="Seleccionar carpeta de Polaridad (ej. JMO-250800005)")
        if folder_selected:
            self.polaridad_folder_var.set(folder_selected)
            match = re.search(r'JMO-(\d{9})', os.path.basename(folder_selected), re.IGNORECASE)
            if match and not self.ot_var.get():
                # Corrección para consistencia: sugerir sin prefijo
                self.ot_var.set(match.group(1))
    
    def run_analysis(self, analysis_type):
        ot_number_raw = self.ot_var.get().strip()
        total_str = self.total_cables_var.get()
        
        # Validaciones comunes
        if not ot_number_raw:
            messagebox.showerror("Error", "Por favor, ingresa el número de O.T.", parent=self)
            return
        if not total_str.isdigit() or int(total_str) <= 0:
            messagebox.showerror("Error", "Por favor, ingresa un número válido para el total de cables.", parent=self)
            return
        
        ot_number = f"JMO-{ot_number_raw}" if not ot_number_raw.startswith("JMO-") else ot_number_raw
        total = int(total_str)
        
        self.progress_bar["value"] = 0
        self.result_label.config(text=f"Iniciando análisis de {analysis_type.upper()} para MPO...")
        
        # Validaciones y arranque de hilos específicos
        if analysis_type == "ilrl":
            file_path = self.ilrl_file_var.get()
            if not file_path:
                messagebox.showerror("Error", "Por favor, selecciona el archivo de IL/RL para MPO.", parent=self)
                return
            thread = threading.Thread(target=self._run_ilrl_mpo_thread, args=(file_path, ot_number, total), daemon=True)
            thread.start()
        elif analysis_type == "geo":
            file_path = self.geo_file_var.get()
            if not file_path:
                messagebox.showerror("Error", "Por favor, selecciona el archivo de Geometría para MPO.", parent=self)
                return
            thread = threading.Thread(target=self._run_geo_mpo_thread, args=(file_path, ot_number, total), daemon=True)
            thread.start()
        elif analysis_type == "polaridad":
            folder_path = self.polaridad_folder_var.get()
            if not folder_path:
                messagebox.showerror("Error", "Por favor, selecciona la carpeta de Polaridad para MPO.", parent=self)
                return
            thread = threading.Thread(target=self._run_polaridad_mpo_thread, args=(folder_path, ot_number, total), daemon=True)
            thread.start()

    def _cargar_ot_configuration(self, ot_number):
        try:
            conn = sqlite3.connect(self.app.config['db_path'])
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM ot_configurations WHERE ot_number = ?", (ot_number,))
            row = cursor.fetchone()
            conn.close()
            return dict(row) if row else None
        except Exception as e:
            messagebox.showerror("Error de Base de Datos", f"No se pudo cargar la configuración de la OT: {e}", parent=self)
            return None

# En la clase AnalisisMPOPage, reemplaza este método completo:

    def _run_ilrl_mpo_thread(self, file_path, ot_number, total_cables):
        try:
            self.update_progress(5)
            ot_config = self._cargar_ot_configuration(ot_number)
            if not ot_config:
                self.after(0, self.show_result, None, f"No se encontró configuración para la OT {ot_number}. Por favor, configúrela primero.", "IL/RL MPO")
                return

            self.update_progress(15)
            df = pd.read_excel(file_path, sheet_name="Results")
            h = {k: ot_config.get(v, d) for k, v, d in [('serie', 'ilrl_serie_header', 'Serial number'), ('estado', 'ilrl_estado_header', 'Alarm Status'), ('conector', 'ilrl_conector_header', 'connector label'), ('fecha', 'ilrl_fecha_header', 'Date'), ('hora', 'ilrl_hora_header', 'Time')]}
            
            if any(header not in df.columns for header in h.values()):
                missing = [k for k, v in h.items() if v not in df.columns]
                self.after(0, self.show_result, None, f"El archivo no contiene los encabezados esperados: {', '.join(missing)}.", "IL/RL MPO")
                return
            
            self.update_progress(30)
            df[h['serie']] = df[h['serie']].astype(str) # Esta es la columna con el N/S completo (ej. ...0028-F)

            # --- INICIO DE LA NUEVA LÓGICA DE PRIORIZACIÓN DE RETRABAJOS (-F) ---
            
            # 1. *** LÍNEA CORREGIDA ***
            #    Añadimos dayfirst=True para que entienda '29/09/2025' como Día/Mes/Año.
            date_series = pd.to_datetime(df[h['fecha']], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y')
            
            time_series = df[h['hora']].astype(str).str.replace('a. m.', 'AM', regex=False).str.replace('p. m.', 'PM', regex=False).str.strip()
            full_datetime_str = date_series + ' ' + time_series
            df['timestamp'] = pd.to_datetime(full_datetime_str, format='%d/%m/%Y %I:%M:%S %p', errors='coerce')
            
            # Esta línea ahora solo borrará las filas *realmente* corruptas (como las del cable 0001)
            df.dropna(subset=['timestamp'], inplace=True)

            # 2. Crear una columna 'base_serie' que contenga solo los 13 dígitos
            df['base_serie'] = df[h['serie']].str.extract(r'(\d{13})')
            df.dropna(subset=['base_serie'], inplace=True) # Descartar filas que no tengan un N/S de 13 dígitos

            # 3. Obtener la lista de cables base únicos
            all_base_series = df['base_serie'].dropna().unique()

            reporte_data = []
            rechazados = []
            
            num_conectores_a = ot_config.get('num_conectores_a', 1)
            fibras_por_conector_a = ot_config.get('fibers_per_connector_a', 12)
            num_conectores_b = ot_config.get('num_conectores_b', 1)
            fibras_por_conector_b = ot_config.get('fibers_per_connector_b', 12)
            total_fibras_esperadas = (num_conectores_a * fibras_por_conector_a) + (num_conectores_b * fibras_por_conector_b)

            # 4. Iterar sobre los cables base, no sobre los N/S completos
            for i, base_serie in enumerate(all_base_series):
                # Obtener todas las mediciones para este cable base (ej. original y -F)
                df_cable_group = df[df['base_serie'] == base_serie]

                # 5. Encontrar qué N/S completo (original o -F) tiene la fecha más reciente
                latest_timestamp = pd.NaT
                best_sub_group_df = None
                best_full_serial = None

                # Iterar sobre las variaciones (ej. "...0028" y "...0028-F")
                for full_serial in df_cable_group[h['serie']].unique():
                    sub_group_df = df_cable_group[df_cable_group[h['serie']] == full_serial]
                    current_max_timestamp = sub_group_df['timestamp'].max() # Fecha más reciente de esta variación

                    if best_sub_group_df is None or current_max_timestamp > latest_timestamp:
                        latest_timestamp = current_max_timestamp
                        best_sub_group_df = sub_group_df
                        best_full_serial = full_serial # Guardar el N/S "ganador"

                # 6. 'best_sub_group_df' ahora contiene solo las filas del N/S más reciente
                if best_sub_group_df is None:
                    continue 
                
                df_cable = best_sub_group_df.copy()
                serie_para_reporte = best_full_serial # Usaremos este N/S para el reporte

                # --- FIN DE LA NUEVA LÓGICA ---

                # 7. El resto de la lógica de análisis se aplica solo al grupo "ganador"
                estado_final = "APROBADO"
                detalles_list = []

                df_cable.sort_values(by='timestamp', ascending=False, inplace=True)

                df_lado_a = df_cable[df_cable[h['conector']] == 'A'].head(fibras_por_conector_a * num_conectores_a)
                df_lado_b = df_cable[df_cable[h['conector']] == 'B'].head(fibras_por_conector_b * num_conectores_b)
                
                df_final_cable = pd.concat([df_lado_a, df_lado_b])
                
                if df_final_cable.empty: continue

                df_final_cable.loc[:, h['estado']] = df_final_cable[h['estado']].str.strip().str.upper()

                fails = len(df_final_cable[df_final_cable[h['estado']] != 'PASS'])
                if fails > 0:
                    estado_final = "RECHAZADO"
                    detalles_list.append(f"{fails} medicion(es) con FALLA")
                else:
                    detalles_list.append(f"{len(df_final_cable)} medicion(es) con PASS")
                
                fibras_a_reales = len(df_lado_a)
                fibras_a_esperadas = num_conectores_a * fibras_por_conector_a
                detalles_list.append(f"Fibras Lado A: {fibras_a_reales}/{fibras_a_esperadas}")
                if fibras_a_reales != fibras_a_esperadas: estado_final = "RECHAZADO"
                
                fibras_b_reales = len(df_lado_b)
                fibras_b_esperadas = num_conectores_b * fibras_por_conector_b
                detalles_list.append(f"Fibras Lado B: {fibras_b_reales}/{fibras_b_esperadas}")
                if fibras_b_reales != fibras_b_esperadas: estado_final = "RECHAZADO"
                
                total_mediciones = len(df_final_cable)
                detalles_list.append(f"Total Fibras: {total_mediciones}/{total_fibras_esperadas}")
                if total_mediciones != total_fibras_esperadas: estado_final = "RECHAZADO"
                
                ultima_medicion_str = ""
                fechas_por_punta = []
                if not df_lado_a.empty:
                    max_fecha_a = df_lado_a['timestamp'].max()
                    fechas_por_punta.append(f"Lado A: {max_fecha_a.strftime('%d/%m/%Y %H:%M:%S') if pd.notna(max_fecha_a) else 'N/A'}")
                if not df_lado_b.empty:
                    max_fecha_b = df_lado_b['timestamp'].max()
                    fechas_por_punta.append(f"Lado B: {max_fecha_b.strftime('%d/%m/%Y %H:%M:%S') if pd.notna(max_fecha_b) else 'N/A'}")
                ultima_medicion_str = ". ".join(fechas_por_punta)
                
                detalle_str = ". ".join(detalles_list)
                
                # 8. Guardar en el reporte usando el N/S "ganador"
                reporte_data.append({
                    'Número de Serie': serie_para_reporte, 
                    'Estado Final': estado_final, 
                    'Detalle': detalle_str,
                    'Última Medición': ultima_medicion_str
                })
                if estado_final == "RECHAZADO":
                    rechazados.append([serie_para_reporte, estado_final, detalle_str, ultima_medicion_str])
                
                self.update_progress(30 + int((i + 1) / len(all_base_series) * 65))

            ruta_reporte, error = self._generar_reporte_excel_mpo(
                reporte_data, total_cables, rechazados, file_path, ot_number, "ILRL_MPO"
            )
            self.after(0, self.show_result, ruta_reporte, error, "IL/RL MPO")
            
        except Exception as e:
            self.after(0, self.show_result, None, f"Error crítico en el análisis IL/RL MPO:\n{traceback.format_exc()}", "IL/RL MPO")
            
    def _run_geo_mpo_thread(self, file_path, ot_number, total_cables):
        try:
            self.update_progress(5)
            ot_config = self._cargar_ot_configuration(ot_number)
            if not ot_config:
                self.after(0, self.show_result, None, f"No se encontró configuración para la OT {ot_number}.", "Geometría MPO")
                return

            self.update_progress(15)
            df = pd.read_excel(file_path, sheet_name="MT12", header=None)
            header_row_index = next((i for i, row in df.iterrows() if str(row.iloc[0]).strip().lower() == 'name'), -1)
            
            if header_row_index == -1:
                self.after(0, self.show_result, None, "No se encontró encabezado 'Name' en la hoja 'MT12'.", "Geometría MPO")
                return
            
            df.columns = df.iloc[header_row_index].str.strip().str.lower()
            df = df.iloc[header_row_index + 1:].rename(columns={'pass/fail': 'result'})
            self.update_progress(30)

            # --- FUNCIÓN MEJORADA PARA RECONOCER RETRABAJOS Y JRMO ---
            def parse_geo_name_mpo(name_str):
                # --- CORRECCIÓN: Regex ajustada para J(R)MO ---
                match = re.match(r'(J(?:R)?MO\d{13})-(R?\d+)(-[FK])?', str(name_str).upper())
                if match:
                    # Retorna: base del serial (ej. JRMO...001), punta (ej. "1" o "R1"), y si es fanout
                    return match.group(1), match.group(2), match.group(3)
                return None, None, None

            df[['serie_base', 'lado', 'retrabajo_info']] = df['name'].apply(lambda x: pd.Series(parse_geo_name_mpo(x)))
            
            # Limpiar filas que no se pudieron parsear
            df.dropna(subset=['serie_base', 'lado'], inplace=True)

            if 'date & time' in df.columns:
                df['timestamp'] = pd.to_datetime(df['date & time'], errors='coerce')
            else:
                 df['timestamp'] = pd.to_datetime(df['date'].astype(str) + ' ' + df['time'].astype(str), errors='coerce')

            all_series = df['serie_base'].dropna().unique()
            reporte_data = []
            rechazados = []
            
            num_lados_esperados = ot_config.get('num_conectores_a', 1) + ot_config.get('num_conectores_b', 1)
            lados_esperados = [str(i + 1) for i in range(num_lados_esperados)]

            for i, serie in enumerate(all_series):
                df_cable = df[df['serie_base'] == serie].copy().sort_values(by='timestamp', ascending=False)
                
                # --- INICIO DE LA NUEVA LÓGICA DE PRIORIZACIÓN DE RETRABAJOS ---
                originales = {}
                retrabajos = {}
                for _, medicion in df_cable.iterrows():
                    punta_original = str(medicion['lado'])
                    punta_normalizada = punta_original.replace('R', '')

                    if 'R' in punta_original:
                        if punta_normalizada not in retrabajos:
                            retrabajos[punta_normalizada] = {'Punta_Original': punta_original, 'Resultado': medicion['result'], 'Fecha': medicion['timestamp'], 'Nombre_Completo': medicion['name']}
                    else:
                        if punta_normalizada not in originales:
                            originales[punta_normalizada] = {'Punta_Original': punta_original, 'Resultado': medicion['result'], 'Fecha': medicion['timestamp'], 'Nombre_Completo': medicion['name']}

                mediciones_definitivas = {}
                for punta, medicion in retrabajos.items():
                    mediciones_definitivas[punta] = medicion
                
                for punta, medicion in originales.items():
                    if punta not in mediciones_definitivas:
                        mediciones_definitivas[punta] = medicion
                # --- FIN DE LA NUEVA LÓGICA DE PRIORIZACIÓN ---

                estado_final = "APROBADO"
                detalles_list = []
                fechas_list = []

                for lado_esperado in lados_esperados:
                    if lado_esperado in mediciones_definitivas:
                        medicion = mediciones_definitivas[lado_esperado]
                        estado_lado = str(medicion['Resultado']).upper()
                        nombre_original = medicion['Nombre_Completo']
                        fecha_lado = medicion['Fecha']

                        if estado_lado != 'PASS':
                            estado_final = "RECHAZADO"
                        
                        detalles_list.append(f"Lado {medicion['Punta_Original']}: {estado_lado}")
                        fechas_list.append(f"Lado {medicion['Punta_Original']}: {fecha_lado.strftime('%d/%m/%Y %H:%M:%S') if pd.notna(fecha_lado) else 'N/A'}")
                    else:
                        estado_final = "RECHAZADO"
                        detalles_list.append(f"Lado {lado_esperado}: FALTANTE")

                detalle_str = ". ".join(detalles_list)
                fechas_str = ". ".join(fechas_list)
                
                reporte_data.append({
                    'Número de Serie': serie, 
                    'Estado Final': estado_final, 
                    'Detalle': detalle_str,
                    'Última Medición': fechas_str
                })
                if estado_final == "RECHAZADO":
                    rechazados.append([serie, estado_final, detalle_str, fechas_str])

                self.update_progress(30 + int((i + 1) / len(all_series) * 65))

            ruta_reporte, error = self._generar_reporte_excel_mpo(
                reporte_data, total_cables, rechazados, file_path, ot_number, "GEO_MPO"
            )
            self.after(0, self.show_result, ruta_reporte, error, "Geometría MPO")

        except Exception as e:
            self.after(0, self.show_result, None, f"Error crítico en el análisis de Geometría MPO:\n{traceback.format_exc()}", "Geometría MPO")

    def _run_polaridad_mpo_thread(self, folder_path, ot_number, total_cables):
        try:
            self.update_progress(10)
            
            mediciones = {} 
            
            for estado_carpeta, estado_resultado in [('PASS', 'APROBADO'), ('FAIL', 'RECHAZADO')]:
                subcarpeta = os.path.join(folder_path, estado_carpeta)
                if not os.path.isdir(subcarpeta):
                    continue
                
                for filename in os.listdir(subcarpeta):
                    if filename.lower().endswith('.xlsx') and not filename.startswith('~$'):
                        # --- CORRECCIÓN: Regex ajustada para J(R)MO ---
                        match_serie = re.search(r'(J(?:R)?MO\d{13})', filename, re.IGNORECASE)
                        match_fecha = re.search(r'(\d{4}-\d{2}-\d{2} \d{2}-\d{2}-\d{2})', filename)
                        
                        if match_serie and match_fecha:
                            serie = match_serie.group(1).upper()
                            fecha_str = match_fecha.group(1)
                            fecha_dt = datetime.strptime(fecha_str, '%Y-%m-%d %H-%M-%S')

                            if serie not in mediciones or fecha_dt > mediciones[serie]['fecha']:
                                mediciones[serie] = {
                                    'fecha': fecha_dt,
                                    'estado': estado_resultado,
                                    'archivo': filename
                                }
            
            self.update_progress(60)

            reporte_data = []
            rechazados = []
            for serie, data in mediciones.items():
                reporte_data.append({
                    'Número de Serie': serie,
                    'Estado Final': data['estado'],
                    'Último Archivo': data['archivo'],
                    'Fecha Medición': data['fecha'].strftime('%Y-%m-%d %H:%M:%S')
                })
                if data['estado'] == 'RECHAZADO':
                    rechazados.append([serie, 'RECHAZADO', data['archivo']])

            self.update_progress(90)
            
            ruta_reporte, error = self._generar_reporte_excel_mpo(
                reporte_data, total_cables, rechazados, folder_path, ot_number, "Polaridad_MPO"
            )
            self.after(0, self.show_result, ruta_reporte, error, "Polaridad MPO")

        except Exception as e:
            self.after(0, self.show_result, None, f"Error crítico en el análisis de Polaridad MPO:\n{traceback.format_exc()}", "Polaridad MPO")

    def _generar_reporte_excel_mpo(self, reporte_data, total_esperado, rechazados, ruta_base, ot_number, tipo_reporte):
        # Para archivos, el reporte se guarda junto al archivo. Para carpetas, dentro de la carpeta.
        if os.path.isfile(ruta_base):
            carpeta_destino = os.path.dirname(ruta_base)
        else:
            carpeta_destino = ruta_base
            
        carpeta_analisis = os.path.join(carpeta_destino, "ANALISIS DE O.T")
        os.makedirs(carpeta_analisis, exist_ok=True)
        
        ot_sufijo = ot_number.replace("JMO-", "")
        nombre_archivo = f"Reporte_Analisis_{tipo_reporte}_{ot_sufijo}.xlsx"
        ruta_destino = os.path.join(carpeta_analisis, nombre_archivo)
        
        try:
            if os.path.exists(ruta_destino):
                os.remove(ruta_destino)

            wb = Workbook()
            
            # --- Estilos reutilizables ---
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # --- Hoja 1: Resultados OT ---
            ws1 = wb.active
            ws1.title = "Resultados OT"
            df_reporte = pd.DataFrame(reporte_data)
            
            headers = []
            if not df_reporte.empty:
                headers = list(df_reporte.columns)
                ws1.append(headers)
                if 'Número de Serie' in df_reporte.columns:
                     df_reporte.sort_values(by='Número de Serie', inplace=True)
                for r in df_reporte.itertuples(index=False):
                    ws1.append(list(r))
                
                # Aplicar tabla con filtros
                tabla_resultados = Table(displayName="ResultadosOT", ref=f"A1:{get_column_letter(ws1.max_column)}{ws1.max_row}")
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tabla_resultados.tableStyleInfo = style
                ws1.add_table(tabla_resultados)

                # Formato condicional para APROBADO/RECHAZADO
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                ws1.conditional_formatting.add(f"B2:B{ws1.max_row}", CellIsRule(operator='equal', formula=['APROBADO'], fill=green_fill))
                ws1.conditional_formatting.add(f"B2:B{ws1.max_row}", CellIsRule(operator='equal', formula=['RECHAZADO'], fill=red_fill))

            # --- Hoja 2: Resumen y Faltantes ---
            ws2 = wb.create_sheet("Resumen")
            encontrados = len(reporte_data)
            faltantes = total_esperado - encontrados
            ws2.append(['Métrica', 'Valor'])
            ws2.append(['Total Esperado', total_esperado])
            ws2.append(['Total Encontrados', encontrados])
            ws2.append(['Total Faltantes', faltantes])
            for cell in ws2["A"]: cell.font = Font(bold=True) # Negrita para la primera columna
            
            # --- Hoja 3: Rechazados ---
            if rechazados:
                ws3 = wb.create_sheet("Rechazados")
                if headers:
                    ws3.append(headers)
                rechazados.sort(key=lambda x: x[0])
                for fila in rechazados:
                    ws3.append(fila)
                
                tabla_rechazados = Table(displayName="Rechazados", ref=f"A1:{get_column_letter(ws3.max_column)}{ws3.max_row}")
                style_rechazados = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False,
                                                  showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tabla_rechazados.tableStyleInfo = style_rechazados
                ws3.add_table(tabla_rechazados)

            # Auto-ajustar columnas y aplicar estilos a encabezados en todas las hojas
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                for col in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) if max_length < 50 else 50
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Estilo de encabezado
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment

            wb.save(ruta_destino)
            return ruta_destino, None
        
        except PermissionError:
            error_msg = (f"No se pudo guardar el reporte porque el archivo está siendo usado por otro proceso.\n\n"
                         f"Por favor, cierra el archivo:\n{nombre_archivo}\n\ny vuelve a intentarlo.")
            return None, error_msg
        except Exception as e:
            return None, f"No se pudo guardar el reporte:\n{traceback.format_exc()}"
    
    def update_progress(self, value):
        self.progress_bar["value"] = value

    def show_result(self, ruta_generada, errores, tipo_analisis):
        self.progress_bar["value"] = 100
        if ruta_generada:
            self.result_label.config(text=f"Reporte de {tipo_analisis} generado con éxito.")
            if messagebox.askyesno("Éxito", f"Reporte de {tipo_analisis} generado con éxito.\n\n"
                                          f"Guardado en:\n{ruta_generada}\n\n"
                                          "¿Deseas abrir el archivo ahora?", parent=self):
                self.abrir_archivo(ruta_generada)
        else:
            self.result_label.config(text=f"El análisis de {tipo_analisis} falló.")
            messagebox.showerror("Análisis Fallido", f"No se pudo generar el reporte de {tipo_analisis}.\n\n"
                                                   f"Detalles:\n{errores}", parent=self)
        if errores and ruta_generada:
             messagebox.showwarning("Advertencia", f"El reporte se generó, pero se encontraron los siguientes problemas:\n\n{errores}", parent=self)

    def abrir_archivo(self, ruta_archivo):
        try:
            if sys.platform == 'win32':
                os.startfile(ruta_archivo)
            elif sys.platform == 'darwin':
                subprocess.run(['open', ruta_archivo], check=True)
            else:
                subprocess.run(['xdg-open', ruta_archivo], check=True)
        except Exception as e:
            print(f"Advertencia: No se pudo abrir el archivo automáticamente: {e}")

class RecordsPage(ttk.Frame):
    def __init__(self, parent, app_instance):
        super().__init__(parent, padding=10)
        self.app = app_instance
        self.create_widgets()

    def create_widgets(self):
        filter_frame = ttk.Frame(self)
        filter_frame.pack(fill='x', pady=5)
        ttk.Label(filter_frame, text="Filtrar por OT o Serie:").pack(side='left', padx=5)
        self.filter_entry = ttk.Entry(filter_frame, width=30)
        self.filter_entry.pack(side='left', padx=5)
        self.filter_entry.bind("<KeyRelease>", lambda e: self.load_records())

        cols = ("ID", "Fecha", "No. Serie", "OT", "Estado Final", "ILRL", "GEO", "Polaridad")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", style='primary.Treeview')
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor='w')
        
        self.tree.column("ID", width=40, stretch=False)
        self.tree.pack(fill='both', expand=True, pady=10)

        self.tree.tag_configure('APROBADO', foreground='green')
        self.tree.tag_configure('RECHAZADO', foreground='red')
        self.tree.tag_configure('NO ENCONTRADO', foreground='orange')
        
    def load_records(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        filtro = self.filter_entry.get().strip()
        try:
            conn = sqlite3.connect(self.app.config['db_path'])
            cursor = conn.cursor()
            query = "SELECT id, entry_date, serial_number, ot_number, overall_status, ilrl_status, geo_status, polaridad_status FROM cable_verifications"
            params = []
            if filtro:
                query += " WHERE serial_number LIKE ? OR ot_number LIKE ?"
                params.extend([f'%{filtro}%', f'%{filtro}%'])
            query += " ORDER BY id DESC"
            
            cursor.execute(query, params)
            for row in cursor.fetchall():
                row_list = list(row)
                while len(row_list) < 8:
                    row_list.append('N/A')
                self.tree.insert("", "end", values=tuple(row_list), tags=(row[4],))
            conn.close()
        except Exception as e:
            messagebox.showerror("Error de Base de Datos", f"No se pudieron cargar los registros: {e}")

class SettingsWindow(tk.Toplevel):
    def __init__(self, app_instance):
        super().__init__(app_instance)
        self.app = app_instance
        self.title("Configurar Rutas")
        self.geometry("800x750") # Aumentamos la altura a 750 para que quepan las nuevas BD
        self.transient(self.app)
        self.grab_set()

        self.bind_class("TEntry", "<FocusIn>", open_keyboard)

        # Usamos un Canvas para permitir Scroll en caso de pantallas pequeñas
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True)
        canvas = tk.Canvas(main_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        frame = ttk.Frame(canvas, padding=20)
        
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        frame.columnconfigure(1, weight=1)

        # --- SECCIÓN LC/SC ---
        ttk.Label(frame, text="Rutas SC/LC", font=("Helvetica", 10, "bold")).grid(row=0, column=0, columnspan=3, sticky='w', pady=(0,5))
        ttk.Label(frame, text="Ruta ILRL (Principal):").grid(row=1, column=0, sticky='w', pady=2)
        self.ilrl_path = tk.StringVar(value=self.app.config.get('ruta_base_ilrl', ''))
        ttk.Entry(frame, textvariable=self.ilrl_path, width=60).grid(row=1, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_folder(self.ilrl_path)).grid(row=1, column=2, padx=5)

        ttk.Label(frame, text="Ruta ILRL (Secundaria):").grid(row=2, column=0, sticky='w', pady=2)
        self.ilrl_path_2 = tk.StringVar(value=self.app.config.get('ruta_base_ilrl_2', ''))
        ttk.Entry(frame, textvariable=self.ilrl_path_2, width=60).grid(row=2, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_folder(self.ilrl_path_2)).grid(row=2, column=2, padx=5)

        ttk.Label(frame, text="Ruta Geometría (Principal):").grid(row=3, column=0, sticky='w', pady=2)
        self.geo_path = tk.StringVar(value=self.app.config.get('ruta_base_geo', ''))
        ttk.Entry(frame, textvariable=self.geo_path, width=60).grid(row=3, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_folder(self.geo_path)).grid(row=3, column=2, padx=5)

        ttk.Label(frame, text="Ruta Geometría (Secundaria):").grid(row=4, column=0, sticky='w', pady=2)
        self.geo_path_2 = tk.StringVar(value=self.app.config.get('ruta_base_geo_2', ''))
        ttk.Entry(frame, textvariable=self.geo_path_2, width=60).grid(row=4, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_folder(self.geo_path_2)).grid(row=4, column=2, padx=5)
        
        ttk.Separator(frame).grid(row=5, column=0, columnspan=3, pady=10, sticky='ew')

        # --- SECCIÓN MPO ---
        ttk.Label(frame, text="Rutas MPO", font=("Helvetica", 10, "bold")).grid(row=6, column=0, columnspan=3, sticky='w', pady=(0,5))
        ttk.Label(frame, text="Ruta ILRL (MPO):").grid(row=7, column=0, sticky='w', pady=2)
        self.ilrl_mpo_path = tk.StringVar(value=self.app.config.get('ruta_base_ilrl_mpo', ''))
        ttk.Entry(frame, textvariable=self.ilrl_mpo_path, width=60).grid(row=7, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_folder(self.ilrl_mpo_path)).grid(row=7, column=2, padx=5)

        ttk.Label(frame, text="Ruta Geometría (MPO):").grid(row=8, column=0, sticky='w', pady=2)
        self.geo_mpo_path = tk.StringVar(value=self.app.config.get('ruta_base_geo_mpo', ''))
        ttk.Entry(frame, textvariable=self.geo_mpo_path, width=60).grid(row=8, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_folder(self.geo_mpo_path)).grid(row=8, column=2, padx=5)

        ttk.Label(frame, text="Ruta Polaridad (MPO):").grid(row=9, column=0, sticky='w', pady=2)
        self.polaridad_mpo_path = tk.StringVar(value=self.app.config.get('ruta_base_polaridad_mpo', ''))
        ttk.Entry(frame, textvariable=self.polaridad_mpo_path, width=60).grid(row=9, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_folder(self.polaridad_mpo_path)).grid(row=9, column=2, padx=5)
        
        ttk.Separator(frame).grid(row=10, column=0, columnspan=3, pady=10, sticky='ew')

        # --- SECCIÓN FANOUT ---
        ttk.Label(frame, text="Rutas FANOUT", font=("Helvetica", 10, "bold")).grid(row=11, column=0, columnspan=3, sticky='w', pady=(0,5))
        ttk.Label(frame, text="Ruta Geo LC/FC (Fanout):").grid(row=12, column=0, sticky='w', pady=2)
        self.geo_fanout_path = tk.StringVar(value=self.app.config.get('ruta_base_geo_fanout_lc', ''))
        ttk.Entry(frame, textvariable=self.geo_fanout_path, width=60).grid(row=12, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_folder(self.geo_fanout_path)).grid(row=12, column=2, padx=5)

        ttk.Separator(frame).grid(row=13, column=0, columnspan=3, pady=10, sticky='ew')
        
        # --- SECCIÓN UNIBOOT ---
        ttk.Label(frame, text="Rutas UNIBOOT", font=("Helvetica", 10, "bold")).grid(row=14, column=0, columnspan=3, sticky='w', pady=(0,5))
        ttk.Label(frame, text="Ruta ILRL (Uniboot):").grid(row=15, column=0, sticky='w', pady=2)
        self.ilrl_uniboot_path = tk.StringVar(value=self.app.config.get('ruta_base_ilrl_uniboot', ''))
        ttk.Entry(frame, textvariable=self.ilrl_uniboot_path, width=60).grid(row=15, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_folder(self.ilrl_uniboot_path)).grid(row=15, column=2, padx=5)

        ttk.Label(frame, text="Ruta Geometría (Uniboot):").grid(row=16, column=0, sticky='w', pady=2)
        self.geo_uniboot_path = tk.StringVar(value=self.app.config.get('ruta_base_geo_uniboot', ''))
        ttk.Entry(frame, textvariable=self.geo_uniboot_path, width=60).grid(row=16, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_folder(self.geo_uniboot_path)).grid(row=16, column=2, padx=5)

        ttk.Separator(frame).grid(row=17, column=0, columnspan=3, pady=10, sticky='ew')

        # --- SECCIÓN DE BASES DE DATOS (SYNOLOGY) ---
        ttk.Label(frame, text="Rutas Bases de Datos (LC/SC)", font=("Helvetica", 10, "bold")).grid(row=18, column=0, columnspan=3, sticky='w', pady=(0,5))
        
        ttk.Label(frame, text="Línea JWS1-1:").grid(row=19, column=0, sticky='w', pady=2)
        self.db_path_1 = tk.StringVar(value=self.app.config.get('db_path_jws1_1', ''))
        ttk.Entry(frame, textvariable=self.db_path_1, width=60).grid(row=19, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_db_file(self.db_path_1)).grid(row=19, column=2, padx=5)

        ttk.Label(frame, text="Línea JWS1-2:").grid(row=20, column=0, sticky='w', pady=2)
        self.db_path_2 = tk.StringVar(value=self.app.config.get('db_path_jws1_2', ''))
        ttk.Entry(frame, textvariable=self.db_path_2, width=60).grid(row=20, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_db_file(self.db_path_2)).grid(row=20, column=2, padx=5)

        ttk.Label(frame, text="Línea JWS1-3:").grid(row=21, column=0, sticky='w', pady=2)
        self.db_path_3 = tk.StringVar(value=self.app.config.get('db_path_jws1_3', ''))
        ttk.Entry(frame, textvariable=self.db_path_3, width=60).grid(row=21, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_db_file(self.db_path_3)).grid(row=21, column=2, padx=5)

        ttk.Separator(frame).grid(row=22, column=0, columnspan=3, pady=10, sticky='ew')

        ttk.Label(frame, text="BD Local/General (MPO/Fanout):").grid(row=23, column=0, sticky='w', pady=2)
        self.db_path = tk.StringVar(value=self.app.config.get('db_path', ''))
        ttk.Entry(frame, textvariable=self.db_path, width=60).grid(row=23, column=1, sticky='ew', padx=5)
        ttk.Button(frame, text="...", command=lambda: self.browse_db_file(self.db_path)).grid(row=23, column=2, padx=5)
        
        # --- Configuración de Identidad de la Línea ---
        ttk.Separator(frame).grid(row=24, column=0, columnspan=3, pady=10, sticky='ew')
        ttk.Label(frame, text="Línea de esta PC (Estación):", font=("Helvetica", 10, "bold")).grid(row=25, column=0, sticky='w', pady=2)
        self.linea_actual_var = tk.StringVar(value=self.app.config.get('linea_actual', 'JWS1-1'))
        cb_linea = ttk.Combobox(frame, textvariable=self.linea_actual_var, values=["JWS1-1", "JWS1-2", "JWS1-3"], state="readonly", width=15)
        cb_linea.grid(row=25, column=1, sticky='w', padx=5)
        
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=26, column=0, columnspan=3, pady=20)
        ttk.Button(btn_frame, text="Guardar", command=self.save_and_close, style='success.TButton').pack(side='left', padx=10)
        ttk.Button(btn_frame, text="Cancelar", command=self.destroy, style='danger.TButton').pack(side='left', padx=10)
        
    def browse_folder(self, path_var):
        initial = path_var.get() if os.path.isdir(path_var.get()) else "/"
        directory = filedialog.askdirectory(initialdir=initial)
        if directory:
            path_var.set(directory)

    def browse_db_file(self, path_var):
        initial_dir = os.path.dirname(path_var.get()) if os.path.exists(path_var.get()) else "/"
        filepath = filedialog.asksaveasfilename(
            defaultextension=".db",
            filetypes=[("Database files", "*.db"), ("All files", "*.*")],
            initialdir=initial_dir,
            initialfile=os.path.basename(path_var.get()) if path_var.get() else "verifications.db"
        )
        if filepath:
            path_var.set(filepath)

    def save_and_close(self):
        new_config = {
            "linea_actual": self.linea_actual_var.get(),
            "ruta_base_ilrl": self.ilrl_path.get(),
            "ruta_base_ilrl_2": self.ilrl_path_2.get(),
            "ruta_base_geo": self.geo_path.get(),
            "ruta_base_geo_2": self.geo_path_2.get(),
            "ruta_base_ilrl_mpo": self.ilrl_mpo_path.get(),
            "ruta_base_geo_mpo": self.geo_mpo_path.get(),
            "ruta_base_polaridad_mpo": self.polaridad_mpo_path.get(),
            "db_path": self.db_path.get(),
            "db_path_jws1_1": self.db_path_1.get(),
            "db_path_jws1_2": self.db_path_2.get(),
            "db_path_jws1_3": self.db_path_3.get(),
            "ruta_base_geo_fanout_lc": self.geo_fanout_path.get(),
            "ruta_base_ilrl_uniboot": self.ilrl_uniboot_path.get(),
            "ruta_base_geo_uniboot": self.geo_uniboot_path.get()
        }
        self.app.save_config(new_config)
        messagebox.showinfo("Guardado", "La configuración se ha guardado correctamente.")
        self.destroy()

class MPOConfigWindow(tk.Toplevel):
    def __init__(self, app_instance, parent_page):
        super().__init__(app_instance)
        self.app = app_instance
        self.parent_page = parent_page
        self.title("Configuración de Mediciones MPO")
        self.geometry("350x300")
        self.transient(self.app)
        self.grab_set()

        # Título
        lbl_title = ttk.Label(self, text="Seleccionar Mediciones Activas", font=("Helvetica", 12, "bold"))
        lbl_title.pack(pady=15)

        frame_switches = ttk.Frame(self)
        frame_switches.pack(fill="both", expand=True, padx=40)

        # Variables booleanas cargadas desde config
        self.var_ilrl = tk.BooleanVar(value=self.app.config.get('check_mpo_ilrl', True))
        self.var_geo = tk.BooleanVar(value=self.app.config.get('check_mpo_geo', True))
        self.var_pol = tk.BooleanVar(value=self.app.config.get('check_mpo_polaridad', True))

        # Switches (Checkbuttons) con estilo 'round-toggle' si usas ttkbootstrap
        chk_ilrl = ttk.Checkbutton(frame_switches, text="Medición IL/RL", variable=self.var_ilrl, bootstyle="round-toggle")
        chk_ilrl.pack(anchor="w", pady=10, fill='x')
        
        chk_geo = ttk.Checkbutton(frame_switches, text="Medición Geometría", variable=self.var_geo, bootstyle="round-toggle")
        chk_geo.pack(anchor="w", pady=10, fill='x')

        chk_pol = ttk.Checkbutton(frame_switches, text="Medición Polaridad", variable=self.var_pol, bootstyle="round-toggle")
        chk_pol.pack(anchor="w", pady=10, fill='x')

        # Botones de Acción
        btn_frame = ttk.Frame(self, padding=20)
        btn_frame.pack(side="bottom", fill="x")

        ttk.Button(btn_frame, text="Guardar Cambios", command=self.save_config, style="success.TButton").pack(side="right", padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=self.destroy, style="danger.TButton").pack(side="right", padx=5)

    def save_config(self):
        # Actualizamos solo las llaves relevantes en la configuración
        new_settings = self.app.config.copy()
        new_settings["check_mpo_ilrl"] = self.var_ilrl.get()
        new_settings["check_mpo_geo"] = self.var_geo.get()
        new_settings["check_mpo_polaridad"] = self.var_pol.get()
        
        self.app.save_config(new_settings)
        messagebox.showinfo("Configuración", "Preferencias actualizadas correctamente.", parent=self)
        self.destroy()

class VerificacionMPO_Page(ttk.Frame):

    def __init__(self, parent, app_instance):
        super().__init__(parent, padding=10)
        self.app = app_instance
        self.last_ilrl_result = None
        self.last_geo_result = None
        self.last_polaridad_result = None
        self.create_widgets()
        
    # En la clase VerificacionMPO_Page, reemplaza este método:
    def create_widgets(self):
        container = ttk.Frame(self, style='TFrame', padding=20)
        container.pack(expand=True, fill='both')
        top_frame = ttk.Frame(container)
        top_frame.pack(fill='x', pady=10)
        input_frame = ttk.Frame(top_frame)
        input_frame.pack(side='left', fill='x', expand=True)
        ttk.Label(input_frame, text="Número de OT:", font=("Helvetica", 11, "bold")).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.ot_entry = ttk.Entry(input_frame, width=30, font=("Helvetica", 11))
        self.ot_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)
        ttk.Label(input_frame, text="Número de Serie (13 dígitos):", font=("Helvetica", 11, "bold")).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.serie_entry = ttk.Entry(input_frame, width=30, font=("Helvetica", 11))
        self.serie_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)
        self.serie_entry.bind("<KeyRelease>", self.verificar_cable_automatico)

        # --- LÍNEA AÑADIDA ---
        # Conectamos nuestra nueva función al evento <FocusIn> (cuando se selecciona el campo)
        self.serie_entry.bind("<FocusIn>", self.on_serie_focus_in)
        # --------------------
        input_frame.columnconfigure(1, weight=1)
        action_buttons_frame = ttk.Frame(top_frame)
        action_buttons_frame.pack(side='left', padx=20)
        
        ot_details_button = ttk.Button(action_buttons_frame, text="Ver Detalles OT", command=self.open_ot_details_window, style='primary.Outline.TButton', padding=10)
        ot_details_button.pack(pady=5, fill='x')
        
        ot_config_button = ttk.Button(action_buttons_frame, text="Configurar OT", command=self.open_ot_config_window, style='info.TButton', padding=10)
        ot_config_button.pack(pady=5, fill='x')

        # --- NUEVO BOTÓN: Configuración de Búsqueda ---
        btn_search_config = ttk.Button(action_buttons_frame, text="⚙ Config. Búsqueda", command=self.open_search_config, style='secondary.TButton', padding=10)
        btn_search_config.pack(pady=5, fill='x')
        # ----------------------------------------------
        verify_button = ttk.Button(action_buttons_frame, text="Verificar (Manual)", command=self.verificar_cable, style='success.TButton', padding=10)
        verify_button.pack(pady=5, fill='x')

        self.result_text = tk.Text(container, height=15, width=80, wrap="word", font=("Courier New", 10), state=tk.DISABLED, relief="flat", bg="#f0f0f0")
        self.result_text.pack(fill='both', expand=True, pady=10)
        self.setup_text_tags()
        self.show_waiting_message()

    def open_search_config(self):
        MPOConfigWindow(self.app, self)
    
    def on_serie_focus_in(self, event=None):
        """
        Este método se activa cuando el campo de texto del número de serie
        recibe el foco (al hacer clic o al escanear).
        Selecciona todo el texto que haya en él.
        """
        self.serie_entry.select_range(0, tk.END)

    # En la clase VerificacionMPO_Page
    # En la clase VerificacionMPO_Page, reemplaza este método:

    def setup_text_tags(self):
        tags = {
            "header": {"font": ("Helvetica", 14, "bold"), "foreground": "#0056b3"},
            "bold": {"font": ("Courier New", 10, "bold")},
            "info": {"font": ("Courier New", 10, "italic"), "foreground": "grey"},
            "APROBADO": {"foreground": "#28a745"}, "PASS": {"foreground": "#28a745"},
            "RECHAZADO": {"foreground": "#dc3545"}, "FAIL": {"foreground": "#dc3545"},
            "NO ENCONTRADO": {"foreground": "#fd7e14"},
            "DATOS INCOMPLETOS": {"foreground": "#FF8C00", "font": ("Courier New", 10, "bold")},
            "DESACTIVADO": {"foreground": "#999999", "font": ("Courier New", 10, "italic")},
            "ERROR": {"foreground": "#dc3545", "font": ("Courier New", 10, "bold")},
            "final_status_large": {"font": ("Courier New", 14, "bold")}
        }
        for tag, options in tags.items():
            self.result_text.tag_configure(tag, **options)
            
        # --- BLOQUE MODIFICADO ---
        for test_type in ["ilrl", "geo", "polaridad"]:
            # 1. Hipervínculo para el ESTADO (abre detalles)
            link_tag = f"{test_type}_link"
            self.result_text.tag_configure(link_tag, foreground="#0056b3", underline=True)
            self.result_text.tag_bind(link_tag, "<Button-1>", lambda e, t=test_type: self.show_details_window(t))
            self.result_text.tag_bind(link_tag, "<Enter>", lambda e: self.result_text.config(cursor="hand2"))
            self.result_text.tag_bind(link_tag, "<Leave>", lambda e: self.result_text.config(cursor=""))

            # 2. Hipervínculo para el ARCHIVO (abre explorador)
            file_link_tag = f"{test_type}_file_link"
            self.result_text.tag_configure(file_link_tag, foreground="#4682B4", underline=True) # Un azul diferente (SteelBlue)
            self.result_text.tag_bind(file_link_tag, "<Button-1>", lambda e, t=test_type: self.open_file_location(t))
            self.result_text.tag_bind(file_link_tag, "<Enter>", lambda e: self.result_text.config(cursor="hand2"))
            self.result_text.tag_bind(file_link_tag, "<Leave>", lambda e: self.result_text.config(cursor=""))
        # --- FIN DEL BLOQUE ---

    def show_waiting_message(self):
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        self.result_text.insert(tk.END, "Esperando un número de serie válido (13 dígitos)...", "info")
        self.result_text.config(state=tk.DISABLED)
    
    def open_ot_details_window(self):
        ot_input = self.ot_entry.get().strip().upper()
        if not ot_input:
            messagebox.showwarning("Falta OT", "Por favor, ingrese un número de O.T.", parent=self)
            return
            
        # --- NORMALIZACIÓN: Usar siempre JMO- para buscar en BD ---
        # Extraemos solo los números (ej. 260200001)
        ot_numeros = re.sub(r'[^0-9]', '', ot_input)
        # Creamos la clave estándar para la base de datos
        ot_clave_bd = f"JMO-{ot_numeros}"
            
        ot_data = self._cargar_ot_configuration(ot_clave_bd)
        if not ot_data:
            # Si no existe, intentamos buscar tal cual lo escribió el usuario por seguridad
            ot_data = self._cargar_ot_configuration(ot_input)
            
        if not ot_data:
            messagebox.showinfo("No Encontrado", f"No se encontró configuración para la O.T.: {ot_clave_bd}", parent=self)
            return
        OTDetailsWindow(self, ot_data)

    def open_ot_config_window(self):
        ot_input = self.ot_entry.get().strip().upper()
        if not ot_input:
            messagebox.showwarning("Falta OT", "Ingrese un número de OT antes de configurar.", parent=self)
            return
            
        # --- NORMALIZACIÓN: Guardar siempre como JMO- en BD ---
        ot_numeros = re.sub(r'[^0-9]', '', ot_input)
        ot_clave_bd = f"JMO-{ot_numeros}"
        
        OTConfigWindow(self, self.app, ot_clave_bd)
    
    # En la clase VerificacionMPO_Page, reemplaza este método:

    def verificar_cable_automatico(self, event=None):
        serie_raw = self.serie_entry.get().strip()
        
        # --- MODIFICACIÓN: Regex ajustado para aceptar JMO o JRMO ---
        # Explicación: J(R)?MO significa que la 'R' es opcional.
        if re.match(r'J(R)?MO\d{13}', serie_raw, re.IGNORECASE):
            # Extraemos solo los números
            numeros_serie = re.sub(r'[^0-9]', '', serie_raw)
            
            if not self.ot_entry.get().strip():
                self.ot_entry.delete(0, tk.END)
                self.ot_entry.insert(0, f"JMO-{numeros_serie[:9]}")
            
            self.verificar_cable()
            
        elif len(serie_raw) == 13 and serie_raw.isdigit():
            if not self.ot_entry.get().strip():
                self.ot_entry.delete(0, tk.END)
                self.ot_entry.insert(0, f"JMO-{serie_raw[:9]}")
            self.verificar_cable()
        elif len(serie_raw) > 0:
            self.show_waiting_message()

    # En la clase VerificacionMPO_Page, reemplaza este método:

    def verificar_cable(self, event=None):
        ot_raw = self.ot_entry.get().strip().upper()
        serie_raw = self.serie_entry.get().strip()

        if not ot_raw or not serie_raw:
            self.show_waiting_message()
            return

        # 1. Normalización numérica
        serie_numerica = re.sub(r'[^0-9]', '', serie_raw)
        ot_numerica_input = re.sub(r'[^0-9]', '', ot_raw)

        if len(serie_numerica) != 13:
            messagebox.showerror("Formato Inválido", "El número de serie debe contener 13 dígitos.")
            self.show_waiting_message()
            return

        # Validación: Los primeros 9 dígitos del serial deben coincidir con la OT
        if ot_numerica_input != serie_numerica[:9]:
            messagebox.showerror("Error de Coincidencia", "La OT del número de serie no corresponde a la OT trabajada.")
            return

        # 2. Cargar Configuración (Usando la clave normalizada JMO-)
        ot_clave_bd = f"JMO-{ot_numerica_input}"
        ot_config = self._cargar_ot_configuration(ot_clave_bd)
        
        if not ot_config:
            messagebox.showwarning("Configuración Faltante", f"No se encontró configuración para {ot_clave_bd}.\nPor favor configúrela primero.", parent=self)
            return
            
        # 3. Preparar Interfaz
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        
        # Determinar prefijo real para mostrar (Estético y para logs)
        # Si el usuario puso JRMO o el serial es JRMO, usamos ese. Si no, JMO.
        if ot_raw.startswith("JRMO") or "JRMO" in serie_raw.upper():
            ot_display = f"JRMO-{ot_numerica_input}"
            prefijo_serie = "JRMO-"
        else:
            ot_display = f"JMO-{ot_numerica_input}"
            prefijo_serie = "JMO-"
            
        serie_completa = f"{prefijo_serie}{serie_numerica}"
        
        self.result_text.insert(tk.END, f"Verificando {serie_completa} en OT {ot_display}...\n", "header")
        self.result_text.insert(tk.END, "-"*70 + "\n\n")
        self.update_idletasks()
        
        # 4. Ejecutar Búsquedas
        # IMPORTANTE: Pasamos 'ot_numerica_input' (solo números) para que la búsqueda sea agnóstica al prefijo
        
        # IL/RL
        if self.app.config.get('check_mpo_ilrl', True):
            # Enviamos solo los números para que busque carpetas JMO o JRMO indistintamente
            self.last_ilrl_result = self.buscar_y_procesar_ilrl_mpo(ot_numerica_input, serie_completa, ot_config)
        else:
            self.last_ilrl_result = {'status': 'DESACTIVADO', 'details': 'Desactivado por usuario.', 'raw_data': []}

        # Geometría
        if self.app.config.get('check_mpo_geo', True):
            self.last_geo_result = self.buscar_y_procesar_geo_mpo(ot_numerica_input, serie_completa, ot_config)
        else:
            self.last_geo_result = {'status': 'DESACTIVADO', 'details': 'Desactivado por usuario.', 'raw_data': []}

        # Polaridad
        if self.app.config.get('check_mpo_polaridad', True):
            self.last_polaridad_result = self.buscar_y_procesar_polaridad_mpo(ot_numerica_input, serie_completa)
        else:
            self.last_polaridad_result = {'status': 'DESACTIVADO', 'details': 'Desactivado por usuario.', 'raw_data': {}}
        
        # Mostrar Resultados
        self.mostrar_resultado_mpo("IL/RL", self.last_ilrl_result)
        self.mostrar_resultado_mpo("Geometría", self.last_geo_result)
        self.mostrar_resultado_mpo("Polaridad", self.last_polaridad_result)
        
        # Semáforo Final
        ilrl_ok = self.last_ilrl_result['status'] in ["APROBADO", "DESACTIVADO"]
        geo_ok = self.last_geo_result['status'] in ["APROBADO", "DESACTIVADO"]
        pol_ok = self.last_polaridad_result['status'] in ["PASS", "APROBADO", "DESACTIVADO"]
        
        final_status = "APROBADO" if ilrl_ok and geo_ok and pol_ok else "RECHAZADO"
        
        self.result_text.insert(tk.END, "\n" + "-"*70 + "\n")
        self.result_text.insert(tk.END, "ESTADO FINAL: ", ("bold", "final_status_large"))
        self.result_text.insert(tk.END, f"{final_status}\n", (final_status, "final_status_large"))
        
        if winsound:
            try:
                if final_status == "APROBADO": winsound.Beep(1200, 200)
                elif final_status == "RECHAZADO": winsound.Beep(400, 500)
            except: pass

        self.result_text.config(state=tk.DISABLED)
        
        # Log (Guardamos la OT display para que se vea si fue retrabajo)
        log_data = {'serial_number': serie_completa, 'ot_number': ot_display, 'overall_status': final_status,
                    'ilrl_status': self.last_ilrl_result['status'], 'ilrl_details': json.dumps(self.last_ilrl_result, default=str),
                    'geo_status': self.last_geo_result['status'], 'geo_details': json.dumps(self.last_geo_result, default=str),
                    'polaridad_status': self.last_polaridad_result['status'], 'polaridad_details': json.dumps(self.last_polaridad_result, default=str)}
        self._log_mpo_verification(log_data)

    def _log_mpo_verification(self, log_data):
        try:
            conn = sqlite3.connect(self.app.config['db_path'])
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO cable_verifications (
                    entry_date, serial_number, ot_number, overall_status,
                    ilrl_status, ilrl_details, geo_status, geo_details,
                    polaridad_status, polaridad_details
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                log_data['serial_number'], log_data['ot_number'], log_data['overall_status'],
                log_data['ilrl_status'], log_data['ilrl_details'], log_data['geo_status'],
                log_data['geo_details'], log_data['polaridad_status'], log_data['polaridad_details']
            ))
            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("Error de Base de Datos", f"No se pudo registrar la verificación MPO: {e}", parent=self)

    # En la clase VerificacionMPO_Page, reemplaza este método:

    # En la clase VerificacionMPO_Page, reemplaza este método:

    def mostrar_resultado_mpo(self, tipo, resultado):
        """Muestra una línea de resultado en el widget de texto."""
        # Mapeo para los tags de DETALLES (Estado)
        tipo_map = {
            "IL/RL": "ilrl_link",
            "Geometría": "geo_link",
            "Polaridad": "polaridad_link"
        }
        # Mapeo para los tags de ARCHIVO (Explorador)
        file_tipo_map = {
            "IL/RL": "ilrl_file_link",
            "Geometría": "geo_file_link",
            "Polaridad": "polaridad_file_link"
        }
        
        link_tag = tipo_map.get(tipo, "") 
        file_link_tag = file_tipo_map.get(tipo, "") # <-- Nuevo tag de archivo

        self.result_text.insert(tk.END, f"{tipo}:\n", "bold")
        self.result_text.insert(tk.END, f"  Estado: ")

        status = resultado.get('status', 'ERROR')
        details = resultado.get('details', 'No hay detalles disponibles.')

        # Aplicar el tag de ESTADO (para detalles) solo al 'status'
        self.result_text.insert(tk.END, f"{status}", (status, link_tag))
        
        # --- INICIO DE LA MODIFICACIÓN: Aplicar el tag de ARCHIVO ---
        if "Archivo:" in details:
            details_text, file_text = details.rsplit("Archivo:", 1)
            file_text = "Archivo:" + file_text
            
            self.result_text.insert(tk.END, f"\n  Detalles: {details_text.strip()}")
            # Aplicar el tag de ARCHIVO (para explorador) solo al texto del archivo
            self.result_text.insert(tk.END, f"\n  {file_text.strip()}", (file_link_tag)) 
            self.result_text.insert(tk.END, "\n\n")
        else:
            self.result_text.insert(tk.END, f"\n  Detalles: {details}\n\n")
        # --- FIN DE LA MODIFICACIÓN ---

    def _cargar_ot_configuration(self, ot_number):
        try:
            conn = sqlite3.connect(self.app.config['db_path'])
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM ot_configurations WHERE ot_number = ?", (ot_number,))
            row = cursor.fetchone()
            conn.close()
            return dict(row) if row else None
        except Exception as e:
            messagebox.showerror("Error de Base de Datos", f"No se pudo cargar la configuración de la OT: {e}", parent=self)
            return None

    # En la clase VerificacionMPO_Page, reemplaza este método completo:

    def buscar_y_procesar_ilrl_mpo(self, ot_num, serie, config):
        base_path = self.app.config['ruta_base_ilrl_mpo']
        
        # 1. BÚSQUEDA DE CARPETA (Agnóstica a JMO/JRMO)
        # Buscamos cualquier carpeta que contenga el número de la OT (ej: 260200001)
        carpetas_candidatas = [os.path.join(base_path, d) for d in os.listdir(base_path) 
                               if str(ot_num) in d and os.path.isdir(os.path.join(base_path, d))]
        
        if not carpetas_candidatas:
            return {'status': 'NO ENCONTRADO', 'details': f'Carpeta de OT {ot_num} no encontrada en ILRL.', 'raw_data': []}
        
        # Si hay más de una (raro), tomamos la modificada más recientemente
        ruta_ot_ilrl = max(carpetas_candidatas, key=os.path.getmtime)
        
        # 2. BÚSQUEDA DE ARCHIVO (Agnóstica)
        # Buscamos archivos Excel que contengan el número de la OT
        archivos = [os.path.join(ruta_ot_ilrl, f) for f in os.listdir(ruta_ot_ilrl) 
                    if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$') and str(ot_num) in f]
        
        if not archivos:
            return {'status': 'NO ENCONTRADO', 'details': f'Archivo .xlsx no encontrado en {os.path.basename(ruta_ot_ilrl)}', 'raw_data': []}
            
        archivo_a_procesar = max(archivos, key=os.path.getmtime)
        
        # 3. LECTURA Y PROCESAMIENTO
        try:
            df = pd.read_excel(archivo_a_procesar, sheet_name="Results")
            
            # --- LIMPIEZA DE COLUMNAS (CRÍTICO) ---
            # Esto elimina espacios al inicio/final de los nombres de columnas
            # Así "conector number " será igual a "conector number"
            df.columns = [str(c).strip() for c in df.columns]
            
            # Mapeo de encabezados desde la configuración
            h = {k: config.get(v, d) for k, v, d in [
                ('serie', 'ilrl_serie_header', 'Serial number'), 
                ('estado', 'ilrl_estado_header', 'Alarm Status'), 
                ('conector', 'ilrl_conector_header', 'connector label'), 
                ('fecha', 'ilrl_fecha_header', 'Date'), 
                ('hora', 'ilrl_hora_header', 'Time')
            ]}
            
            # Validación de existencia de encabezados
            missing = [h_val for h_val in h.values() if h_val not in df.columns]
            if missing:
                return {'status': 'ERROR', 'details': f"Faltan encabezados en Excel: {', '.join(missing)}. Revise Configuración.", 'raw_data': []}

            # Preparación de datos
            df[h['serie']] = df[h['serie']].astype(str)
            df['base_serie'] = df[h['serie']].str.extract(r'(\d{13})')
            
            # Filtramos por el serial numérico buscado
            serie_numerica_buscada = re.sub(r'[^0-9]', '', serie)
            df_cable_group_all = df[df['base_serie'] == serie_numerica_buscada].copy()
            
            if df_cable_group_all.empty:
                return {'status': 'NO ENCONTRADO', 'details': f'Serie {serie} no encontrada en el archivo.', 'raw_data': []}

            # Construcción de Timestamp para comparar fechas
            date_series = pd.to_datetime(df_cable_group_all[h['fecha']], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y')
            time_series = df_cable_group_all[h['hora']].astype(str).str.replace('a. m.', 'AM', regex=False).str.replace('p. m.', 'PM', regex=False).str.strip()
            full_datetime_str = date_series + ' ' + time_series
            df_cable_group_all['timestamp'] = pd.to_datetime(full_datetime_str, format='%d/%m/%Y %I:%M:%S %p', errors='coerce')
            
            # Si hay filas con fechas inválidas, intentamos trabajar con las que sirvan
            df_cable_group_all.dropna(subset=['timestamp'], inplace=True)
            if df_cable_group_all.empty:
                 return {'status': 'ERROR', 'details': 'Fechas inválidas en las mediciones encontradas.', 'raw_data': []}

            # 4. PRIORIZACIÓN (RETRABAJOS vs ORIGINAL)
            # Buscamos cuál variación del serial (ej. ...001 o ...001-F) es la más reciente
            latest_timestamp = pd.NaT
            best_sub_group_df = None
            serie_para_reporte = "" 

            for full_serial in df_cable_group_all[h['serie']].unique():
                sub_group_df = df_cable_group_all[df_cable_group_all[h['serie']] == full_serial]
                current_max_timestamp = sub_group_df['timestamp'].max() 

                if best_sub_group_df is None or current_max_timestamp > latest_timestamp:
                    latest_timestamp = current_max_timestamp
                    best_sub_group_df = sub_group_df
                    serie_para_reporte = full_serial 
            
            df_cable = best_sub_group_df.copy()

            # 5. VERIFICACIÓN DE FIBRAS (Lado A y B)
            num_conectores_a = config.get('num_conectores_a', 1)
            fibras_por_conector_a = config.get('fibers_per_connector_a', 12)
            num_conectores_b = config.get('num_conectores_b', 1)
            fibras_por_conector_b = config.get('fibers_per_connector_b', 12)
            total_fibras_esperadas = (num_conectores_a * fibras_por_conector_a) + (num_conectores_b * fibras_por_conector_b)
            
            # Ordenamos por fecha descendente para tomar las últimas mediciones
            df_cable.sort_values(by='timestamp', ascending=False, inplace=True)
            
            # Filtramos Lado A y Lado B
            df_lado_a = df_cable[df_cable[h['conector']] == 'A'].head(fibras_por_conector_a * num_conectores_a)
            df_lado_b = df_cable[df_cable[h['conector']] == 'B'].head(fibras_por_conector_b * num_conectores_b)
            
            df_final_cable = pd.concat([df_lado_a, df_lado_b])

            # Construcción de datos para mostrar
            raw_data = []
            
            # Lado A
            mediciones_a = []
            estado_a = "APROBADO"
            for _, row in df_lado_a.iterrows():
                res = str(row[h['estado']]).strip().upper()
                mediciones_a.append({'fibra': len(mediciones_a)+1, 'resultado': res})
                if res != 'PASS': estado_a = "RECHAZADO"
            if not mediciones_a: estado_a = "NO ENCONTRADO"
            raw_data.append({'conector': 'A', 'estado': estado_a, 'mediciones': mediciones_a})

            # Lado B
            mediciones_b = []
            estado_b = "APROBADO"
            for _, row in df_lado_b.iterrows():
                res = str(row[h['estado']]).strip().upper()
                mediciones_b.append({'fibra': len(mediciones_b)+1, 'resultado': res})
                if res != 'PASS': estado_b = "RECHAZADO"
            if not mediciones_b: estado_b = "NO ENCONTRADO"
            raw_data.append({'conector': 'B', 'estado': estado_b, 'mediciones': mediciones_b})

            # Evaluación Final Global
            total_mediciones = len(df_final_cable)
            overall_pass = (estado_a == "APROBADO") and (estado_b == "APROBADO")
            
            if total_mediciones != total_fibras_esperadas: 
                overall_pass = False
                details = f"Faltan fibras: {total_mediciones}/{total_fibras_esperadas}. Archivo: {os.path.basename(archivo_a_procesar)}"
            else:
                # Contamos cuántos PASS reales tenemos en el grupo final
                pass_count = len(df_final_cable[df_final_cable[h['estado']].str.strip().str.upper() == 'PASS'])
                details = f"{pass_count}/{total_mediciones} fibras OK. Archivo: {os.path.basename(archivo_a_procesar)}"

            status = 'APROBADO' if overall_pass else 'RECHAZADO'
            
            return {'status': status, 'details': details, 'raw_data': raw_data, 'file_path': archivo_a_procesar, 'serial_number': serie_para_reporte}

        except Exception as e:
             return {'status': 'ERROR', 'details': f'Error procesando: {e}', 'raw_data': []}

    def buscar_y_procesar_geo_mpo(self, ot_num, serie, config):
        ruta_base_geo = self.app.config['ruta_base_geo_mpo']
        
        if not os.path.isdir(ruta_base_geo): 
            return {'status': 'NO ENCONTRADO', 'details': 'Carpeta de Geometría MPO no encontrada.', 'raw_data': []}
        
        archivos_encontrados = [os.path.join(self.app.config['ruta_base_geo_mpo'], f) 
                                for f in os.listdir(self.app.config['ruta_base_geo_mpo']) 
                                if str(ot_num) in f and f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
        
        if not archivos_encontrados: 
            return {'status': 'NO ENCONTRADO', 'details': f'Ningún archivo de Geometría para la OT "{ot_num}".', 'raw_data': []}
        
        archivo_a_procesar = max(archivos_encontrados, key=os.path.getmtime)

        try:
            full_df = pd.read_excel(archivo_a_procesar, sheet_name="MT12", header=None)
            header_row_index = next((i for i, row in full_df.iterrows() if str(row.iloc[0]).strip().lower() == 'name'), -1)
            
            if header_row_index == -1: 
                return {'status': 'ERROR', 'details': "No se encontró encabezado 'Name' en la hoja 'MT12'.", 'raw_data': []}
            
            df = full_df.iloc[header_row_index:].copy()
            df.columns = df.iloc[0].str.strip().str.lower()
            df = df[1:].rename(columns={'pass/fail': 'result'})
            
            df['temp_serie_numerica'] = df['name'].astype(str).str.extract(r'(\d{13})')
            serie_numerica_buscada = re.sub(r'[^0-9]', '', serie)
            
            # --- CORRECCIÓN 2: Corregido el nombre de la variable (df_cable_cable -> df_cable) ---
            df_cable = df[df['temp_serie_numerica'] == serie_numerica_buscada].copy()
            # -------------------------------------------------------------------------------------

            if df_cable.empty: 
                return {'status': 'NO ENCONTRADO', 'details': f'Serie no encontrada en {os.path.basename(archivo_a_procesar)}', 'raw_data': []}
            
            def parse_geo_name_mpo(name_str):
                match = re.match(r'(JMO\d{13})-(R?\d+)(-[FK])?', str(name_str).upper())
                if match: return match.group(1), match.group(2)
                return None, None
            
            parsed_data = df_cable['name'].apply(lambda x: pd.Series(parse_geo_name_mpo(x)))
            if parsed_data.empty or parsed_data.shape[1] < 2:
                 return {'status': 'ERROR', 'details': f'No se pudieron parsear los nombres de los conectores.', 'raw_data': []}

            df_cable[['serie_base', 'lado']] = parsed_data

            if 'date & time' in df_cable.columns:
                df_cable['timestamp'] = pd.to_datetime(df_cable['date & time'], dayfirst=True, errors='coerce')
            else:
                df_cable['timestamp'] = pd.to_datetime(df_cable['date'].astype(str) + ' ' + df_cable['time'].astype(str), dayfirst=True, errors='coerce')
            
            df_cable = df_cable.sort_values(by='timestamp', ascending=False)
            
            originales = {}
            retrabajos = {}
            for _, medicion in df_cable.iterrows():
                punta_original = str(medicion['lado'])
                punta_normalizada = punta_original.replace('R', '')
                if 'R' in punta_original:
                    if punta_normalizada not in retrabajos:
                        retrabajos[punta_normalizada] = {'Punta_Original': punta_original, 'Resultado': medicion['result'], 'Fuente': medicion['name']}
                else:
                    if punta_normalizada not in originales:
                        originales[punta_normalizada] = {'Punta_Original': punta_original, 'Resultado': medicion['result'], 'Fuente': medicion['name']}
            
            mediciones_definitivas = {}
            for punta, medicion in retrabajos.items():
                mediciones_definitivas[punta] = medicion
            for punta, medicion in originales.items():
                if punta not in mediciones_definitivas:
                    mediciones_definitivas[punta] = medicion

            # --- LÍNEA DE DEPURACIÓN ---
            print("\n*** MEDICIONES DEFINITIVAS ELEGIDAS ***")
            print(mediciones_definitivas)
            print("************************************\n")
            # ------------------------------------

            status = 'APROBADO'
            total_conectores_esperados = config.get('num_conectores_a', 0) + config.get('num_conectores_b', 0)
            lados_esperados = [str(i + 1) for i in range(total_conectores_esperados)]
            pass_count = 0
            
            for lado_req in lados_esperados:
                if lado_req in mediciones_definitivas:
                    if str(mediciones_definitivas[lado_req]['Resultado']).upper() == 'PASS':
                        pass_count += 1
                    else:
                        status = 'RECHAZADO'
                else:
                    status = 'RECHAZADO'
            
            if len(mediciones_definitivas) < total_conectores_esperados:
                status = 'RECHAZADO'
                details = f"Faltan mediciones. Encontradas: {len(mediciones_definitivas)}, Esperadas: {total_conectores_esperados}."
            else:
                details = f"{pass_count}/{total_conectores_esperados} mediciones OK."

            details += f" Archivo: {os.path.basename(archivo_a_procesar)}"
            raw_data = [{'conector': v['Punta_Original'], 'resultado': v['Resultado'], 'serie_completo': v['Fuente']} for v in mediciones_definitivas.values()]
            
            # --- LÍNEA MODIFICADA: Añadimos serial_number al retorno ---
            # Como todas las filas de df_cable pertenecen al mismo cable base, tomamos el primero
            primer_serial_completo = df_cable.iloc[0]['name'] if not df_cable.empty else serie
            
            return {'status': status, 'details': details, 'raw_data': raw_data, 'file_path': archivo_a_procesar, 'serial_number': primer_serial_completo}

        except Exception as e:
            return {'status': 'ERROR', 'details': f'Fallo al procesar {os.path.basename(archivo_a_procesar)}: {traceback.format_exc()}', 'raw_data': []}
        
    def buscar_y_procesar_polaridad_mpo(self, ot, serie):
        ruta_base_polaridad = self.app.config['ruta_base_polaridad_mpo']
        if not os.path.isdir(ruta_base_polaridad): return {'status': 'NO ENCONTRADO', 'details': 'Carpeta de Polaridad MPO no encontrada.', 'raw_data': {}}
        serie_sin_prefijo = re.sub(r'[^0-9]', '', serie)
        archivos_encontrados = [os.path.join(root, f) for root, _, files in os.walk(ruta_base_polaridad) for f in files if serie_sin_prefijo in f and f.lower().endswith('.xlsx') and not f.startswith('~$')]
        if not archivos_encontrados: return {'status': 'NO ENCONTRADO', 'details': f'Ningún archivo de Polaridad para la serie "{serie}".', 'raw_data': {}}
        archivo_a_procesar = max(archivos_encontrados, key=os.path.getmtime)
        try:
            df = pd.read_excel(archivo_a_procesar, sheet_name=0, header=None)
            serial_number_raw = df.iloc[3, 1] if df.shape[0] > 3 and df.shape[1] > 1 else None
            status_raw = df.iloc[12, 1] if df.shape[0] > 12 and df.shape[1] > 1 else None
            file_name = os.path.basename(archivo_a_procesar)
            date_match = re.search(r'(\d{4}-\d{2}-\d{2} \d{2}-\d{2}-\d{2})', file_name)
            date_raw = date_match.group(1).replace('-', ' ') if date_match else "N/A"
            if not serial_number_raw or not status_raw: return {'status': 'ERROR', 'details': f'Datos clave no encontrados en {file_name}', 'raw_data': {}}
            status = str(status_raw).strip().upper()
            details = f"Resultado de {file_name}. Fecha: {date_raw}"
            raw_data = {'serial_number': str(serial_number_raw), 'status': status, 'test_date': date_raw, 'file_name': file_name}
            return {'status': status, 'details': details, 'raw_data': raw_data, 'file_path': archivo_a_procesar}
        except Exception as e: return {'status': 'ERROR', 'details': f'Fallo al procesar {os.path.basename(archivo_a_procesar)}: {e}', 'raw_data': {}}

    def show_details_window(self, analysis_type):
        data = getattr(self, f"last_{analysis_type}_result", None)
        title = f"Detalles de Análisis {analysis_type.upper()} (MPO)"
        if not data or not data.get('raw_data'):
            messagebox.showinfo(title, "No hay datos detallados para mostrar.")
            return
        DetailsWindow(self, title, data, analysis_type)
    
    # En la clase VerificacionMPO_Page, añade esta nueva función:

    # En la clase VerificacionMPO_Page, reemplaza este método:

    # En la clase VerificacionMPO_Page, reemplaza este método:

    # En la clase VerificacionMPO_Page, reemplaza este método:

    # En la clase VerificacionMPO_Page, reemplaza este método:

    def open_file_location(self, analysis_type):
        """Abre la carpeta que contiene el archivo de reporte."""
        data = getattr(self, f"last_{analysis_type}_result", None)
        if not data or not data.get('file_path'):
            messagebox.showinfo("Error", "No se encontró la ruta del archivo.", parent=self)
            return
            
        file_path = data['file_path']
        if not os.path.exists(file_path):
            messagebox.showerror("Error", f"La ruta del archivo no existe:\n{file_path}", parent=self)
            return
            
        try:
            # --- INICIO DE LA CORRECCIÓN ---
            # 1. Obtener el DIRECTORIO que contiene el archivo
            folder_path = os.path.dirname(file_path)
            
            # 2. Asegurarse de que la ruta sea absoluta
            path_absoluto = os.path.abspath(folder_path)

            if not os.path.isdir(path_absoluto):
                messagebox.showerror("Error", f"La ruta de la carpeta no existe:\n{path_absoluto}", parent=self)
                return

            # 3. Usar os.startfile() - es la forma más robusta en Windows
            #    para abrir una carpeta en el explorador.
            os.startfile(path_absoluto)
            # --- FIN DE LA CORRECCIÓN ---
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la carpeta del archivo:\n{e}", parent=self)

class VerificacionFanout_Page(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.setup_ui()

    def setup_ui(self):
        # Título
        tk.Label(self, text="Verificación Individual (Fanout)", font=("Helvetica", 16, "bold"), fg="#2c3e50").pack(pady=20)

        # Contenedor principal
        main_frame = tk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=40)

        # Controles de Entrada (O.T. y Serie)
        input_frame = tk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=10)

        tk.Label(input_frame, text="Número de O.T.:", font=("Helvetica", 12)).grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.ot_entry = ttk.Entry(input_frame, font=("Helvetica", 12), width=25)
        self.ot_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(input_frame, text="Número de Serie:", font=("Helvetica", 12)).grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.serie_entry = ttk.Entry(input_frame, font=("Helvetica", 12), width=25)
        self.serie_entry.grid(row=1, column=1, padx=5, pady=5)
        self.serie_entry.bind('<Return>', lambda e: self.verificar_cable())

        # Botones
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=15)
        
        ttk.Button(btn_frame, text="Verificar Cable", command=self.verificar_cable).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Configurar O.T.", command=self.configurar_ot).pack(side=tk.LEFT, padx=10)

        # Área de resultados
        self.result_text = tk.Text(main_frame, height=20, width=80, font=("Consolas", 11), state=tk.DISABLED, bg="#f8f9fa")
        self.result_text.pack(pady=20, fill=tk.BOTH, expand=True)

    def _cargar_ot_configuration(self, ot_number):
        """Carga la configuración desde la tabla real (ot_configurations)"""
        try:
            conn = sqlite3.connect(self.app.config['db_path'])
            conn.row_factory = sqlite3.Row  # Esto convierte los resultados en un diccionario
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM ot_configurations WHERE ot_number = ?", (ot_number,))
            row = cursor.fetchone()
            conn.close()
            return dict(row) if row else None
        except Exception as e:
            print(f"Error cargando config: {e}")
        return None

    def verificar_cable(self, event=None):
        ot_raw = self.ot_entry.get().strip().upper()
        serie_raw = self.serie_entry.get().strip()

        if not ot_raw or not serie_raw:
            return

        # 1. Normalización
        serie_numerica = re.sub(r'[^0-9]', '', serie_raw)
        ot_numerica_input = re.sub(r'[^0-9]', '', ot_raw)
        ot_clave_bd = f"JMO-{ot_numerica_input}"

        if len(serie_numerica) != 13:
            messagebox.showerror("Formato Inválido", "El número de serie debe contener 13 dígitos.")
            return

        if ot_numerica_input != serie_numerica[:9]:
            messagebox.showerror("Error", "La OT del número de serie no coincide con la OT ingresada.")
            return

        # 2. Cargar Configuración
        ot_config = self._cargar_ot_configuration(ot_clave_bd)
        if not ot_config:
            messagebox.showwarning("Configuración Faltante", f"No se encontró configuración para {ot_clave_bd}.")
            return

        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        self.result_text.insert(tk.END, f"Verificando FANOUT {serie_raw}...\n", "bold")
        self.result_text.insert(tk.END, "-"*70 + "\n")
        self.update_idletasks()

        # 3. Ejecutar las 4 Búsquedas
        # IL/RL (100%)
        res_ilrl = self.buscar_y_procesar_ilrl_fanout(ot_numerica_input, serie_raw, ot_config)
        
        # Polaridad (100%)
        res_pol = self.buscar_y_procesar_polaridad_fanout(ot_numerica_input, serie_raw)
        
        # Geometría MPO (100%) -> Método propio para Fanout (1 punta, sin sufijos)
        res_geo_mpo = self.buscar_y_procesar_geo_mpo_fanout(ot_numerica_input, serie_raw)
        
        # Geometría LC/FC (10%)
        res_geo_lc = self.buscar_y_procesar_geo_lc_fanout(ot_numerica_input, serie_raw)

        # 4. Evaluar la Lógica del 10% para LC
        lote_liberado = ot_config.get("fanout_lote_liberado", False)
        
        if res_geo_lc['status'] == 'NO ENCONTRADO':
            if lote_liberado:
                res_geo_lc['status'] = "EXENTO (Muestreo OK)"
                res_geo_lc['details'] = "La O.T. ya fue liberada por calidad (10% cumplido)."
            else:
                res_geo_lc['status'] = "RECHAZADO (NO MUESTREADO)"
                res_geo_lc['details'] = "Cable sin medición y la O.T. AÚN NO ha sido liberada."

        # 5. Mostrar Resultados en Pantalla
        def imprimir_res(nombre, resultado):
            self.result_text.insert(tk.END, f"{nombre}:\n", "bold")
            self.result_text.insert(tk.END, f"  Estado:  {resultado['status']}\n")
            self.result_text.insert(tk.END, f"  Detalle: {resultado['details']}\n\n")

        imprimir_res("1. IL/RL (100%)", res_ilrl)
        imprimir_res("2. Polaridad (100%)", res_pol)
        imprimir_res("3. Geometría MPO (100%)", res_geo_mpo)
        imprimir_res("4. Geometría LC/FC (10%)", res_geo_lc)

        # 6. Semáforo Final
        ok_ilrl = res_ilrl['status'] in ["APROBADO"]
        ok_pol = res_pol['status'] in ["APROBADO", "PASS"]
        ok_geo_mpo = res_geo_mpo['status'] in ["APROBADO"]
        ok_geo_lc = res_geo_lc['status'] in ["APROBADO", "EXENTO (Muestreo OK)"]

        final_status = "APROBADO" if (ok_ilrl and ok_pol and ok_geo_mpo and ok_geo_lc) else "RECHAZADO"

        self.result_text.insert(tk.END, "-"*70 + "\n")
        self.result_text.insert(tk.END, f"ESTADO FINAL: {final_status}\n", "bold")
        
        if winsound:
            try:
                if final_status == "APROBADO": winsound.Beep(1200, 200)
                else: winsound.Beep(400, 500)
            except: pass

        self.result_text.config(state=tk.DISABLED)

    def configurar_ot(self):
        ot_input = self.ot_entry.get().strip().upper()
        if not ot_input:
            messagebox.showwarning("Falta OT", "Ingrese un número de OT antes de configurar.", parent=self)
            return
            
        # Normalizar a JMO- para guardar en BD
        ot_numeros = re.sub(r'[^0-9]', '', ot_input)
        ot_clave_bd = f"JMO-{ot_numeros}"
        
        # Abrimos la ventana de configuración estándar
        OTConfigWindow(self, self.app, ot_clave_bd)
    
    # --- NUEVOS MÉTODOS PARA FASE 2 ---

    def buscar_y_procesar_geo_lc_fanout(self, ot_num, serie_buscada):
        """
        Lee el reporte de Geometría DIMENSION para conectores LC/FC.
        Datos a partir de la Fila 13.
        Col B (1) = N.S. Secuencial, Col C (2) = Conector, Col I (8) = Pass/Fail
        """
        ruta_base = self.app.config.get('ruta_base_geo_fanout_lc', '') 
        
        if not ruta_base or not os.path.isdir(ruta_base):
            return {'status': 'ERROR', 'details': 'Ruta base de Geometría LC/FC no configurada o no existe.', 'raw_data': []}

        archivos_encontrados = [
            os.path.join(ruta_base, f) 
            for f in os.listdir(ruta_base) 
            if str(ot_num) in f and f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')
        ]

        if not archivos_encontrados:
            return {'status': 'NO ENCONTRADO', 'details': f'Sin archivo de Geometría LC para OT {ot_num}.', 'raw_data': []}

        archivo_a_procesar = max(archivos_encontrados, key=os.path.getmtime)
        
        # --- NUEVA LÓGICA DE SERIE: Extraer el secuencial (ej. "0001" -> "1") ---
        serie_numerica = re.sub(r'[^0-9]', '', serie_buscada)
        if len(serie_numerica) >= 4:
            secuencial_buscado = str(int(serie_numerica[-4:]))
        else:
            secuencial_buscado = serie_numerica

        try:
            # Leemos el Excel sin encabezados para evitar problemas con el formato
            df = pd.read_excel(archivo_a_procesar, sheet_name=0, header=None)
            
            # En pandas, la fila 13 de Excel es el índice 12
            if len(df) <= 12:
                return {'status': 'ERROR', 'details': 'El archivo tiene menos de 13 filas (formato incorrecto).', 'raw_data': []}
                
            # Cortamos el DataFrame para quedarnos solo con los datos puros
            df_datos = df.iloc[12:].copy()
            
            raw_data = []
            status_global = "APROBADO"
            puntas_encontradas = 0

            # Iteramos sobre las filas de datos
            for index, row in df_datos.iterrows():
                # Seguridad: Si la fila no tiene suficientes columnas, la ignoramos
                if len(row) < 9: continue
                
                # Columna B es el índice 1
                row_sn = str(row[1]).strip() 
                # Evitar que pandas lea el número como "1.0"
                if row_sn.endswith('.0'): row_sn = row_sn[:-2]
                
                # Si el secuencial coincide con la fila actual
                if row_sn == secuencial_buscado:
                    puntas_encontradas += 1
                    
                    # Columna C es el índice 2 (Número de conector)
                    conector_id = str(row[2]).strip()
                    if conector_id.endswith('.0'): conector_id = conector_id[:-2]
                    
                    # Columna I es el índice 8 (Pass/Fail)
                    resultado_texto = str(row[8]).strip().upper()
                    
                    res_punta = "RECHAZADO" if "FAIL" in resultado_texto else "APROBADO"
                    if res_punta == "RECHAZADO":
                        status_global = "RECHAZADO"

                    raw_data.append({
                        'conector': f"Punta {conector_id}",
                        'resultado': res_punta,
                        'serie_completo': serie_buscada
                    })

            # Si el cable no está en el Excel (porque solo prueban el 10%)
            if puntas_encontradas == 0:
                return {'status': 'NO ENCONTRADO', 'details': f'Secuencial {secuencial_buscado} NO muestreado en LC.', 'raw_data': []}

            details = f"Geometría LC: {puntas_encontradas} puntas medidas OK. Archivo: {os.path.basename(archivo_a_procesar)}"
            
            return {'status': status_global, 'details': details, 'raw_data': raw_data, 'file_path': archivo_a_procesar}

        except Exception as e:
            return {'status': 'ERROR', 'details': f'Error leyendo Geometría LC: {e}', 'raw_data': []}

    def buscar_y_procesar_ilrl_fanout(self, ot_num, serie_buscada, config):
        """
        Lee el reporte IL/RL SANTEC para Fanout.
        Busca el N.S. secuencial (ej. '1') y valida que la polaridad sea 'A-B'.
        """
        base_path = self.app.config.get('ruta_base_ilrl_mpo', '')
        
        if not base_path or not os.path.isdir(base_path):
            return {'status': 'ERROR', 'details': 'Ruta base de IL/RL no configurada.', 'raw_data': []}
            
        carpetas_candidatas = [
            os.path.join(base_path, d) for d in os.listdir(base_path) 
            if str(ot_num) in d and os.path.isdir(os.path.join(base_path, d))
        ]
        
        if not carpetas_candidatas:
            return {'status': 'NO ENCONTRADO', 'details': f'Carpeta de OT {ot_num} no encontrada en ILRL.', 'raw_data': []}
        
        ruta_ot_ilrl = max(carpetas_candidatas, key=os.path.getmtime)
        
        archivos = [
            os.path.join(ruta_ot_ilrl, f) for f in os.listdir(ruta_ot_ilrl) 
            if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$') and str(ot_num) in f
        ]
        
        if not archivos:
            return {'status': 'NO ENCONTRADO', 'details': f'Archivo .xlsx no encontrado.', 'raw_data': []}
            
        archivo_a_procesar = max(archivos, key=os.path.getmtime)
        
        try:
            df = pd.read_excel(archivo_a_procesar, sheet_name="Results")
            df.columns = [str(c).strip() for c in df.columns]
            
            h = {k: config.get(v, d) for k, v, d in [
                ('serie', 'ilrl_serie_header', 'Serial number'), 
                ('estado', 'ilrl_estado_header', 'Alarm Status'), 
                ('conector', 'ilrl_conector_header', 'connector label')
            ]}
            
            missing = [h_val for h_val in h.values() if h_val not in df.columns]
            if missing:
                return {'status': 'ERROR', 'details': f"Faltan encabezados: {', '.join(missing)}", 'raw_data': []}

            # --- NUEVA LÓGICA DE BÚSQUEDA ---
            # Extraemos los últimos 4 dígitos del N.S. y lo convertimos a entero para quitar ceros a la izquierda
            serie_numerica_completa = re.sub(r'[^0-9]', '', serie_buscada)
            if len(serie_numerica_completa) >= 4:
                secuencial_buscado = str(int(serie_numerica_completa[-4:])) # "0001" -> "1"
            else:
                secuencial_buscado = serie_numerica_completa

            # Filtramos el DataFrame buscando ese número secuencial exacto
            df[h['serie']] = df[h['serie']].astype(str).str.strip()
            df_cable = df[df[h['serie']] == secuencial_buscado].copy()
            
            if df_cable.empty:
                return {'status': 'NO ENCONTRADO', 'details': f'Secuencial {secuencial_buscado} no hallado.', 'raw_data': []}

            raw_data = []
            status_global = "APROBADO"
            pass_count = 0
            
            # Evaluamos las lecturas (Polaridad A-B y Status PASS)
            for index, row in df_cable.iterrows():
                estado = str(row.get(h['estado'], '')).strip().upper()
                polaridad = str(row.get(h['conector'], '')).strip().upper()
                
                # Validamos ambas condiciones
                if estado == 'PASS' and polaridad == 'A-B':
                    pass_count += 1
                    res_linea = "APROBADO"
                else:
                    status_global = "RECHAZADO"
                    res_linea = f"RECHAZADO ({estado} | {polaridad})"
                    
                raw_data.append({'conector': polaridad, 'estado': res_linea})
            
            if status_global == "APROBADO":
                details = f"Secuencial {secuencial_buscado}: {pass_count} lecturas 'A-B' OK."
            else:
                details = f"Secuencial {secuencial_buscado}: Rechazado por Estado o Polaridad incorrecta."
                
            return {'status': status_global, 'details': details, 'raw_data': raw_data, 'file_path': archivo_a_procesar}

        except Exception as e:
            return {'status': 'ERROR', 'details': f'Error en ILRL Fanout: {e}', 'raw_data': []}


    def buscar_y_procesar_polaridad_fanout(self, ot_num, serie_buscada):
        """
        Busca el reporte de Polaridad. Reutiliza la misma ruta y lógica que Polaridad MPO.
        """
        ruta_base = self.app.config.get('ruta_base_polaridad_mpo', '')
        
        if not ruta_base or not os.path.isdir(ruta_base):
            return {'status': 'ERROR', 'details': 'Ruta de Polaridad no configurada.', 'raw_data': {}}
            
        carpetas_ot = [
            os.path.join(ruta_base, d) for d in os.listdir(ruta_base) 
            if str(ot_num) in d and os.path.isdir(os.path.join(ruta_base, d))
        ]
        rutas_a_escanear = carpetas_ot if carpetas_ot else [ruta_base]
        
        serie_sin_prefijo = re.sub(r'[^0-9]', '', serie_buscada)
        archivos_candidatos = []
        
        for ruta in rutas_a_escanear:
            for root, _, files in os.walk(ruta):
                for f in files:
                    if serie_sin_prefijo in f and f.lower().endswith('.xlsx') and not f.startswith('~$'):
                        archivos_candidatos.append(os.path.join(root, f))
        
        if not archivos_candidatos: 
            return {'status': 'NO ENCONTRADO', 'details': f'Sin archivo de Polaridad para "{serie_buscada}".', 'raw_data': {}}
            
        archivo_a_procesar = max(archivos_candidatos, key=os.path.getmtime)
        
        try:
            folder_name = os.path.basename(os.path.dirname(archivo_a_procesar)).upper()
            status = "RECHAZADO" if ("FAIL" in folder_name or "RECHAZADO" in folder_name) else "APROBADO"
            details = f"Archivo en carpeta: {folder_name}"
            raw_data = {'archivo': os.path.basename(archivo_a_procesar), 'carpeta': folder_name}
            
            return {'status': status, 'details': details, 'raw_data': raw_data, 'file_path': archivo_a_procesar}
            
        except Exception as e:
            return {'status': 'ERROR', 'details': f'Error en Polaridad Fanout: {e}', 'raw_data': {}}
    
    def buscar_y_procesar_geo_mpo_fanout(self, ot_num, serie_buscada):
        """
        Busca el reporte de Geometría MPO específico para Fanout.
        Espera encontrar solo 1 medición sin sufijos (ej. JMO2603000010001).
        """
        ruta_base_geo = self.app.config.get('ruta_base_geo_mpo', '')
        
        if not ruta_base_geo or not os.path.isdir(ruta_base_geo): 
            return {'status': 'ERROR', 'details': 'Ruta de Geometría MPO no configurada.', 'raw_data': []}
        
        archivos_encontrados = [
            os.path.join(ruta_base_geo, f) for f in os.listdir(ruta_base_geo) 
            if str(ot_num) in f and f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')
        ]
        
        if not archivos_encontrados: 
            return {'status': 'NO ENCONTRADO', 'details': f'Sin archivo de Geometría MPO para OT {ot_num}.', 'raw_data': []}
        
        archivo_a_procesar = max(archivos_encontrados, key=os.path.getmtime)

        try:
            # Tolerancia al nombre de la pestaña
            try:
                full_df = pd.read_excel(archivo_a_procesar, sheet_name="MT12", header=None)
            except ValueError:
                full_df = pd.read_excel(archivo_a_procesar, sheet_name=0, header=None)

            serie_numerica_buscada = re.sub(r'[^0-9]', '', serie_buscada)
            
            raw_data = []
            status = "NO ENCONTRADO"
            details = "No se encontraron mediciones."

            for index, row in full_df.iterrows():
                row_str = " ".join([str(x).upper() for x in row.values])
                
                # Buscamos la coincidencia exacta de los 13 dígitos
                if serie_numerica_buscada in row_str:
                    line_status = "RECHAZADO" if "FAIL" in row_str else "APROBADO"
                    status = line_status
                    details = f"Medición MPO (1 punta) {line_status}."
                    
                    raw_data.append({
                        'conector': 'MPO (Lado A)',
                        'resultado': line_status,
                        'serie_completo': serie_buscada
                    })
                    break # Solo esperamos 1 punta, así que detenemos la búsqueda

            if status == "NO ENCONTRADO":
                details = f"Serie {serie_buscada} no hallada en {os.path.basename(archivo_a_procesar)}"

            return {'status': status, 'details': details, 'raw_data': raw_data, 'file_path': archivo_a_procesar}

        except Exception as e:
            return {'status': 'ERROR', 'details': f'Fallo al procesar: {e}', 'raw_data': []}

class AnalisisFanoutPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.setup_ui()

    def setup_ui(self):
        tk.Label(self, text="Liberación de Lote Fanout (AQL 10%)", font=("Helvetica", 16, "bold"), fg="#2c3e50").pack(pady=20)

        controls_frame = tk.Frame(self)
        controls_frame.pack(pady=10)

        tk.Label(controls_frame, text="Número de O.T.:", font=("Helvetica", 12)).grid(row=0, column=0, padx=5, sticky="e")
        self.ot_var = tk.StringVar()
        ttk.Entry(controls_frame, textvariable=self.ot_var, font=("Helvetica", 12), width=20).grid(row=0, column=1, padx=5)

        tk.Label(controls_frame, text="Cantidad Total de Cables:", font=("Helvetica", 12)).grid(row=1, column=0, padx=5, pady=10, sticky="e")
        self.total_var = tk.StringVar()
        ttk.Entry(controls_frame, textvariable=self.total_var, font=("Helvetica", 12), width=10).grid(row=1, column=1, padx=5, sticky="w")

        ttk.Button(self, text="Validar Lote Automáticamente", command=self.validar_lote_thread).pack(pady=20)

        self.result_text = tk.Text(self, height=18, width=85, font=("Consolas", 11), state=tk.DISABLED, bg="#f8f9fa")
        self.result_text.pack(pady=10, padx=20)

    def validar_lote_thread(self):
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        self.result_text.config(state=tk.DISABLED)
        threading.Thread(target=self.ejecutar_validacion, daemon=True).start()

    def ejecutar_validacion(self):
        ot_input = self.ot_var.get().strip().upper()
        total_str = self.total_var.get().strip()

        if not ot_input or not total_str.isdigit():
            self._log("ERROR: Ingrese O.T. y una cantidad válida numérica.", "error")
            return

        total_esperado = int(total_str)
        diez_por_ciento = max(1, int(total_esperado * 0.10))

        ot_numerica = re.sub(r'[^0-9]', '', ot_input)
        ot_clave_bd = f"JMO-{ot_numerica}"

        ot_config = self.app.pages["VerificacionFanout"]._cargar_ot_configuration(ot_clave_bd)
        if not ot_config:
            self._log(f"ERROR: No se encontró configuración en BD para {ot_clave_bd}. Configúrela primero.", "error")
            return

        self._log(f"Iniciando Validación (NUEVA LÓGICA DE VOLUMEN) para OT: {ot_clave_bd}")
        self._log(f"Meta Volumen: {total_esperado} mediciones aprobadas por estación")
        self._log(f"Meta Volumen (10%): {diez_por_ciento} mediciones aprobadas en Geo LC\n")
        self._log("-" * 75)

        # --- OBTENER CANTIDADES DE APROBADOS (Validación por Volumen) ---
        self._log("Consultando IL/RL (Estación final que define los N.S.)...")
        set_ilrl = self._obtener_set_ilrl(ot_numerica, ot_config)
        self._log(f"  -> Aprobados: {len(set_ilrl)} / {total_esperado}")

        self._log("Consultando Polaridad...")
        set_pol = self._obtener_set_polaridad(ot_numerica)
        self._log(f"  -> Aprobados: {len(set_pol)} / {total_esperado}")

        self._log("Consultando Geometría MPO...")
        set_geo_mpo = self._obtener_set_geo_mpo(ot_numerica)
        self._log(f"  -> Aprobados: {len(set_geo_mpo)} / {total_esperado}")

        self._log("Consultando Geometría LC/FC...")
        set_geo_lc = self._obtener_set_geo_lc(ot_numerica)
        self._log(f"  -> Aprobados: {len(set_geo_lc)} / {diez_por_ciento}")

        self._log("-" * 75 + "\n")

        # --- EVALUACIÓN DE VOLÚMENES ---
        pasa_ilrl = len(set_ilrl) >= total_esperado
        pasa_pol = len(set_pol) >= total_esperado
        pasa_geo_mpo = len(set_geo_mpo) >= total_esperado
        pasa_geo_lc = len(set_geo_lc) >= diez_por_ciento

        if pasa_ilrl and pasa_pol and pasa_geo_mpo and pasa_geo_lc:
            self._log("\n¡ÉXITO! Se cumplieron las cantidades requeridas en todas las estaciones.", "bold")
            self._log("Generando Master de Liberación con los N.S. extraídos de IL/RL...", "bold")
            
            # Generamos el Excel basándonos EXCLUSIVAMENTE en los cables de IL/RL
            self._generar_excel_liberacion(ot_numerica, set_ilrl)
            self.marcar_lote_liberado(ot_clave_bd, silencioso=False)
        else:
            self._log("\nLOTE NO LIBERADO. Faltan mediciones aprobadas en una o más estaciones.", "error")
            
            # Diagnóstico de faltantes por estación
            if not pasa_ilrl: 
                self._log(f"  • IL/RL: Faltan {total_esperado - len(set_ilrl)} mediciones aprobadas.", "error")
            if not pasa_pol: 
                self._log(f"  • Polaridad: Faltan {total_esperado - len(set_pol)} mediciones aprobadas.", "error")
            if not pasa_geo_mpo: 
                self._log(f"  • Geo MPO: Faltan {total_esperado - len(set_geo_mpo)} mediciones aprobadas.", "error")
            if not pasa_geo_lc: 
                self._log(f"  • Geo LC: Faltan {diez_por_ciento - len(set_geo_lc)} mediciones aprobadas.", "error")

    def _log(self, mensaje, tag=None):
        self.result_text.config(state=tk.NORMAL)
        if tag:
            self.result_text.insert(tk.END, mensaje + "\n", tag)
            self.result_text.tag_config("bold", font=("Consolas", 11, "bold"), foreground="green")
            self.result_text.tag_config("error", font=("Consolas", 11, "bold"), foreground="red")
        else:
            self.result_text.insert(tk.END, mensaje + "\n")
        self.result_text.config(state=tk.DISABLED)
        self.result_text.see(tk.END)
    
    def _extraer_4_digitos(self, sn_raw):
        """Convierte cualquier formato (ej. 1.0, 1, 2603000010001) a '0001' de forma segura"""
        sn_str = str(sn_raw).strip()
        if sn_str.endswith('.0'):
            sn_str = sn_str[:-2]
            
        sn_num = re.sub(r'[^0-9]', '', sn_str)
        if not sn_num: 
            return None
        if len(sn_num) >= 4:
            return sn_num[-4:]
        else:
            return f"{int(sn_num):04d}"

    # ================== MÉTODOS DE EXTRACCIÓN ==================

    def _obtener_set_ilrl(self, ot_num, config):
        try:
            ruta_base = self.app.config.get('ruta_base_ilrl_mpo', '')
            carpetas = [os.path.join(ruta_base, d) for d in os.listdir(ruta_base) if str(ot_num) in d]
            if not carpetas: 
                self._log("    [Alerta IL/RL] No se encontró la carpeta para esta O.T.")
                return set()
            
            carpeta_reciente = max(carpetas, key=os.path.getmtime)
            # IGNORAMOS archivos de Reporte creados por el sistema
            archivos = [os.path.join(carpeta_reciente, f) for f in os.listdir(carpeta_reciente) 
                        if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')
                        and "Reporte" not in f and "Analisis" not in f]
            if not archivos: 
                self._log("    [Alerta IL/RL] La carpeta está vacía o no tiene archivos crudos.")
                return set()
            
            archivos_ot = [f for f in archivos if str(ot_num) in os.path.basename(f)]
            archivo_a_procesar = max(archivos_ot, key=os.path.getmtime) if archivos_ot else max(archivos, key=os.path.getmtime)
            
            try: df = pd.read_excel(archivo_a_procesar, sheet_name="Results")
            except ValueError: df = pd.read_excel(archivo_a_procesar, sheet_name=0)
                
            df.columns = [str(c).strip() for c in df.columns]
            
            col_serie = config.get('ilrl_serie_header', 'Serial number')
            col_estado = config.get('ilrl_estado_header', 'Alarm Status')
            col_conector = config.get('ilrl_conector_header', 'connector label')
            
            if col_serie not in df.columns or col_estado not in df.columns: 
                self._log(f"    [Alerta IL/RL] Faltan encabezados en el Excel.")
                return set()
            
            pass_series, fail_series = set(), set()
            for _, row in df.iterrows():
                sn_norm = self._extraer_4_digitos(row[col_serie])
                if not sn_norm: continue
                
                polaridad = str(row.get(col_conector, 'A-B')).strip().upper()
                if polaridad == 'A-B':
                    if str(row[col_estado]).strip().upper() != 'PASS':
                        fail_series.add(sn_norm)
                    else:
                        pass_series.add(sn_norm)
            
            return pass_series - fail_series
        except Exception as e: 
            self._log(f"    [Error IL/RL] Fallo de lectura: {str(e)}")
            return set()

    def _obtener_set_polaridad(self, ot_num):
        try:
            ruta_base = self.app.config.get('ruta_base_polaridad_mpo', '')
            carpetas = [os.path.join(ruta_base, d) for d in os.listdir(ruta_base) if str(ot_num) in d]
            rutas_escanear = carpetas if carpetas else [ruta_base]
            
            valid_series = set()
            for ruta in rutas_escanear:
                for root, _, files in os.walk(ruta):
                    if "FAIL" in root.upper() or "RECHAZADO" in root.upper(): continue
                    for f in files:
                        if ot_num in f and f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$') and "Reporte" not in f and "Analisis" not in f:
                            match = re.search(r'(\d{13})', f)
                            if match: 
                                valid_series.add(self._extraer_4_digitos(match.group(1)))
            return valid_series
        except Exception as e: 
            self._log(f"    [Error Polaridad] Fallo de lectura: {str(e)}")
            return set()

    def _obtener_set_geo_mpo(self, ot_num):
        try:
            ruta_base = self.app.config.get('ruta_base_geo_mpo', '')
            archivos = [os.path.join(ruta_base, f) for f in os.listdir(ruta_base) 
                        if str(ot_num) in f and f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')
                        and "Reporte" not in f and "Analisis" not in f]
            if not archivos: 
                self._log("    [Alerta Geo MPO] No se encontraron archivos raw para esta O.T.")
                return set()
            
            # Un diccionario para que el último test (retrabajo exitoso) reemplace fallas previas
            cable_status = {}
            
            # Recorremos TODOS los archivos de la OT por si hay más de uno
            for archivo in archivos:
                try: df = pd.read_excel(archivo, sheet_name="MT12", header=None)
                except ValueError: df = pd.read_excel(archivo, sheet_name=0, header=None)

                for _, row in df.iterrows():
                    row_str = " ".join([str(x).upper() for x in row.values])
                    # Buscamos exactamente tu O.T. unida a 4 dígitos (Ej. 260300001 + 0005)
                    match = re.search(r'(' + str(ot_num) + r'\d{4})', row_str)
                    if match:
                        sn_norm = self._extraer_4_digitos(match.group(1))
                        if sn_norm:
                            if "FAIL" in row_str:
                                cable_status[sn_norm] = "FAIL"
                            else:
                                cable_status[sn_norm] = "PASS"
                                
            # Los que al final del día quedaron como PASS, se aprueban
            pass_series = {sn for sn, status in cable_status.items() if status == "PASS"}
            return pass_series
        except Exception as e: 
            self._log(f"    [Error Geo MPO] Fallo de lectura: {str(e)}")
            return set()

    def _obtener_set_geo_lc(self, ot_num):
        try:
            ruta_base = self.app.config.get('ruta_base_geo_fanout_lc', '')
            archivos = [os.path.join(ruta_base, f) for f in os.listdir(ruta_base) 
                        if str(ot_num) in f and f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')
                        and "Reporte" not in f and "Analisis" not in f]
            if not archivos: return set()
            
            archivo_a_procesar = max(archivos, key=os.path.getmtime)
            df = pd.read_excel(archivo_a_procesar, sheet_name=0, header=None)
            if len(df) <= 12: return set()
            df_datos = df.iloc[12:].copy()

            pass_series, fail_series = set(), set()
            for _, row in df_datos.iterrows():
                if len(row) < 9: continue
                sn_norm = self._extraer_4_digitos(row[1])
                if not sn_norm: continue
                
                if "FAIL" in str(row[8]).upper(): fail_series.add(sn_norm)
                else: pass_series.add(sn_norm)
            return pass_series - fail_series
        except Exception as e: 
            self._log(f"    [Error Geo LC] Fallo de lectura: {str(e)}")
            return set()

    # ================== GENERADOR DE REPORTE Y BASE DE DATOS ==================

    def _generar_excel_liberacion(self, ot_numerica, set_aprobados):
        try:
            # 1. Encontrar el Escritorio REAL (Considerando OneDrive y Español)
            home = os.path.expanduser("~")
            desktop_base = os.path.join(home, "Desktop")
            
            posibles_rutas = [
                os.path.join(home, "OneDrive", "Escritorio"),
                os.path.join(home, "OneDrive", "Desktop"),
                os.path.join(home, "Escritorio"),
                os.path.join(home, "Desktop")
            ]
            
            for item in os.listdir(home):
                if item.startswith("OneDrive -"):
                    posibles_rutas.insert(0, os.path.join(home, item, "Escritorio"))
                    posibles_rutas.insert(0, os.path.join(home, item, "Desktop"))

            for ruta in posibles_rutas:
                if os.path.exists(ruta):
                    desktop_base = ruta
                    break
            
            # 2. Crear la carpeta en el Escritorio real
            desktop_path = os.path.join(desktop_base, "Resultados de Fanout")
            os.makedirs(desktop_path, exist_ok=True)

            # 3. Construir la tabla con los N.S. Aprobados
            datos = []
            fecha_actual = pd.Timestamp.now().strftime("%d/%m/%Y %H:%M")
            
            for seq in sorted(set_aprobados):
                ns_final = f"E{ot_numerica}{seq}"
                datos.append({
                    "Número de Serie": ns_final,
                    "Estatus": "APROBADO",
                    "Fecha de Liberación": fecha_actual
                })

            df = pd.DataFrame(datos)
            
            # 4. Guardar el archivo Excel
            nombre_archivo = f"Liberacion_Fanout_JMO-{ot_numerica}.xlsx"
            ruta_completa = os.path.join(desktop_path, nombre_archivo)
            
            df.to_excel(ruta_completa, index=False)
            
            self._log(f"\nReporte Oficial creado en:\n{ruta_completa}", "bold")
            
            # 5. Abrir la carpeta automáticamente
            try:
                os.startfile(desktop_path)
            except Exception as e:
                pass
            
        except Exception as e:
            self._log(f"\nError al intentar crear el Excel: {str(e)}", "error")

    def marcar_lote_liberado(self, ot_clave, silencioso=False):
        try:
            conn = sqlite3.connect(self.app.config['db_path'])
            cursor = conn.cursor()
            try: cursor.execute("ALTER TABLE ot_configurations ADD COLUMN fanout_lote_liberado INTEGER DEFAULT 0")
            except sqlite3.OperationalError: pass
                
            cursor.execute("SELECT ot_number FROM ot_configurations WHERE ot_number = ?", (ot_clave,))
            row = cursor.fetchone()
            
            if row:
                cursor.execute("UPDATE ot_configurations SET fanout_lote_liberado = 1 WHERE ot_number = ?", (ot_clave,))
                conn.commit()
                if not silencioso: messagebox.showinfo("Liberado", "Lote Liberado. El reporte Excel se guardó en tu Escritorio.")
            else:
                if not silencioso: messagebox.showerror("Error", "O.T. no configurada.")
            conn.close()
        except Exception as e:
            if not silencioso: messagebox.showerror("Error BD", f"Error: {str(e)}")

class RegistroWHFanout_Page(ttk.Frame):
    def __init__(self, parent, app_instance):
        # Heredamos de ttk.Frame y aplicamos padding para mantener el mismo diseño
        super().__init__(parent, padding=20)
        self.app = app_instance

        # --- Creamos los widgets para esta página ---
        container = ttk.Frame(self, style='TFrame')
        container.pack(expand=True, fill='both', pady=20)

        # Un texto descriptivo
        ttk.Label(
            container,
            text="Módulo de Registro en Almacén (Fanout)",
            font=("Helvetica", 16, "bold")
        ).pack(pady=10)

        ttk.Label(
            container,
            text="Presiona el botón para abrir el archivo de Excel y lanzar el formulario de registro para empaque de productos Fanout.",
            font=("Helvetica", 11),
            wraplength=500
        ).pack(pady=10)

        # El botón que abrirá el archivo de Excel
        open_button = ttk.Button(
            container,
            text="Abrir Excel de Registro (Fanout)",
            command=self.abrir_registro_fanout,
            style='success.TButton',
            padding=15
        )
        open_button.pack(pady=30)

    def abrir_registro_fanout(self):
        """
        Busca y abre el MISMO archivo de registro que se usa para MPO, 
        ya que el usuario indicó que se compartirá la lógica en el mismo Excel.
        """
        file_name = "MPORegistroWH.xlsm"
        
        try:
            # Buscamos en el directorio base donde se ejecuta el programa
            base_path = os.path.dirname(sys.argv[0])
            file_path = os.path.join(base_path, file_name)

            if os.path.exists(file_path):
                os.startfile(file_path)
            else:
                messagebox.showerror(
                    "Archivo no Encontrado",
                    f"No se pudo encontrar el archivo '{file_name}'.\n\nAsegúrate de que esté en la misma carpeta que el programa."
                )
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al intentar abrir el archivo:\n\n{e}")

class OTConfigWindow(tk.Toplevel):
    def __init__(self, parent_page, app_instance, ot_number):
        super().__init__(parent_page)
        self.parent_page = parent_page
        self.app = app_instance
        self.ot_number = ot_number
        self.title(f"Configuración para OT: {ot_number}")
        self.geometry("800x700")
        self.transient(parent_page)
        self.grab_set()
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True)
        canvas = tk.Canvas(main_frame, highlightthickness=0)
        yscrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=yscrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        yscrollbar.pack(side="right", fill="y")
        def _on_canvas_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfigure(0, width=event.width)
        canvas.bind("<Configure>", _on_canvas_configure)
        frame = ttk.Frame(scrollable_frame, padding=20)
        frame.pack(fill='x', expand=True)
        
        # Diccionario de variables con nombres en INGLÉS para consistencia
        self.vars = {
            "drawing_number": tk.StringVar(), "link": tk.StringVar(),
            "num_conectores_a": tk.StringVar(value="1"), "fibers_per_connector_a": tk.StringVar(value="12"),
            "num_conectores_b": tk.StringVar(value="1"), "fibers_per_connector_b": tk.StringVar(value="12"),
            "ilrl_ot_header": tk.StringVar(value='Work number'), "ilrl_serie_header": tk.StringVar(value='Serial number'),
            "ilrl_fecha_header": tk.StringVar(value='Date'), "ilrl_hora_header": tk.StringVar(value='Time'),
            "ilrl_estado_header": tk.StringVar(value='Alarm Status'), "ilrl_conector_header": tk.StringVar(value='connector label')
        }
        self.load_existing_config()
        self.create_config_ui(frame)
        self.after(100, self.draw_mpo_cable_config)

    def create_config_ui(self, frame):
        ot_frame = ttk.LabelFrame(frame, text="Configuración General de OT", padding=10)
        ot_frame.pack(fill='x', expand=True, pady=5)
        ot_frame.columnconfigure(1, weight=1)
        ttk.Label(ot_frame, text="Número de Dibujo:").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        ttk.Entry(ot_frame, textvariable=self.vars["drawing_number"]).grid(row=0, column=1, sticky='ew', padx=5, pady=2)
        ttk.Label(ot_frame, text="Link a Dibujo:").grid(row=1, column=0, sticky='w', padx=5, pady=2)
        ttk.Entry(ot_frame, textvariable=self.vars["link"]).grid(row=1, column=1, sticky='ew', padx=5, pady=2)
        
        cable_frame = ttk.LabelFrame(frame, text="Configuración del Cable MPO", padding=10)
        cable_frame.pack(fill='x', expand=True, pady=5)
        cable_frame.columnconfigure(1, weight=1)
        cable_frame.columnconfigure(3, weight=1)
        
        # Lado A
        ttk.Label(cable_frame, text="Lado A - Conectores:").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        ttk.Entry(cable_frame, textvariable=self.vars["num_conectores_a"]).grid(row=0, column=1, sticky='ew', padx=5, pady=2)
        ttk.Label(cable_frame, text="Lado A - Fibras/Conector:").grid(row=0, column=2, sticky='w', padx=5, pady=2)
        ttk.Entry(cable_frame, textvariable=self.vars["fibers_per_connector_a"]).grid(row=0, column=3, sticky='ew', padx=5, pady=2)
        
        # Lado B
        ttk.Label(cable_frame, text="Lado B - Conectores:").grid(row=1, column=0, sticky='w', padx=5, pady=2)
        ttk.Entry(cable_frame, textvariable=self.vars["num_conectores_b"]).grid(row=1, column=1, sticky='ew', padx=5, pady=2)
        ttk.Label(cable_frame, text="Lado B - Fibras/Conector:").grid(row=1, column=2, sticky='w', padx=5, pady=2)
        ttk.Entry(cable_frame, textvariable=self.vars["fibers_per_connector_b"]).grid(row=1, column=3, sticky='ew', padx=5, pady=2)
        
        header_frame = ttk.LabelFrame(frame, text="Encabezados de Reporte ILRL", padding=10)
        header_frame.pack(fill='x', expand=True, pady=5)
        header_frame.columnconfigure(1, weight=1)
        headers = [("O.T.:", "ilrl_ot_header"), ("Nro. de Serie:", "ilrl_serie_header"), ("Fecha:", "ilrl_fecha_header"), ("Hora:", "ilrl_hora_header"), ("Estado:", "ilrl_estado_header"), ("Conector:", "ilrl_conector_header")]
        for i, (label, key) in enumerate(headers):
            ttk.Label(header_frame, text=f"Encabezado de {label}").grid(row=i, column=0, sticky=tk.W, pady=2, padx=5)
            ttk.Entry(header_frame, textvariable=self.vars[key]).grid(row=i, column=1, sticky=tk.EW, padx=5)
            
        canvas_frame = ttk.LabelFrame(frame, text="Vista Previa", padding=10)
        canvas_frame.pack(fill='both', expand=True, pady=5)
        self.cable_canvas = Canvas(canvas_frame, bg="white", height=200)
        self.cable_canvas.pack(fill='x', expand=True)
        
        ttk.Button(frame, text="Guardar Configuración", command=self.save_config, style="success.TButton").pack(pady=20)
        
        for key in ["num_conectores_a", "fibers_per_connector_a", "num_conectores_b", "fibers_per_connector_b"]:
            self.vars[key].trace_add("write", self.draw_mpo_cable_config)

    def load_existing_config(self):
        ot_data = self.parent_page._cargar_ot_configuration(self.ot_number)
        if ot_data:
            for key, var in self.vars.items():
                var.set(ot_data.get(key, var.get()))

    def draw_mpo_cable_config(self, *args):
        self.cable_canvas.delete("all")
        try:
            num_ca, fibras_ca, num_cb, fibras_cb = [int(self.vars[k].get()) for k in ["num_conectores_a", "fibers_per_connector_a", "num_conectores_b", "fibers_per_connector_b"]]
        except (ValueError, tk.TclError): return
        w, h = self.cable_canvas.winfo_width(), self.cable_canvas.winfo_height()
        if w < 50 or h < 50: return
        cy = h / 2
        self.cable_canvas.create_line(50, cy, w - 50, cy, width=10, fill="grey")
        total_h_a = num_ca * 20 + (num_ca - 1) * 5
        start_y_a = cy - total_h_a / 2
        for i in range(num_ca):
            y0 = start_y_a + i * 25
            self.cable_canvas.create_rectangle(20, y0, 50, y0 + 20, fill="black", outline="black")
            self.cable_canvas.create_text(35, y0 + 10, text=f"A{i+1}", fill="white", font=("Helvetica", 8, "bold"))
        total_h_b = num_cb * 20 + (num_cb - 1) * 5
        start_y_b = cy - total_h_b / 2
        for i in range(num_cb):
            y0 = start_y_b + i * 25
            self.cable_canvas.create_rectangle(w - 50, y0, w - 20, y0 + 20, fill="black", outline="black")
            self.cable_canvas.create_text(w - 35, y0 + 10, text=f"B{i+1}", fill="white", font=("Helvetica", 8, "bold"))
        self.cable_canvas.create_text(10, h - 10, anchor='sw', text=f"Total Fibras A: {num_ca * fibras_ca}", font=("Helvetica", 9))
        self.cable_canvas.create_text(w - 10, h - 10, anchor='se', text=f"Total Fibras B: {num_cb * fibras_cb}", font=("Helvetica", 9))

    def save_config(self):
        try:
            ot_data = {key: var.get() for key, var in self.vars.items()}
            ot_data['ot_number'] = self.ot_number
            for key in ["num_conectores_a", "fibers_per_connector_a", "num_conectores_b", "fibers_per_connector_b"]:
                ot_data[key] = int(ot_data[key])
            if self.app.guardar_ot_configuration(ot_data):
                messagebox.showinfo("Éxito", "Configuración guardada correctamente.", parent=self)
                self.destroy()
        except ValueError: 
            messagebox.showerror("Error de Entrada", "Los campos de conectores y fibras deben ser números enteros.", parent=self)
        except Exception as e: 
            messagebox.showerror("Error", f"No se pudo guardar la configuración: {e}", parent=self)

class OTDetailsWindow(tk.Toplevel):
    def __init__(self, parent, ot_data):
        super().__init__(parent)
        self.title(f"Detalles de OT: {ot_data['ot_number']}")
        self.geometry("600x450")
        self.transient(parent); self.grab_set()
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=f"Detalles para {ot_data['ot_number']}", font=("Helvetica", 16, "bold")).pack(anchor='w', pady=(0, 10))
        ttk.Label(frame, text=f"Número de Dibujo: {ot_data['drawing_number']}").pack(anchor='w')
        if ot_data.get('link'):
            link_label = ttk.Label(frame, text=f"Link: {ot_data['link']}", foreground="blue", cursor="hand2")
            link_label.pack(anchor='w')
            link_label.bind("<Button-1>", lambda e: webbrowser.open_new(ot_data['link']))
        ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=10)
        ttk.Label(frame, text="Configuración del Cable:", font=("Helvetica", 12, "bold")).pack(anchor='w', pady=(5,5))
        ttk.Label(frame, text=f"  Lado A: {ot_data['num_conectores_a']} conector(es) de {ot_data['fibers_per_connector_a']} fibras").pack(anchor='w')
        ttk.Label(frame, text=f"  Lado B: {ot_data['num_conectores_b']} conector(es) de {ot_data['fibers_per_connector_b']} fibras").pack(anchor='w')
        ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=10)
        ttk.Label(frame, text="Encabezados de Reporte ILRL:", font=("Helvetica", 12, "bold")).pack(anchor='w', pady=(5,5))
        for key, label in [('ilrl_ot_header', 'OT'), ('ilrl_serie_header', 'Serie'), ('ilrl_fecha_header', 'Fecha'), ('ilrl_hora_header', 'Hora'), ('ilrl_estado_header', 'Estado'), ('ilrl_conector_header', 'Conector')]:
            ttk.Label(frame, text=f"  {label}: '{ot_data[key]}'").pack(anchor='w')
        ttk.Button(frame, text="Cerrar", command=self.destroy, style='secondary.TButton').pack(pady=20)

class DetailsWindow(tk.Toplevel):
    # En la clase DetailsWindow, reemplaza el método __init__:

    def __init__(self, parent, title, data, analysis_type):
        super().__init__(parent)
        self.title(title); self.geometry("750x400"); self.transient(parent); self.grab_set() # Hice la ventana un poco más ancha
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=f"Detalles para: {data.get('details', 'N/A').split(' Archivo:')[0]}", wraplength=700).pack(anchor='w', pady=(0, 5))
        ttk.Label(frame, text=f"Archivo: {os.path.basename(data.get('file_path', 'N/A'))}", wraplength=700).pack(anchor='w', pady=(0, 10))
        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        cols, values_list = self.get_details_data(data, analysis_type)
        if not cols: return
        
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        
        # --- BLOQUE MODIFICADO: Ajuste dinámico de columnas ---
        for col in cols:
            tree.heading(col, text=col)
            if "Fuente" in col or "Serie" in col:
                tree.column(col, anchor='w', width=220)
            elif "Resultado" in col or "Conector" in col or "Punta" in col or "Línea" in col or "Fibra" in col:
                tree.column(col, anchor='center', width=80)
            else:
                tree.column(col, anchor='w', width=120)
        # --- FIN DEL BLOQUE ---
        
        tree.tag_configure('PASS', foreground='green'); tree.tag_configure('APROBADO', foreground='green')
        tree.tag_configure('FAIL', foreground='red'); tree.tag_configure('RECHAZADO', foreground='red')
        
        for values in values_list:
            tag_val = ""
            if len(values) > 1: tag_val = values[1] if analysis_type != 'polaridad' else (values[1] if values[0] == 'status' else '')
            tag = 'PASS' if str(tag_val) in ['PASS', 'APROBADO'] else 'FAIL' if str(tag_val) in ['FAIL', 'RECHAZADO'] else ''
            tree.insert("", "end", values=values, tags=(tag,))
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        tree.pack(side='left', fill='both', expand=True)
        ttk.Button(frame, text="Cerrar", command=self.destroy, style='secondary.TButton').pack(pady=20)

    # En la clase DetailsWindow, reemplaza este método:

    # En la clase DetailsWindow, reemplaza el método get_details_data:

    def get_details_data(self, data, analysis_type):
        raw_data = data.get('raw_data', [])
        # Extraer el serial number del nivel superior del diccionario 'data'
        serial_num = data.get('serial_number', 'N/A')

        if analysis_type == "ilrl":
            if data.get('error_type') == 'fechas_invalidas':
                cols = ("Conector", "Resultado", "Fecha (Original)", "Hora (Original)", "Número de Serie")
                values_list = [(
                    item.get('conector'), 
                    item.get('resultado'), 
                    item.get('fecha_original'), 
                    item.get('hora_original'),
                    serial_num # Añadir el N/S
                ) for item in raw_data]
                return cols, values_list

            # SC/LC (simple)
            if raw_data and isinstance(raw_data[0], dict) and 'linea' in raw_data[0]:
                cols = ("Línea", "Resultado", "Número de Serie")
                values_list = [(item.get('linea'), item.get('resultado'), serial_num) for item in raw_data]
                return cols, values_list
            # MPO (normal)
            else:
                cols = ("Conector", "Fibra", "Resultado", "Número de Serie")
                values_list = []
                for conn in raw_data:
                    for fib in conn.get('mediciones', []):
                        values_list.append((conn.get('conector'), fib.get('fibra'), fib.get('resultado'), serial_num))
                return cols, values_list
        
        elif analysis_type == "geo":
            # SC/LC
            if raw_data and isinstance(raw_data[0], dict) and 'punta' in raw_data[0]:
                cols = ("Punta", "Resultado", "Número de Serie (Fuente)")
                values_list = [(item.get('punta'), item.get('resultado'), item.get('fuente')) for item in raw_data]
                return cols, values_list
            # MPO
            else:
                cols = ("Conector", "Resultado", "Número de Serie (Fuente)")
                values_list = [(item.get('conector'), item.get('resultado'), item.get('serie_completo')) for item in raw_data]
                return cols, values_list
        
        elif analysis_type == "polaridad":
            return ("Propiedad", "Valor"), list(data.get('raw_data', {}).items())
        
        return (), []

class VerificacionUniboot_Page(ttk.Frame):
    def __init__(self, parent, app_instance):
        super().__init__(parent, padding=10)
        self.app = app_instance 
        self.last_ilrl_result = None 
        self.last_geo_result = None
        self.create_widgets()

    def create_widgets(self):
        container = ttk.Frame(self, style='TFrame', padding=20)
        container.pack(expand=True, fill='both')
        
        input_frame = ttk.Frame(container)
        input_frame.pack(fill='x', pady=10)
        
        ttk.Label(input_frame, text="Número de OT:", font=("Helvetica", 11, "bold")).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.ot_entry = ttk.Entry(input_frame, width=30, font=("Helvetica", 11))
        self.ot_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)
        
        ttk.Label(input_frame, text="Número de Serie (13 dígitos):", font=("Helvetica", 11, "bold")).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.serie_entry = ttk.Entry(input_frame, width=30, font=("Helvetica", 11))
        self.serie_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)
        self.serie_entry.bind("<KeyRelease>", self.verificar_cable_automatico)
        
        input_frame.columnconfigure(1, weight=1)

        verify_button = ttk.Button(container, text="Verificar Cable (Manual)", command=self.verificar_cable, style='success.TButton', padding=10)
        verify_button.pack(pady=20)

        self.result_text = tk.Text(container, height=15, width=80, wrap="word", font=("Courier New", 10), state=tk.DISABLED, relief="flat", bg="#f0f0f0")
        self.result_text.pack(fill='both', expand=True, pady=10)

        self.result_text.tag_configure("header", font=("Helvetica", 14, "bold"), foreground="#0056b3")
        self.result_text.tag_configure("bold", font=("Courier New", 10, "bold"))
        self.result_text.tag_configure("info", font=("Courier New", 10, "italic"), foreground="grey")
        self.result_text.tag_configure("APROBADO", foreground="#28a745")
        self.result_text.tag_configure("RECHAZADO", foreground="#dc3545")
        self.result_text.tag_configure("NO ENCONTRADO", foreground="#fd7e14")
        self.result_text.tag_configure("ERROR", foreground="#dc3545", font=("Courier New", 10, "bold"))
        
        # Tags dinámicos interactivos
        self.result_text.tag_configure("ilrl_link", foreground="#0056b3", underline=True)
        self.result_text.tag_bind("ilrl_link", "<Button-1>", lambda e: self.show_details_window("ilrl"))
        self.result_text.tag_bind("ilrl_link", "<Enter>", lambda e: self.result_text.config(cursor="hand2"))
        self.result_text.tag_bind("ilrl_link", "<Leave>", lambda e: self.result_text.config(cursor=""))
        
        self.result_text.tag_configure("geo_link", foreground="#0056b3", underline=True)
        self.result_text.tag_bind("geo_link", "<Button-1>", lambda e: self.show_details_window("geo"))
        self.result_text.tag_bind("geo_link", "<Enter>", lambda e: self.result_text.config(cursor="hand2"))
        self.result_text.tag_bind("geo_link", "<Leave>", lambda e: self.result_text.config(cursor=""))

        self.result_text.tag_configure("ilrl_file_link", foreground="#4682B4", underline=True)
        self.result_text.tag_bind("ilrl_file_link", "<Button-1>", lambda e: self.open_file_location("ilrl"))
        self.result_text.tag_bind("ilrl_file_link", "<Enter>", lambda e: self.result_text.config(cursor="hand2"))
        self.result_text.tag_bind("ilrl_file_link", "<Leave>", lambda e: self.result_text.config(cursor=""))
        
        self.result_text.tag_configure("geo_file_link", foreground="#4682B4", underline=True)
        self.result_text.tag_bind("geo_file_link", "<Button-1>", lambda e: self.open_file_location("geo"))
        self.result_text.tag_bind("geo_file_link", "<Enter>", lambda e: self.result_text.config(cursor="hand2"))
        self.result_text.tag_bind("geo_file_link", "<Leave>", lambda e: self.result_text.config(cursor=""))

        self.result_text.tag_configure("final_status_large", font=("Courier New", 14, "bold"))
        self.show_waiting_message()

    def show_waiting_message(self):
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        self.result_text.insert(tk.END, "Esperando un número de serie valido para UNIBOOT (13 dígitos)...", "info")
        self.result_text.config(state=tk.DISABLED)

    def _log_verification(self, log_data):
        try:
            conn = sqlite3.connect(self.app.config['db_path'])
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO cable_verifications (
                    entry_date, serial_number, ot_number, overall_status,
                    ilrl_status, ilrl_details, geo_status, geo_details, digital_seal
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                log_data['serial_number'], log_data['ot_number'], log_data['overall_status'],
                log_data['ilrl_status'], json.dumps(log_data['ilrl_details']),
                log_data['geo_status'], json.dumps(log_data['geo_details']),
                log_data.get('digital_seal', 'N/A') # <--- Inserción del sello
            ))
            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("Error BD", f"Error al registrar verificación Uniboot: {e}")

    def verificar_cable_automatico(self, event=None):
        serie_raw = self.serie_entry.get().strip()
        ot_numero = self.ot_entry.get().strip()
        
        if re.match(r'J(R)?MO\d{13}', serie_raw, re.IGNORECASE):
            numeros_serie = re.sub(r'[^0-9]', '', serie_raw)
            if not ot_numero:
                 self.ot_entry.insert(0, f"JMO-{numeros_serie[:9]}")
            self.verificar_cable()
        elif len(serie_raw) == 13 and serie_raw.isdigit():
            self.verificar_cable()
        else:
            self.show_waiting_message()

    def verificar_cable(self, event=None):
        ot_numero = self.ot_entry.get().strip().upper()
        serie_raw = self.serie_entry.get().strip()
        
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)

        if not ot_numero or not serie_raw:
            self.result_text.insert(tk.END, "ERROR: Por favor ingrese OT y Número de Serie.", "ERROR")
            self.result_text.config(state=tk.DISABLED)
            return

        serie_numerica = re.sub(r'[^0-9]', '', serie_raw)
        
        if len(serie_numerica) != 13:
            messagebox.showerror("Formato Inválido", "El número de serie debe contener 13 dígitos.")
            self.result_text.config(state=tk.DISABLED)
            return

        prefijo_serie = "JRMO-" if "JRMO" in serie_raw.upper() else "JMO-"
        serie_cable = f"{prefijo_serie}{serie_numerica}"
        
        ot_parte_input = re.sub(r'[^0-9]', '', ot_numero)
        serie_ot_parte = serie_numerica[:9]

        if ot_parte_input != serie_ot_parte:
            messagebox.showerror("Error", "La OT del N.S. no corresponde a la OT trabajada.")
            self.result_text.config(state=tk.DISABLED)
            return
        
        current_mode = self.app.cable_mode.get()
        self.result_text.insert(tk.END, f"Verificando UNIBOOT {serie_cable} en OT {ot_numero} (Modo: {current_mode})...\n", "header")
        self.result_text.insert(tk.END, "-"*60 + "\n\n")

        # 1. IL/RL (Lógica LC/SC)
        self.last_ilrl_result = self.buscar_y_procesar_ilrl_uniboot(ot_numero, serie_cable, current_mode)
        
        # 2. Geometría (Lógica Fanout)
        self.last_geo_result = self.buscar_y_procesar_geo_uniboot(ot_numero, serie_cable)
        
        self.mostrar_resultado("IL/RL", self.last_ilrl_result)
        self.mostrar_resultado("Geometría", self.last_geo_result)

        # --- Semáforo Final ---
        final_status = "NO ENCONTRADO"
        if self.last_ilrl_result['status'] not in ['NO ENCONTRADO', 'ERROR'] or self.last_geo_result['status'] not in ['NO ENCONTRADO', 'ERROR']:
            if self.last_ilrl_result['status'] == 'APROBADO' and self.last_geo_result['status'] == 'APROBADO':
                final_status = 'APROBADO'
            else:
                final_status = 'RECHAZADO'
        
        # --- Semáforo Final ---
        final_status = "NO ENCONTRADO"
        if self.last_ilrl_result['status'] not in ['NO ENCONTRADO', 'ERROR'] or self.last_geo_result['status'] not in ['NO ENCONTRADO', 'ERROR']:
            if self.last_ilrl_result['status'] == 'APROBADO' and self.last_geo_result['status'] == 'APROBADO':
                final_status = 'APROBADO'
            else:
                final_status = 'RECHAZADO'
        
        # --- CREACIÓN DEL SELLO DIGITAL ---
        # El Sello Digital será exactamente el Número de Serie (único e irrepetible)
        sello_digital = serie_cable
        
        self.result_text.insert(tk.END, "\n" + "-"*60 + "\n")
        self.result_text.insert(tk.END, "ESTADO FINAL: ", ("bold", "final_status_large"))
        self.result_text.insert(tk.END, f"{final_status}\n", (final_status, "final_status_large"))
        
        # Mostramos el sello en pantalla
        self.result_text.insert(tk.END, f"SELLO DIGITAL:  {sello_digital}\n", "header")
        self.result_text.insert(tk.END, "-"*60 + "\n")
        
        if winsound:
            try:
                if final_status == "APROBADO": winsound.Beep(1200, 200)
                elif final_status == "RECHAZADO": winsound.Beep(400, 500)
            except: pass
        
        self.result_text.config(state=tk.DISABLED)
        
        # Añadimos el sello al diccionario de log
        log_data = {
            'serial_number': serie_cable, 'ot_number': ot_numero, 'overall_status': final_status,
            'ilrl_status': self.last_ilrl_result['status'], 'ilrl_details': self.last_ilrl_result,
            'geo_status': self.last_geo_result['status'], 'geo_details': self.last_geo_result,
            'digital_seal': sello_digital # <--- Guardamos el sello (N.S.) en la base de datos
        }
        self._log_verification(log_data)
        
        log_data = {
            'serial_number': serie_cable, 'ot_number': ot_numero, 'overall_status': final_status,
            'ilrl_status': self.last_ilrl_result['status'], 'ilrl_details': self.last_ilrl_result,
            'geo_status': self.last_geo_result['status'], 'geo_details': self.last_geo_result
        }
        self._log_verification(log_data)

    def mostrar_resultado(self, tipo, resultado):
        link_tag = "ilrl_link" if tipo == "IL/RL" else "geo_link"
        file_link_tag = "ilrl_file_link" if tipo == "IL/RL" else "geo_file_link"

        self.result_text.insert(tk.END, f"Análisis {tipo}:\n", "bold")
        self.result_text.insert(tk.END, f"  Estado: ")
        self.result_text.insert(tk.END, f"{resultado['status']}", (resultado['status'], link_tag))
        
        details = resultado['details']
        if "Archivo:" in details:
            details_text, file_text = details.rsplit("Archivo:", 1)
            file_text = "Archivo:" + file_text
            self.result_text.insert(tk.END, f"\n  Detalles: {details_text.strip()}")
            self.result_text.insert(tk.END, f"\n  {file_text.strip()}", (file_link_tag)) 
            self.result_text.insert(tk.END, "\n\n")
        else:
            self.result_text.insert(tk.END, f"\n  Detalles: {details}\n\n")

    # ================== MÉTODOS HÍBRIDOS (LC/SC + FANOUT) ==================

    def buscar_y_procesar_ilrl_uniboot(self, ot, serie, mode):
        """Usa la lógica individual de LC/SC buscando los últimos 4 dígitos en la subcarpeta OK."""
        ruta_base = self.app.config.get('ruta_base_ilrl_uniboot', '')
        if not ruta_base or not os.path.isdir(ruta_base):
            return {'status': 'ERROR', 'details': 'Ruta ILRL Uniboot no configurada.', 'raw_data': []}

        serie_terminacion = serie[-4:]
        candidatos = []
        
        # --- MODIFICACIÓN UNIBOOT: Obligamos a buscar específicamente en la subcarpeta "OK" ---
        ruta_ot_ok = os.path.join(ruta_base, ot, "OK")
        
        if os.path.isdir(ruta_ot_ok):
            for root, _, files in os.walk(ruta_ot_ok):
                for f in files:
                    if f.endswith('.xlsx') and not f.startswith('~$') and serie_terminacion in f:
                        candidatos.append(os.path.join(root, f))

        if not candidatos:
            return {'status': 'NO ENCONTRADO', 'details': f'Ningún archivo con terminación "{serie_terminacion}" en la carpeta OK.', 'raw_data': []}
        
        archivo_a_procesar = max(candidatos, key=os.path.getmtime)
        
        try:
            expected_count = 2 if mode == "Simplex" else 4
            df = pd.read_excel(archivo_a_procesar, header=None)
            rows = df.iloc[12:]
            col7_vals = rows[7].dropna().astype(str).str.upper()
            col8_vals = rows[8].dropna().astype(str).str.upper()
            
            result_col = 8 if col8_vals.isin(['PASS', 'FAIL']).sum() >= col7_vals.isin(['PASS', 'FAIL']).sum() else 7
            results = rows[result_col].dropna().astype(str).str.upper().tolist()
            valid_results = [r for r in results if r in ['PASS', 'FAIL']]
            
            if not valid_results: 
                return {'status': 'RECHAZADO', 'details': 'No se encontraron mediciones.', 'raw_data': []}
                
            all_pass = all(r == 'PASS' for r in valid_results)
            count_ok = len(valid_results) == expected_count
            status = 'APROBADO' if all_pass and count_ok else 'RECHAZADO'
            
            details = f"{len(valid_results)}/{expected_count} mediciones encontradas."
            if not count_ok: details += f" Se esperaban {expected_count} para modo {mode}."
            if not all_pass: details += f" {valid_results.count('FAIL')} con FALLA."
            details += f" Archivo: {os.path.basename(archivo_a_procesar)}"
            
            raw_data = [{'linea': i + 1, 'resultado': res} for i, res in enumerate(valid_results)]
            return {'status': status, 'details': details, 'raw_data': raw_data, 'file_path': archivo_a_procesar}
        except Exception as e:
            return {'status': 'ERROR', 'details': f'Error ILRL: {e}', 'raw_data': []}

    def buscar_y_procesar_geo_uniboot(self, ot_num, serie_buscada):
        """Usa la lógica de Geometría de Fanout, extrayendo el Secuencial del lote y manejando Retrabajos (-R)."""
        ruta_base = self.app.config.get('ruta_base_geo_uniboot', '') 
        
        if not ruta_base or not os.path.isdir(ruta_base):
            return {'status': 'ERROR', 'details': 'Ruta de Geo Uniboot no configurada.', 'raw_data': []}

        archivos_encontrados = [
            os.path.join(ruta_base, f) for f in os.listdir(ruta_base) 
            if str(ot_num) in f and f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')
        ]

        if not archivos_encontrados:
            return {'status': 'NO ENCONTRADO', 'details': f'Sin archivo de Geometría Uniboot para OT {ot_num}.', 'raw_data': []}

        archivo_a_procesar = max(archivos_encontrados, key=os.path.getmtime)
        
        # Extraemos los últimos 4 dígitos y los convertimos a entero para comparar numéricamente (ej. "0001" -> 1)
        serie_numerica = re.sub(r'[^0-9]', '', serie_buscada)
        secuencial_buscado = int(serie_numerica[-4:]) if len(serie_numerica) >= 4 else int(serie_numerica)

        try:
            df = pd.read_excel(archivo_a_procesar, sheet_name=0, header=None)
            if len(df) <= 12:
                return {'status': 'ERROR', 'details': 'El archivo tiene menos de 13 filas.', 'raw_data': []}
                
            df_datos = df.iloc[12:].copy()
            
            # Diccionario para fusionar retrabajos: { '1': {'id_original': '1-R', 'res': 'APROBADO', 'es_retrabajo': True} }
            puntas_dict = {}

            for index, row in df_datos.iterrows():
                if len(row) < 9: continue
                
                row_sn = str(row[1]).strip() 
                if row_sn.endswith('.0'): row_sn = row_sn[:-2]
                
                # Intentamos convertir la celda a entero para que "0001", "1.0" y "1" coincidan
                try:
                    row_sn_int = int(row_sn)
                except ValueError:
                    continue # Ignorar si la celda tiene texto basura
                
                if row_sn_int == secuencial_buscado:
                    conector_id_original = str(row[2]).strip().upper()
                    if conector_id_original.endswith('.0'): conector_id_original = conector_id_original[:-2]
                    
                    # Identificamos la base de la punta (ej. "1-R" -> "1")
                    punta_base = conector_id_original.replace('-R', '').replace('R', '').strip()
                    es_retrabajo = '-R' in conector_id_original or 'R' in conector_id_original
                    
                    res_punta = "RECHAZADO" if "FAIL" in str(row[8]).strip().upper() else "APROBADO"
                    
                    # --- LÓGICA DE PRIORIDAD (RETRABAJOS) ---
                    if punta_base not in puntas_dict:
                        puntas_dict[punta_base] = {'id_original': conector_id_original, 'res': res_punta, 'es_retrabajo': es_retrabajo}
                    else:
                        existente = puntas_dict[punta_base]
                        # Si la nueva fila es retrabajo y la anterior no, sobrescribimos
                        if es_retrabajo and not existente['es_retrabajo']:
                            puntas_dict[punta_base] = {'id_original': conector_id_original, 'res': res_punta, 'es_retrabajo': es_retrabajo}
                        # Si ambas son retrabajos o ambas normales, tomamos la última (asumiendo que se midió después)
                        elif es_retrabajo == existente['es_retrabajo']:
                            puntas_dict[punta_base] = {'id_original': conector_id_original, 'res': res_punta, 'es_retrabajo': es_retrabajo}

            if not puntas_dict:
                secuencial_str = f"{secuencial_buscado:04d}" # Regresamos los ceros para el mensaje
                return {'status': 'NO ENCONTRADO', 'details': f'Secuencial {secuencial_str} no medido.', 'raw_data': []}

            # Preparamos el veredicto final con las puntas definitivas
            status_global = "APROBADO"
            raw_data = []
            
            for punta_base, info in puntas_dict.items():
                if info['res'] == "RECHAZADO":
                    status_global = "RECHAZADO"
                    
                raw_data.append({
                    'punta': f"Punta {info['id_original']}",
                    'resultado': info['res'],
                    'fuente': serie_buscada
                })

            details = f"Geometría Uniboot: {len(puntas_dict)} puntas medidas OK. Archivo: {os.path.basename(archivo_a_procesar)}"
            return {'status': status_global, 'details': details, 'raw_data': raw_data, 'file_path': archivo_a_procesar}

        except Exception as e:
            return {'status': 'ERROR', 'details': f'Error Geo Uniboot: {e}', 'raw_data': []}

    # ================== MÉTODOS DE VISUALIZACIÓN ==================
    def show_details_window(self, analysis_type):
        if analysis_type == "ilrl":
            data = self.last_ilrl_result
            title = "Detalles de Análisis IL/RL (Uniboot)"
        else:
            data = self.last_geo_result
            title = "Detalles de Geometría (Uniboot)"
            
        if not data or not data.get('raw_data'):
            messagebox.showinfo(title, "No hay datos detallados para mostrar.")
            return
            
        data['serial_number'] = self.serie_entry.get().strip()
        DetailsWindow(self, title, data, analysis_type)

    def open_file_location(self, analysis_type):
        data = self.last_ilrl_result if analysis_type == 'ilrl' else self.last_geo_result
        if not data or not data.get('file_path'): return
            
        file_path = data['file_path']
        if os.path.exists(file_path):
            os.startfile(os.path.abspath(os.path.dirname(file_path)))
        else:
            messagebox.showerror("Error", f"La ruta no existe:\n{file_path}", parent=self)

class Auditoria_LC_SC_Page(ttk.Frame):
    def __init__(self, parent, app_instance):
        super().__init__(parent, padding=20)
        self.app = app_instance
        self.analisis_ilrl = AnalisisILRL()
        self.analisis_geo = AnalisisGEO()
        self.create_widgets()

    def create_widgets(self):
        container = ttk.Frame(self, style='TFrame')
        container.pack(expand=True, fill='both')

        ctrl_frame = ttk.LabelFrame(container, text="Parámetros de Auditoría Automática", padding=15)
        ctrl_frame.pack(fill='x', pady=(0, 15))

        # ==========================================
        # FILA 1: ENTRADAS DE DATOS (Filtros limpios)
        # ==========================================
        inputs_frame = ttk.Frame(ctrl_frame)
        inputs_frame.pack(fill='x', pady=(0, 15))

        ttk.Label(inputs_frame, text="Número de O.T.:", font=("Helvetica", 11, "bold")).pack(side='left', padx=5)
        self.ot_var = tk.StringVar()
        ttk.Entry(inputs_frame, textvariable=self.ot_var, width=18, font=("Helvetica", 11)).pack(side='left', padx=5)

        ttk.Label(inputs_frame, text="Total Esperado:", font=("Helvetica", 11, "bold")).pack(side='left', padx=(20, 5))
        self.total_var = tk.StringVar()
        ttk.Entry(inputs_frame, textvariable=self.total_var, width=8, font=("Helvetica", 11)).pack(side='left', padx=5)

        # ==========================================
        # FILA 2: BOTONES DE ACCIÓN Y SWITCHES
        # ==========================================
        actions_frame = ttk.Frame(ctrl_frame)
        actions_frame.pack(fill='x')

        ttk.Button(actions_frame, text="▶ Ejecutar Auditoría", command=self.ejecutar_auditoria_thread, style='success.TButton').pack(side='left', padx=(5, 20))
        
        self.btn_excel = ttk.Button(actions_frame, text="📊 Descargar Reporte", command=self.descargar_reporte_excel, style='info.TButton', state=tk.DISABLED)
        self.btn_excel.pack(side='left', padx=5)
        
        self.btn_exportar = ttk.Button(actions_frame, text="☁️ Subir a Feishu", command=self.exportar_feishu_thread, style='info.TButton', state=tk.DISABLED)
        self.btn_exportar.pack(side='left', padx=5)

        self.liberar_var = tk.BooleanVar(value=False)
        self.chk_liberar = ttk.Checkbutton(actions_frame, text="✅ Aprobar Liberación Oficial", variable=self.liberar_var, bootstyle="success-round-toggle")
        self.chk_liberar.pack(side='left', padx=(30, 5)) 

        # --- Título de Resumen ---
        self.summary_label = ttk.Label(container, text="Ingrese la O.T. y cantidad. El sistema consolidará datos de TODAS las líneas de producción.", font=("Helvetica", 12, "italic"), foreground="#555555")
        self.summary_label.pack(anchor='w', pady=5)

        # =========================================================================
        # NUEVO: BARRA DE FILTROS POR COLUMNA (Búsqueda en Tiempo Real)
        # ==========================================
        filtros_frame = ttk.LabelFrame(container, text="🔍 Filtros de Búsqueda Rápida", padding=5)
        filtros_frame.pack(fill='x', pady=(0, 10))

        # Variables de memoria para cada columna (El trace_add hace que filtren mientras escribes)
        self.var_f_cable = tk.StringVar(); self.var_f_cable.trace_add("write", self.aplicar_filtros)
        self.var_f_ilrl = tk.StringVar(); self.var_f_ilrl.trace_add("write", self.aplicar_filtros)
        self.var_f_geo = tk.StringVar(); self.var_f_geo.trace_add("write", self.aplicar_filtros)
        self.var_f_estado = tk.StringVar(); self.var_f_estado.trace_add("write", self.aplicar_filtros)
        self.var_f_sello = tk.StringVar(); self.var_f_sello.trace_add("write", self.aplicar_filtros)

        # Cajas de texto alineadas horizontalmente
        ttk.Label(filtros_frame, text="Cable:").grid(row=0, column=0, padx=(5,2), sticky='w')
        ttk.Entry(filtros_frame, textvariable=self.var_f_cable, width=18).grid(row=0, column=1, padx=2)

        ttk.Label(filtros_frame, text="IL/RL:").grid(row=0, column=2, padx=(10,2), sticky='w')
        ttk.Entry(filtros_frame, textvariable=self.var_f_ilrl, width=12).grid(row=0, column=3, padx=2)

        ttk.Label(filtros_frame, text="Geo:").grid(row=0, column=4, padx=(10,2), sticky='w')
        ttk.Entry(filtros_frame, textvariable=self.var_f_geo, width=12).grid(row=0, column=5, padx=2)

        ttk.Label(filtros_frame, text="Estado:").grid(row=0, column=6, padx=(10,2), sticky='w')
        ttk.Entry(filtros_frame, textvariable=self.var_f_estado, width=15).grid(row=0, column=7, padx=2)

        ttk.Label(filtros_frame, text="Sello BD:").grid(row=0, column=8, padx=(10,2), sticky='w')
        ttk.Entry(filtros_frame, textvariable=self.var_f_sello, width=18).grid(row=0, column=9, padx=2)
        
        # Botón para borrar todo rápido
        ttk.Button(filtros_frame, text="✖ Limpiar", command=self.limpiar_filtros, style='secondary.TButton').grid(row=0, column=10, padx=(15,5))
        # =========================================================================

        columns = ("Cable", "IL/RL", "Geometría", "Estado Final", "Sello en BD")
        self.tree = ttk.Treeview(container, columns=columns, show="headings", height=15)
        self.tree = ttk.Treeview(container, columns=columns, show="headings", height=15)
        
        for col in columns:
            self.tree.heading(col, text=col)
            if col == "Cable": self.tree.column(col, anchor='w', width=200)
            elif col == "Sello en BD": self.tree.column(col, anchor='center', width=180)
            else: self.tree.column(col, anchor='center', width=120)

        self.tree.tag_configure('APROBADO', foreground='green', font=("Helvetica", 10, "bold"))
        self.tree.tag_configure('RECHAZADO', foreground='red', font=("Helvetica", 10, "bold"))
        self.tree.tag_configure('FALTANTE', foreground='#d35400', font=("Helvetica", 10, "bold"))
        self.tree.tag_configure('INTRUSO', foreground='purple', font=("Helvetica", 10, "bold"))
        self.tree.tag_configure('SCRAP CONFIRMADO', foreground='gray', font=("Helvetica", 10, "bold", "overstrike"))
        self.tree.tag_configure('SCRAP PENDIENTE', foreground='red', font=("Helvetica", 10, "bold"))

        scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        self.tree.bind("<Double-1>", self.mostrar_detalles_cable)

    def ejecutar_auditoria_thread(self):
        self.summary_label.config(text="Buscando archivos y analizando datos... Por favor espera.", foreground="#0056b3")
        self.tree.delete(*self.tree.get_children())
        threading.Thread(target=self._proceso_auditoria, daemon=True).start()

    def mostrar_detalles_cable(self, event):
        seleccion = self.tree.selection()
        if not seleccion: return
        valores = self.tree.item(seleccion[0])['values']
        if not valores: return

        cable = str(valores[0])
        detalles = getattr(self, 'detalles_auditoria', {}).get(cable)
        if not detalles:
            messagebox.showinfo("Sin Detalles", f"No hay información extra para {cable}.")
            return

        top = tk.Toplevel(self)
        top.title(f"Detalles de Auditoría - {cable}")
        top.geometry("550x450")
        top.transient(self.app)
        top.grab_set()

        def cerrar_si_clic_afuera(e):
            x, y = top.winfo_rootx(), top.winfo_rooty()
            w, h = top.winfo_width(), top.winfo_height()
            if not (x <= e.x_root <= x + w and y <= e.y_root <= y + h):
                top.destroy()
        top.bind("<Button-1>", cerrar_si_clic_afuera)

        ttk.Label(top, text=f"🔍 Informe Detallado", font=("Helvetica", 14, "bold")).pack(pady=(15,5))
        ttk.Label(top, text=cable, font=("Courier New", 12)).pack(pady=(0,15))
        frame = ttk.Frame(top, padding=20, relief="groove", borderwidth=2)
        frame.pack(fill="both", expand=True, padx=20, pady=5)

        ttk.Label(frame, text="Sello Digital BD:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, sticky="w", pady=5)
        ttk.Label(frame, text=detalles['sello'], font=("Courier New", 10)).grid(row=0, column=1, sticky="w", pady=5)
        ttk.Label(frame, text="Estado Final:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, sticky="w", pady=5)
        
        lbl_final = ttk.Label(frame, text=detalles['final'], font=("Helvetica", 10, "bold"))
        lbl_final.grid(row=1, column=1, sticky="w", pady=5)
        if detalles['final'] == 'APROBADO': lbl_final.config(foreground="green")
        elif detalles['final'] == 'FALTANTE': lbl_final.config(foreground="#d35400")
        else: lbl_final.config(foreground="red")

        ttk.Separator(frame).grid(row=2, column=0, columnspan=2, sticky="ew", pady=10)
        ttk.Label(frame, text="Resultados IL/RL", font=("Helvetica", 11, "bold"), foreground="#0056b3").grid(row=3, column=0, columnspan=2, sticky="w", pady=2)
        ttk.Label(frame, text=f"Estado: {detalles['ilrl']['estado']}", font=("Helvetica", 10, "italic")).grid(row=4, column=0, columnspan=2, sticky="w")
        ttk.Label(frame, text=detalles['ilrl']['detalle'], wraplength=450).grid(row=5, column=0, columnspan=2, sticky="w", pady=(2, 10))
        ttk.Label(frame, text="Resultados Geometría (DIMENSION)", font=("Helvetica", 11, "bold"), foreground="#0056b3").grid(row=6, column=0, columnspan=2, sticky="w", pady=2)
        ttk.Label(frame, text=f"Estado: {detalles['geo']['estado']}", font=("Helvetica", 10, "italic")).grid(row=7, column=0, columnspan=2, sticky="w")
        ttk.Label(frame, text=detalles['geo']['detalle'], wraplength=450).grid(row=8, column=0, columnspan=2, sticky="w", pady=(2, 10))

        ttk.Button(top, text="Cerrar", command=top.destroy, style="danger.TButton").pack(pady=15)

    def _proceso_auditoria(self):
        ot_raw = self.ot_var.get().strip().upper()
        total_raw = self.total_var.get().strip()

        if not ot_raw or not total_raw.isdigit():
            self.app.after(0, lambda: messagebox.showwarning("Datos Inválidos", "Ingrese una O.T. y una cantidad total válida."))
            self.app.after(0, lambda: self.summary_label.config(text="Error en los datos de entrada.", foreground="red"))
            return

        ot_num = re.sub(r'[^0-9]', '', ot_raw)
        ot_completa = f"JMO-{ot_num}"
        total = int(total_raw)
        self.detalles_auditoria = {}

        # 1. Recopilar TODAS las bases de datos configuradas en la planta
        rutas_bds = []
        for key in ['db_path_jws1_1', 'db_path_jws1_2', 'db_path_jws1_3', 'db_path']:
            path = self.app.config.get(key, '')
            if path and os.path.exists(path) and path not in rutas_bds:
                rutas_bds.append(path)

        if not rutas_bds:
            self.app.after(0, lambda: messagebox.showerror("Error", "No hay bases de datos configuradas o accesibles en red."))
            self.app.after(0, lambda: self.summary_label.config(text="Error de conexión a BD.", foreground="red"))
            return

        consolidado = []
        aprobados_count, rechazados_count, faltantes_count, scraps_count = 0, 0, 0, 0
        intrusos_encontrados = set()

        try:
            # 1. Extraer TODOS los cables y NORMALIZAR la llave para evitar duplicados JMO/JRMO
            cables_en_bd = {}
            for db_path in rutas_bds:
                try:
                    conn = sqlite3.connect(db_path, timeout=10)
                    conn.row_factory = sqlite3.Row
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT * FROM cable_verifications
                        WHERE ot_number = ? OR serial_number LIKE ?
                        ORDER BY id ASC
                    """, (ot_completa, f"%{ot_num}%"))
                    
                    for row in cursor.fetchall():
                        sn_real = row['serial_number']
                        # Aseguramos que la llave siempre tenga el mismo formato base
                        numeros = re.sub(r'[^0-9]', '', sn_real)
                        sn_normalizado = f"JMO-{numeros[:13]}" if len(numeros) >= 13 else sn_real
                        
                        row_dict = dict(row)
                        
                        if sn_normalizado not in cables_en_bd:
                            cables_en_bd[sn_normalizado] = row_dict
                        else:
                            fecha_existente = cables_en_bd[sn_normalizado].get('entry_date', '1900-01-01 00:00:00')
                            fecha_nueva = row_dict.get('entry_date', '1900-01-01 00:00:00')
                            if fecha_nueva > fecha_existente:
                                cables_en_bd[sn_normalizado] = row_dict
                    conn.close()
                except Exception as db_err:
                    print(f"Aviso: No se pudo leer {db_path} - {db_err}")

            # 2. CALCULAR EL LÍMITE REAL (Total esperado + Repuestos exactos por Scrap)
            scraps_reales = sum(1 for row in cables_en_bd.values() if 'SCRAP' in row.get('overall_status', ''))
            limite_secuencial = total + scraps_reales

            scraps_pendientes_list = []
            
            # 3. Evaluar SOLO desde el cable 1 hasta el límite lógico
            for i in range(1, limite_secuencial + 1):
                secuencial = str(i).zfill(4)
                cable_visual = f"{ot_completa}{secuencial}"
                
                row_encontrado = cables_en_bd.get(cable_visual)

                if row_encontrado:
                    sello_bd = row_encontrado['digital_seal'] if (row_encontrado['digital_seal'] and row_encontrado['digital_seal'] != "N/A") else row_encontrado['serial_number']
                    estado_final = row_encontrado['overall_status']
                    estado_ilrl = row_encontrado['ilrl_status']
                    estado_geo = row_encontrado['geo_status']
                    
                    try: ilrl_det = json.loads(row_encontrado['ilrl_details']).get('details', 'Detalle no disponible')
                    except: ilrl_det = 'Detalle no disponible'
                        
                    try: geo_det = json.loads(row_encontrado['geo_details']).get('details', 'Detalle no disponible')
                    except: geo_det = 'Detalle no disponible'

                    if estado_final in ['SCRAP', 'SCRAP CONFIRMADO']:
                        sello_bd = "SCRAP CONFIRMADO"
                        estado_final = "SCRAP CONFIRMADO"
                        estado_ilrl = "SCRAP CONFIRMADO"
                        estado_geo = "SCRAP CONFIRMADO"
                        ilrl_det = "Confirmado por Calidad."
                        geo_det = "Confirmado por Calidad."
                        scraps_count += 1
                    elif estado_final == 'SCRAP PENDIENTE':
                        sello_bd = "SCRAP PENDIENTE"
                        estado_final = "SCRAP PENDIENTE"
                        estado_ilrl = "PENDIENTE"
                        estado_geo = "PENDIENTE"
                        ilrl_det = "Enviado a Scrap por Producción (Falta confirmar)."
                        geo_det = "Enviado a Scrap por Producción (Falta confirmar)."
                        scraps_count += 1
                        scraps_pendientes_list.append(cable_visual)
                    elif estado_final == 'APROBADO':
                        aprobados_count += 1
                    else:
                        rechazados_count += 1
                else:
                    sello_bd = "SIN SELLO (No escaneado)"
                    estado_final = "FALTANTE"
                    estado_ilrl = "FALTANTE"
                    estado_geo = "FALTANTE"
                    ilrl_det = "Cable jamás escaneado en Verificación."
                    geo_det = "Cable jamás escaneado en Verificación."
                    faltantes_count += 1

                self.detalles_auditoria[cable_visual] = {
                    'sello': sello_bd, 
                    'final': estado_final,
                    'ilrl': {'estado': estado_ilrl, 'detalle': ilrl_det},
                    'geo': {'estado': estado_geo, 'detalle': geo_det}
                }
                consolidado.append((cable_visual, estado_ilrl, estado_geo, estado_final, sello_bd))

            # 4. Actualizar Interfaz (intrusos = 0 porque ya no usamos ese concepto para repuestos de la misma OT)
            self.app.after(0, self._actualizar_ui, consolidado, aprobados_count, rechazados_count, faltantes_count, 0, scraps_count, total)

            # --- ALERTA EMERGENTE DE SCRAPS PENDIENTES ---
            if scraps_pendientes_list:
                mensaje_alerta = f"¡ATENCIÓN!\n\nSe detectaron {len(scraps_pendientes_list)} cable(s) enviados a Scrap por Producción que AÚN NO han sido confirmados por FQC:\n\n"
                mensaje_alerta += "\n".join(scraps_pendientes_list[:10])
                if len(scraps_pendientes_list) > 10:
                    mensaje_alerta += f"\n... y {len(scraps_pendientes_list) - 10} más."
                mensaje_alerta += "\n\nPor favor, diríjase a 'Buscar Sello Digital' y confirme estos cables para autorizar el Scrap oficial."
                
                self.app.after(100, lambda: messagebox.showwarning("Scraps Pendientes de Confirmación", mensaje_alerta))

        except Exception as e:
            error_trace = traceback.format_exc()
            print(error_trace)
            self.app.after(0, lambda: messagebox.showerror("Error Crítico", f"Falló la auditoría consolidada:\n{e}"))

    def _actualizar_ui(self, consolidado, aprobados, rechazados, faltantes, intrusos, scraps, total):
        # --- NUEVO: Guardamos la tabla completa en memoria y llamamos al filtro ---
        self.datos_completos = consolidado
        self.aplicar_filtros()
        # --------------------------------------------------------------------------
        
        resumen = f"📊 RESULTADOS  |  Meta: {total}  |  ✅ Aprobados: {aprobados}  |  ❌ Rechazados: {rechazados}  |  ⚠️ Faltantes: {faltantes}"
        if intrusos > 0: resumen += f"  |  🚨 INTRUSOS: {intrusos}"
        if scraps > 0: resumen += f"  |  🗑️ SCRAP: {scraps}"
        
        estado_lote_final = "RECHAZADO"
        if aprobados >= total and intrusos == 0:
            estado_lote_final = "APROBADO"
            self.summary_label.config(text=resumen + "  (¡LOTE LISTO PARA LIBERAR!)", foreground="green", font=("Helvetica", 12, "bold"))
        elif intrusos > 0:
            self.summary_label.config(text=resumen + "  (¡ALERTA! HAY ARCHIVOS DE OTRA O.T.)", foreground="purple", font=("Helvetica", 12, "bold"))
        else:
            self.summary_label.config(text=resumen, foreground="#d35400", font=("Helvetica", 12, "bold"))
            
        self.datos_feishu_pendientes = {
            # ... lo que ya tenías
        }
        self.btn_exportar.config(state=tk.NORMAL)
        self.btn_excel.config(state=tk.NORMAL)

    def exportar_feishu_thread(self):
        self.btn_exportar.config(state=tk.DISABLED, text="Sincronizando...")
        threading.Thread(target=self._proceso_exportar_feishu, daemon=True).start()

    def _proceso_exportar_feishu(self):
        try:
            # ¡INGRESA TUS CÓDIGOS DE FEISHU AQUÍ!
            APP_ID = "tu_app_id_aqui" 
            APP_SECRET = "tu_app_secret_aqui"
            APP_TOKEN = "tu_app_token_aqui" 
            TABLE_ID = "tu_table_id_aqui" 
            
            feishu = FeishuIntegrator(APP_ID, APP_SECRET, APP_TOKEN, TABLE_ID)
            feishu.create_bitable_record(self.datos_feishu_pendientes)
            
            self.app.after(0, lambda: messagebox.showinfo("Sincronización Exitosa", "Subido a Feishu Bitable."))
            self.app.after(0, lambda: self.btn_exportar.config(text="☁️ Subido a Feishu", style='success.TButton'))
        except Exception as e:
            print(traceback.format_exc())
            self.app.after(0, lambda: messagebox.showerror("Error de Red", f"Fallo al subir:\n{e}"))
            self.app.after(0, lambda: self.btn_exportar.config(state=tk.NORMAL, text="☁️ Reintentar Feishu", style='danger.TButton'))
    
    def descargar_reporte_excel(self):
        if not hasattr(self, 'detalles_auditoria') or not self.detalles_auditoria:
            messagebox.showwarning("Sin Datos", "Primero debes ejecutar una auditoría.")
            return

        ot_num = re.sub(r'[^0-9]', '', self.ot_var.get())
        ot_completa = f"JMO-{ot_num}"
        total_esperado = int(self.total_var.get()) if self.total_var.get().isdigit() else 0

        # 1. Encontrar el Escritorio de forma segura
        home = os.path.expanduser("~")
        desktop_base = os.path.join(home, "Desktop")
        posibles_rutas = [
            os.path.join(home, "OneDrive", "Escritorio"),
            os.path.join(home, "OneDrive", "Desktop"),
            os.path.join(home, "Escritorio"),
            os.path.join(home, "Desktop")
        ]
        for ruta in posibles_rutas:
            if os.path.exists(ruta):
                desktop_base = ruta
                break

        carpeta_reportes = os.path.join(desktop_base, "Reportes de Auditoria FibraTrace")
        os.makedirs(carpeta_reportes, exist_ok=True)

        nombre_archivo = f"Reporte_VERIF_{ot_completa}.xlsx"
        ruta_completa = os.path.join(carpeta_reportes, nombre_archivo)

        try:
            wb = Workbook()
            
            # =========================================================
            # HOJA 1: NUEVA PORTADA DE INFORMACIÓN GENERAL
            # =========================================================
            ws_info = wb.active
            ws_info.title = "Información General"
            ws_info.sheet_view.showGridLines = False # Oculta la cuadrícula para que se vea como documento

            ws_info['A1'] = "REPORTE OFICIAL DE AUDITORÍA"
            ws_info['A1'].font = Font(size=18, bold=True, color="2C3E50")

            ws_info['A3'] = "Orden de Trabajo (O.T.):"
            ws_info['B3'] = ot_completa

            # Extraemos el nombre del auditor que inició sesión
            auditor = getattr(self.app, 'auditor_name', 'Auditor Desconocido')
            ws_info['A4'] = "Auditor Responsable:"
            ws_info['B4'] = auditor

            ws_info['A5'] = "Fecha y Hora de Generación:"
            ws_info['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            ws_info['A6'] = "Línea de Producción Auditada:"
            ws_info['B6'] = "Múltiples (Consolidado Automático BD)"
            
            # --- Si tiene ID de Liberación Oficial, lo añadimos ---
            quiere_liberar = self.liberar_var.get()
            if quiere_liberar:
                ws_info['A7'] = "ID de Liberación Oficial:"
                ws_info['B7'] = "SE ASIGNARÁ EN LA HOJA DE RESUMEN" # Se llenará en la siguiente hoja

            # Darle formato a la tablita
            for i in range(3, 8):
                ws_info[f'A{i}'].font = Font(bold=True)
                ws_info[f'A{i}'].alignment = Alignment(horizontal="right")
                ws_info[f'B{i}'].font = Font(color="0056b3", bold=True)
                ws_info[f'B{i}'].alignment = Alignment(horizontal="left")

            ws_info.column_dimensions['A'].width = 30
            ws_info.column_dimensions['B'].width = 35

            # =========================================================================
            # HOJA 2: RESUMEN EJECUTIVO DE O.T. (Antes era la Hoja 1)
            # =========================================================================
            ws_resumen = wb.create_sheet("Resumen O.T.")
            
            # Procesar métricas para el resumen
            total_procesados = len(self.detalles_auditoria)
            aprobados_count = 0
            rechazados_count = 0
            intrusos = []
            sin_sello = []
            scraped_cables = []

            for cable, datos in self.detalles_auditoria.items():
                estado = datos.get('final')
                if estado == 'APROBADO':
                    aprobados_count += 1
                elif estado == 'RECHAZADO':
                    rechazados_count += 1
                elif estado == 'INTRUSO':
                    intrusos.append(cable)
                elif estado in ['SCRAP', 'SCRAP CONFIRMADO', 'SCRAP PENDIENTE']:
                    scraped_cables.append(cable)
                
                # Buscar cables sin sello
                sello_str = str(datos.get('sello', '')).upper()
                if estado not in ['SCRAP', 'SCRAP CONFIRMADO', 'SCRAP PENDIENTE'] and ("SIN SELLO" in sello_str or "NO REGISTRADO" in sello_str or "ERROR" in sello_str):
                    sin_sello.append(cable)

            # --- LÓGICA FLEXIBLE DE APROBACIÓN DE LOTE (Manejo de Repuestos) ---
            # El lote se aprueba si: Alcanza la meta solicitada AND No hay cables sin verificar AND No hay rechazados
            lote_aprobado = (aprobados_count >= total_esperado) and (rechazados_count == 0) and (len(sin_sello) == 0)

            # --- POKA-YOKE DE LIBERACIÓN ---
            quiere_liberar = self.liberar_var.get()
            if quiere_liberar and not lote_aprobado:
                messagebox.showwarning("Liberación Bloqueada", "No puedes liberar una O.T. con estado RECHAZADO.\n\nEl reporte de auditoría se generará, pero sin la etiqueta de liberación oficial para empaque.", parent=self)
                quiere_liberar = False
                self.liberar_var.set(False)

            # Estilos
            bold_font = Font(bold=True)
            title_font = Font(size=16, bold=True, color="2C3E50")
            
            ws_resumen['A1'] = f"RESUMEN DE AUDITORÍA - {ot_completa}"
            ws_resumen['A1'].font = title_font
            
            # --- Tabla de Métricas ---
            ws_resumen['A3'] = "Cantidad de Cables Esperada:"
            ws_resumen['B3'] = total_esperado
            ws_resumen['A3'].font = bold_font

            ws_resumen['A4'] = "Cantidad Total Encontrada/Procesada:"
            ws_resumen['B4'] = total_procesados
            ws_resumen['A4'].font = bold_font

            ws_resumen['A5'] = "Cables Aprobados Correctamente:"
            ws_resumen['B5'] = aprobados_count
            ws_resumen['A5'].font = bold_font

            ws_resumen['A6'] = "Cantidad de Cables Intrusos:"
            ws_resumen['B6'] = len(intrusos)
            ws_resumen['A6'].font = bold_font
            
            ws_resumen['A7'] = "Cantidad de Cables SIN Sello Digital:"
            ws_resumen['B7'] = len(sin_sello)
            ws_resumen['A7'].font = bold_font

            ws_resumen['A8'] = "Cantidad de Cables en SCRAP:"
            ws_resumen['B8'] = len(scraped_cables)
            ws_resumen['A8'].font = bold_font

            # --- RECUADRO GIGANTE DE ESTADO (Semáforo) ---
            ws_resumen.merge_cells('E2:G6')
            c_estado = ws_resumen['E2']
            c_estado.value = "LOTE APROBADO" if lote_aprobado else "LOTE RECHAZADO"
            c_estado.font = Font(size=24, bold=True, color="FFFFFF")
            c_estado.alignment = Alignment(horizontal="center", vertical="center")
            if lote_aprobado:
                c_estado.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid") # Verde Éxito
            else:
                c_estado.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid") # Rojo Alerta

            # --- NUEVO: ETIQUETA DE LIBERACIÓN OFICIAL Y GENERACIÓN DE ID ---
            if quiere_liberar:
                # 1. Generamos el ID Único
                id_liberacion = self._generar_id_liberacion()

                # 2. Dibujamos el Banner Azul (En la fila 7)
                ws_resumen.merge_cells('E7:G7')
                c_lib = ws_resumen['E7']
                c_lib.value = "O.T. LIBERADA EXITOSAMENTE"
                c_lib.font = Font(size=14, bold=True, color="FFFFFF")
                c_lib.alignment = Alignment(horizontal="center", vertical="center")
                c_lib.fill = PatternFill(start_color="0056b3", end_color="0056b3", fill_type="solid")

                # 3. Dibujamos el ID Único debajo del banner (En la fila 8)
                ws_resumen.merge_cells('E8:G8')
                c_id = ws_resumen['E8']
                c_id.value = f"ID de Liberación: {id_liberacion}"
                c_id.font = Font(size=12, bold=True, color="000000")
                c_id.alignment = Alignment(horizontal="center", vertical="center")
                
                # 4. Alerta visual para el Auditor
                messagebox.showinfo("Lote Liberado Oficialmente", f"Se ha generado el documento de liberación para la O.T. {ot_completa}.\n\nID Asignado: {id_liberacion}\n\nEste código será requerido por Empaque para confirmar el lote.", parent=self)
            # --- Listados de Anomalías ---
            ws_resumen['A10'] = "🚨 Listado de Intrusos (Carpetas Equivocadas):"
            ws_resumen['A10'].font = Font(bold=True, color="5C005C")
            row_idx = 11
            if intrusos:
                for cab in intrusos:
                    ws_resumen[f'A{row_idx}'] = cab
                    row_idx += 1
            else:
                ws_resumen[f'A{row_idx}'] = "Ninguno detectado."
                
            ws_resumen['C10'] = "⚠️ Listado de Cables SIN Sello (Se saltaron Verificación):"
            ws_resumen['C10'].font = Font(bold=True, color="9C6500")
            row_idx_sello = 11
            if sin_sello:
                for cab in sin_sello:
                    ws_resumen[f'C{row_idx_sello}'] = cab
                    row_idx_sello += 1
            else:
                ws_resumen[f'C{row_idx_sello}'] = "Todos los cables están sellados."
            
            # --- NUEVA COLUMNA DE SCRAP ---
            ws_resumen['E10'] = "🗑️ Listado de N.S. Scrapeados:"
            ws_resumen['E10'].font = Font(bold=True, color="808080")
            row_idx_scrap = 11
            if scraped_cables:
                for cab in scraped_cables:
                    ws_resumen[f'E{row_idx_scrap}'] = cab
                    row_idx_scrap += 1
            else:
                ws_resumen[f'E{row_idx_scrap}'] = "Ninguno reportado."

            ws_resumen.column_dimensions['A'].width = 50
            ws_resumen.column_dimensions['C'].width = 55
            ws_resumen.column_dimensions['E'].width = 50

            # =========================================================================
            # HOJA 2: AUDITORIA O.T. (Desglose Detallado)
            # =========================================================================
            ws_detalles = wb.create_sheet("Auditoria O.T.")

            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") 
            header_font = Font(color="FFFFFF", bold=True)
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

            headers = ["Cable (N.S.)", "Sello Digital (BD)", "ESTADO FINAL", "Estado IL/RL", "Detalles IL/RL", "Estado Geometría", "Detalles Geometría"]
            ws_detalles.append(headers)

            for col_num, cell in enumerate(ws_detalles[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            for cable, datos in self.detalles_auditoria.items():
                row = [
                    cable,
                    datos.get('sello', 'N/A'),
                    datos.get('final', 'N/A'),
                    datos.get('ilrl', {}).get('estado', 'N/A'),
                    datos.get('ilrl', {}).get('detalle', 'N/A').replace('\n', ' | '), 
                    datos.get('geo', {}).get('estado', 'N/A'),
                    datos.get('geo', {}).get('detalle', 'N/A').replace('\n', ' | ')
                ]
                ws_detalles.append(row)

            for row in ws_detalles.iter_rows(min_row=2, max_col=7, max_row=ws_detalles.max_row):
                for i, cell in enumerate(row):
                    if i in [0, 1, 2, 3, 5]: 
                        cell.alignment = center_align
                    else: 
                        cell.alignment = left_align

                    if i == 2:
                        if cell.value == "APROBADO":
                            cell.font = Font(color="006100", bold=True)
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value == "RECHAZADO":
                            cell.font = Font(color="9C0006", bold=True)
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif cell.value == "FALTANTE":
                            cell.font = Font(color="9C6500", bold=True)
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        elif cell.value == "INTRUSO":
                            cell.font = Font(color="5C005C", bold=True)
                            cell.fill = PatternFill(start_color="E6B3E6", end_color="E6B3E6", fill_type="solid")
                        elif cell.value == "SCRAP": # <--- AÑADE ESTO
                            cell.font = Font(color="FFFFFF", bold=True)
                            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

            ws_detalles.column_dimensions['A'].width = 20
            ws_detalles.column_dimensions['B'].width = 23
            ws_detalles.column_dimensions['C'].width = 18
            ws_detalles.column_dimensions['D'].width = 15
            ws_detalles.column_dimensions['E'].width = 65 
            ws_detalles.column_dimensions['F'].width = 18
            ws_detalles.column_dimensions['G'].width = 65

            ws_detalles.auto_filter.ref = ws_detalles.dimensions

            # Guardar (sobrescribe en automático)
            wb.save(ruta_completa)
            os.startfile(ruta_completa)
            
        except PermissionError:
            messagebox.showerror("Archivo en Uso", f"El archivo {nombre_archivo} está abierto en Excel.\n\nPor favor ciérralo antes de generar un reporte nuevo para esta O.T.")
        except Exception as e:
            messagebox.showerror("Error", f"Fallo al generar el reporte Excel:\n{e}")
    
    def _generar_id_liberacion(self):
        """Genera un folio único consecutivo REP-FQC-XXXXX y lo guarda en config.json"""
        config_file = self.app.config_file
        try:
            with open(config_file, 'r') as f:
                config = json.load(f)
            
            # Buscamos el último ID usado (Si no existe, empezamos en 0)
            ultimo_id = config.get("last_fqc_id", 0)
            siguiente_id = ultimo_id + 1
            
            # Actualizamos y guardamos
            config["last_fqc_id"] = siguiente_id
            with open(config_file, 'w') as f:
                json.dump(config, f, indent=4)
                
            # Formateamos rellenando con ceros a la izquierda (ej. 00001)
            return f"REP-FQC-{str(siguiente_id).zfill(5)}"
        except Exception as e:
            print(f"Error generando ID: {e}")
            return "REP-FQC-ERROR"
    def aplicar_filtros(self, *args):
        """Dibuja la tabla en tiempo real basándose en lo que el usuario escriba en las cajas de filtro."""
        # 1. Limpiamos la tabla visual
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # 2. Obtenemos lo que el usuario escribió (convertido a minúsculas para que no importe mayúsculas/minúsculas)
        f_cable = self.var_f_cable.get().lower()
        f_ilrl = self.var_f_ilrl.get().lower()
        f_geo = self.var_f_geo.get().lower()
        f_estado = self.var_f_estado.get().lower()
        f_sello = self.var_f_sello.get().lower()
        
        # 3. Filtramos e insertamos
        if hasattr(self, 'datos_completos'):
            for row in self.datos_completos:
                # row es una tupla: (Cable, IL/RL, Geometría, Estado Final, Sello BD)
                if (f_cable in str(row[0]).lower() and
                    f_ilrl in str(row[1]).lower() and
                    f_geo in str(row[2]).lower() and
                    f_estado in str(row[3]).lower() and
                    f_sello in str(row[4]).lower()):
                    
                    # Si cumple con TODOS los filtros, lo pintamos en la tabla
                    self.tree.insert("", "end", values=row, tags=(row[3],))

    def limpiar_filtros(self):
        """Vacia las cajas de búsqueda para mostrar toda la tabla de nuevo."""
        self.var_f_cable.set("")
        self.var_f_ilrl.set("")
        self.var_f_geo.set("")
        self.var_f_estado.set("")
        self.var_f_sello.set("")

class RevisarLote_LC_SC_Page(ttk.Frame):
    def __init__(self, parent, app_instance):
        super().__init__(parent, padding=20)
        self.app = app_instance
        self.create_widgets()

    def create_widgets(self):
        container = ttk.Frame(self, style='TFrame')
        container.pack(expand=True, fill='both')

        ctrl_frame = ttk.LabelFrame(container, text="Parámetros de Revisión de Lote (Empaque)", padding=15)
        ctrl_frame.pack(fill='x', pady=(0, 15))

        # FILA 1: ENTRADAS DE DATOS (Más limpias)
        inputs_frame = ttk.Frame(ctrl_frame)
        inputs_frame.pack(fill='x', pady=(0, 15))

        ttk.Label(inputs_frame, text="Número de O.T.:", font=("Helvetica", 11, "bold")).pack(side='left', padx=5)
        self.ot_var = tk.StringVar()
        ttk.Entry(inputs_frame, textvariable=self.ot_var, width=18, font=("Helvetica", 11)).pack(side='left', padx=5)

        ttk.Label(inputs_frame, text="Cantidad de Lote:", font=("Helvetica", 11, "bold")).pack(side='left', padx=(20, 5))
        self.total_var = tk.StringVar()
        ttk.Entry(inputs_frame, textvariable=self.total_var, width=8, font=("Helvetica", 11)).pack(side='left', padx=5)

        # FILA 2: BOTONES DE ACCIÓN
        actions_frame = ttk.Frame(ctrl_frame)
        actions_frame.pack(fill='x')

        ttk.Button(actions_frame, text="▶ Revisar Lote", command=self.ejecutar_revision_thread, style='success.TButton').pack(side='left', padx=(5, 20))
        self.btn_excel = ttk.Button(actions_frame, text="📊 Descargar Reporte de Lote", command=self.descargar_reporte_excel, style='info.TButton', state=tk.DISABLED)
        self.btn_excel.pack(side='left', padx=5)

        self.summary_label = ttk.Label(container, text="Ingrese la O.T. y la cantidad. El sistema buscará inteligentemente en todas las líneas.", font=("Helvetica", 12, "italic"), foreground="#555555")
        self.summary_label.pack(anchor='w', pady=5)

        columns = ("Cable", "IL/RL", "Geometría", "Estado Final", "Sello en BD")
        self.tree = ttk.Treeview(container, columns=columns, show="headings", height=15)
        
        for col in columns:
            self.tree.heading(col, text=col)
            if col == "Cable": self.tree.column(col, anchor='w', width=200)
            elif col == "Sello en BD": self.tree.column(col, anchor='center', width=180)
            else: self.tree.column(col, anchor='center', width=120)

        self.tree.tag_configure('APROBADO', foreground='green', font=("Helvetica", 10, "bold"))
        self.tree.tag_configure('RECHAZADO', foreground='red', font=("Helvetica", 10, "bold"))
        self.tree.tag_configure('FALTANTE', foreground='#d35400', font=("Helvetica", 10, "bold"))
        self.tree.tag_configure('INTRUSO', foreground='purple', font=("Helvetica", 10, "bold"))
        self.tree.tag_configure('SCRAP', foreground='gray', font=("Helvetica", 10, "bold", "overstrike"))

        scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

    def ejecutar_revision_thread(self):
        self.summary_label.config(text="Procesando lote... Por favor espera.", foreground="#0056b3")
        self.tree.delete(*self.tree.get_children())
        threading.Thread(target=self._proceso_revision, daemon=True).start()

    def _proceso_revision(self):
        ot_raw = self.ot_var.get().strip().upper()
        total_raw = self.total_var.get().strip()

        if not ot_raw or not total_raw.isdigit():
            self.app.after(0, lambda: messagebox.showwarning("Datos Inválidos", "Ingrese una O.T. y cantidad de lote válida."))
            self.app.after(0, lambda: self.summary_label.config(text="Error en los datos de entrada.", foreground="red"))
            return

        ot_num = re.sub(r'[^0-9]', '', ot_raw)  
        ot_completa = f"JMO-{ot_num}"
        total = int(total_raw)
        self.detalles_auditoria = {}

        # 1. Recopilar todas las bases de datos configuradas en la planta
        rutas_bds = []
        for key in ['db_path_jws1_1', 'db_path_jws1_2', 'db_path_jws1_3', 'db_path']:
            path = self.app.config.get(key, '')
            if path and os.path.exists(path) and path not in rutas_bds:
                rutas_bds.append(path)

        if not rutas_bds:
            self.app.after(0, lambda: messagebox.showerror("Error", "No hay bases de datos configuradas o accesibles en red."))
            self.app.after(0, lambda: self.summary_label.config(text="Error de conexión a BD.", foreground="red"))
            return

        consolidado = []
        aprobados_count, rechazados_count, faltantes_count, scraps_count = 0, 0, 0, 0

        try:
            # --- NUEVO: Extraer TODOS los cables de la BD de una sola vez y comparar fechas ---
            cables_en_bd = {}
            for db_path in rutas_bds:
                try:
                    conn = sqlite3.connect(db_path, timeout=10)
                    conn.row_factory = sqlite3.Row
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT * FROM cable_verifications 
                        WHERE ot_number = ? OR serial_number LIKE ? 
                        ORDER BY id ASC
                    """, (ot_completa, f"%{ot_num}%"))
                    
                    for row in cursor.fetchall():
                        sn = row['serial_number']
                        row_dict = dict(row)
                        
                        if sn not in cables_en_bd:
                            cables_en_bd[sn] = row_dict
                        else:
                            fecha_existente = cables_en_bd[sn].get('entry_date', '1900-01-01 00:00:00')
                            fecha_nueva = row_dict.get('entry_date', '1900-01-01 00:00:00')
                            if fecha_nueva > fecha_existente:
                                cables_en_bd[sn] = row_dict
                    conn.close()
                except:
                    pass

            # --- CALCULAR EL LÍMITE REAL (Total esperado + Repuestos exactos por Scrap) ---
            scraps_reales = sum(1 for row in cables_en_bd.values() if 'SCRAP' in row.get('overall_status', ''))
            limite_secuencial = total + scraps_reales

            # 2. Iterar SOLO desde el cable 1 hasta el límite lógico
            for i in range(1, limite_secuencial + 1):
                secuencial = str(i).zfill(4)
                cable_visual = f"{ot_completa}{secuencial}"
                
                row_encontrado = cables_en_bd.get(cable_visual)

                # 4. Procesar el resultado oficial directamente de la Base de Datos
                if row_encontrado:
                    sello_bd = row_encontrado['digital_seal'] if (row_encontrado['digital_seal'] and row_encontrado['digital_seal'] != "N/A") else row_encontrado['serial_number']
                    estado_final = row_encontrado['overall_status']
                    estado_ilrl = row_encontrado['ilrl_status']
                    estado_geo = row_encontrado['geo_status']

                    if estado_final in ['SCRAP', 'SCRAP CONFIRMADO']:
                        sello_bd = "SCRAP CONFIRMADO"
                        estado_final = "SCRAP CONFIRMADO"
                        estado_ilrl = "SCRAP CONFIRMADO"
                        estado_geo = "SCRAP CONFIRMADO"
                        scraps_count += 1
                    elif estado_final == 'SCRAP PENDIENTE':
                        sello_bd = "SCRAP PENDIENTE"
                        estado_final = "SCRAP PENDIENTE"
                        estado_ilrl = "PENDIENTE"
                        estado_geo = "PENDIENTE"
                        scraps_count += 1
                    elif estado_final == 'APROBADO':
                        aprobados_count += 1
                    else:
                        rechazados_count += 1
                else:
                    sello_bd = "SIN SELLO (No escaneado)"
                    estado_final = "FALTANTE"
                    estado_ilrl = "FALTANTE"
                    estado_geo = "FALTANTE"
                    faltantes_count += 1

                # Guardar en memoria para la interfaz y para el Excel de empaque
                self.detalles_auditoria[cable_visual] = {
                    'sello': sello_bd, 
                    'final': estado_final,
                    'ilrl': {'estado': estado_ilrl},
                    'geo': {'estado': estado_geo}
                }
                consolidado.append((cable_visual, estado_ilrl, estado_geo, estado_final, sello_bd))

            # Actualizar la interfaz visual. (Intrusos lo pasamos en 0, pues buscamos la secuencia exacta)
            self.app.after(0, self._actualizar_ui, consolidado, aprobados_count, rechazados_count, faltantes_count, 0, scraps_count, total)

        except Exception as e:
            self.app.after(0, lambda: messagebox.showerror("Error", f"Falló la revisión:\n{e}"))
            self.app.after(0, lambda: self.summary_label.config(text="Error en la consulta.", foreground="red"))

    def _actualizar_ui(self, consolidado, aprobados, rechazados, faltantes, intrusos, scraps, total):
        for row in consolidado:
            self.tree.insert("", "end", values=row, tags=(row[3],))
        
        resumen = f"📦 LOTE | Meta: {total} | ✅ OK: {aprobados} | ❌ Rechazos: {rechazados} | ⚠️ Faltantes: {faltantes}"
        if intrusos > 0: resumen += f" | 🚨 Intrusos: {intrusos}"
        if scraps > 0: resumen += f" | 🗑️ Scrap: {scraps}"
        
        self.summary_label.config(text=resumen, foreground="black", font=("Helvetica", 12, "bold"))
        self.btn_excel.config(state=tk.NORMAL)

    def descargar_reporte_excel(self):
        ot_num = re.sub(r'[^0-9]', '', self.ot_var.get())
        ot_completa = f"JMO-{ot_num}"
        total_esperado = int(self.total_var.get()) if self.total_var.get().isdigit() else 0

        home = os.path.expanduser("~")
        desktop_base = os.path.join(home, "Desktop")
        for ruta in [os.path.join(home, "OneDrive", "Escritorio"), os.path.join(home, "OneDrive", "Desktop"), os.path.join(home, "Escritorio")]:
            if os.path.exists(ruta): desktop_base = ruta; break

        carpeta_reportes = os.path.join(desktop_base, "Reportes de Lote Empaque")
        os.makedirs(carpeta_reportes, exist_ok=True)
        ruta_completa = os.path.join(carpeta_reportes, f"Reporte_LOTE_{ot_completa}.xlsx")

        try:
            wb = Workbook()
            ws_resumen = wb.active
            ws_resumen.title = "Resumen de Lote"
            
            aprobados_count = sum(1 for d in self.detalles_auditoria.values() if d.get('final') == 'APROBADO')
            intrusos = [c for c, d in self.detalles_auditoria.items() if d.get('final') == 'INTRUSO']
            scraped_cables = [c for c, d in self.detalles_auditoria.items() if d.get('final') == 'SCRAP' or d.get('sello') == 'SCRAP']
            sin_sello = [c for c, d in self.detalles_auditoria.items() if d.get('final') != 'SCRAP' and ("SIN SELLO" in str(d.get('sello')).upper() or "ERROR" in str(d.get('sello')).upper())]

            ws_resumen['A1'] = f"REVISIÓN DE LOTE (EMPAQUE) - {ot_completa}"
            ws_resumen['A1'].font = Font(size=16, bold=True, color="2C3E50")
            
            ws_resumen['A3'] = "Cantidad Consultada:"; ws_resumen['B3'] = total_esperado
            ws_resumen['A4'] = "Cables Aprobados:"; ws_resumen['B4'] = aprobados_count
            
            ws_detalles = wb.create_sheet("Extracción de Sellos")
            ws_detalles.append(["Cable (N.S.)", "Sello Digital (BD)", "ESTADO FINAL"])
            
            for col in ws_detalles[1]: 
                col.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                col.font = Font(color="FFFFFF", bold=True)
                col.alignment = Alignment(horizontal="center")

            for cable, datos in self.detalles_auditoria.items():
                ws_detalles.append([cable, datos.get('sello', 'N/A'), datos.get('final', 'N/A')])

            for row in ws_detalles.iter_rows(min_row=2, max_col=3):
                for cell in row: cell.alignment = Alignment(horizontal="center")
                if row[2].value == "APROBADO": row[2].font = Font(color="006100", bold=True)
                elif row[2].value in ["RECHAZADO", "INTRUSO"]: row[2].font = Font(color="9C0006", bold=True)
                elif row[2].value == "SCRAP": row[2].font = Font(color="808080", bold=True)

            ws_detalles.column_dimensions['A'].width = 20
            ws_detalles.column_dimensions['B'].width = 25
            ws_detalles.column_dimensions['C'].width = 18

            wb.save(ruta_completa)
            os.startfile(ruta_completa)
            
        except Exception as e:
            messagebox.showerror("Error", f"Fallo al generar el Excel:\n{e}")

class BuscadorSellos_Page(ttk.Frame):
    def __init__(self, parent, app_instance):
        super().__init__(parent, padding=20)
        self.app = app_instance
        self.create_widgets()

    def create_widgets(self):
        container = ttk.Frame(self, style='TFrame')
        container.pack(expand=True, fill='both')

        # --- Cabecera de Búsqueda ---
        header_frame = ttk.Frame(container)
        header_frame.pack(fill='x', pady=10)

        ttk.Label(header_frame, text="Línea:", font=("Helvetica", 12, "bold")).pack(side='left', padx=5)
        self.linea_var = tk.StringVar(value="JWS1-1")
        linea_cb = ttk.Combobox(header_frame, textvariable=self.linea_var, values=["JWS1-1", "JWS1-2", "JWS1-3"], state="readonly", width=10, font=("Helvetica", 12))
        linea_cb.pack(side='left', padx=(0, 15))

        ttk.Label(header_frame, text="Sello Digital (13 dígitos):", font=("Helvetica", 12, "bold")).pack(side='left', padx=5)
        
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._limitar_caracteres)
        self.search_entry = ttk.Entry(header_frame, textvariable=self.search_var, font=("Helvetica", 14), width=20)
        self.search_entry.pack(side='left', padx=10)
        self.search_entry.bind("<Return>", lambda e: self.buscar_sello())
        
        ttk.Button(header_frame, text="Buscar Sello", command=self.buscar_sello, style='primary.TButton', padding=10).pack(side='left', padx=10)

        # --- SECCIÓN DE SCRAP ---
        scrap_frame = ttk.LabelFrame(container, text="Registro de Scrap (Cables Dañados/Irreparables)", padding=10)
        scrap_frame.pack(fill='x', pady=5)

        ttk.Label(scrap_frame, text="N.S. a Scrapear:", font=("Helvetica", 11, "bold")).pack(side='left', padx=5)
        self.scrap_var = tk.StringVar()
        ttk.Entry(scrap_frame, textvariable=self.scrap_var, font=("Helvetica", 12), width=18).pack(side='left', padx=5)

        ttk.Label(scrap_frame, text="Nombre del Auditor:", font=("Helvetica", 11, "bold")).pack(side='left', padx=(15, 5))
        self.auditor_var = tk.StringVar()
        ttk.Entry(scrap_frame, textvariable=self.auditor_var, font=("Helvetica", 12), width=20).pack(side='left', padx=5)

        ttk.Button(scrap_frame, text="🗑️ Mandar a Scrap", command=self.mandar_a_scrap, style='danger.TButton').pack(side='left', padx=15)

        # --- NUEVA SECCIÓN DE RESTAURAR DE SCRAP ---
        restore_frame = ttk.LabelFrame(container, text="Restaurar de Scrap (Revertir estado)", padding=10)
        restore_frame.pack(fill='x', pady=5)

        ttk.Label(restore_frame, text="N.S. a Restaurar:", font=("Helvetica", 11, "bold")).pack(side='left', padx=5)
        self.restore_var = tk.StringVar()
        ttk.Entry(restore_frame, textvariable=self.restore_var, font=("Helvetica", 12), width=18).pack(side='left', padx=5)

        ttk.Label(restore_frame, text="Nombre del Auditor:", font=("Helvetica", 11, "bold")).pack(side='left', padx=(15, 5))
        self.restore_auditor_var = tk.StringVar()
        ttk.Entry(restore_frame, textvariable=self.restore_auditor_var, font=("Helvetica", 12), width=20).pack(side='left', padx=5)

        ttk.Button(restore_frame, text="♻️ Sacar de Scrap", command=self.sacar_de_scrap, style='success.TButton').pack(side='left', padx=15)
        # ------------------------------

        # --- NUEVA SECCIÓN: ELIMINAR ERROR DE DEDO ---
        delete_frame = ttk.LabelFrame(container, text="Eliminar Registro (Error de Dedo en Producción)", padding=10)
        delete_frame.pack(fill='x', pady=5)

        ttk.Label(delete_frame, text="N.S. a Eliminar:", font=("Helvetica", 11, "bold")).pack(side='left', padx=5)
        self.delete_var = tk.StringVar()
        ttk.Entry(delete_frame, textvariable=self.delete_var, font=("Helvetica", 12), width=18).pack(side='left', padx=5)

        ttk.Label(delete_frame, text="Nombre del Auditor:", font=("Helvetica", 11, "bold")).pack(side='left', padx=(15, 5))
        self.delete_auditor_var = tk.StringVar()
        ttk.Entry(delete_frame, textvariable=self.delete_auditor_var, font=("Helvetica", 12), width=20).pack(side='left', padx=5)

        ttk.Button(delete_frame, text="⚠️ Borrar Registro", command=self.borrar_registro, style='danger.Outline.TButton').pack(side='left', padx=15)
        # ---------------------------------------------

        # --- Área de Resultados ---
        self.result_text = tk.Text(container, height=20, width=90, wrap="word", font=("Courier New", 11), state=tk.DISABLED, bg="#f8f9fa")
        self.result_text.pack(fill='both', expand=True, pady=15)

        self.result_text.tag_configure("titulo", font=("Helvetica", 16, "bold"), foreground="#2c3e50")
        self.result_text.tag_configure("subtitulo", font=("Helvetica", 12, "bold"), foreground="#0056b3")
        self.result_text.tag_configure("APROBADO", foreground="#28a745", font=("Helvetica", 14, "bold"))
        self.result_text.tag_configure("RECHAZADO", foreground="#dc3545", font=("Helvetica", 14, "bold"))
        self.result_text.tag_configure("SCRAP", foreground="#dc3545", font=("Helvetica", 14, "bold"))
        self.result_text.tag_configure("normal", font=("Courier New", 11))
        self.result_text.tag_configure("bold", font=("Courier New", 11, "bold"))
    
    def _limitar_caracteres(self, *args):
        texto = self.search_var.get()
        if len(texto) > 16:
            self.search_var.set(texto[:16])

    def mandar_a_scrap(self):
        serie_raw = self.scrap_var.get().strip()
        auditor = self.auditor_var.get().strip()
        
        if not auditor:
            messagebox.showerror("Falta Auditor", "Es obligatorio ingresar el nombre del auditor que autoriza el scrap.", parent=self)
            return

        serie_numerica = re.sub(r'[^0-9]', '', serie_raw)
        if len(serie_numerica) != 13:
            messagebox.showerror("Formato Inválido", "El número de serie debe contener exactamente 13 dígitos para enviarse a Scrap.", parent=self)
            return

        linea_seleccionada = self.linea_var.get()
        db_key = f"db_path_{linea_seleccionada.lower().replace('-', '_')}"
        db_path = self.app.config.get(db_key, '')
        
        if not db_path or not os.path.exists(db_path):
            messagebox.showerror("Error BD", f"Base de datos no encontrada para {linea_seleccionada}.")
            return

        prefijo_serie = "JRMO-" if "JRMO" in serie_raw.upper() else "JMO-"
        serie_completa = f"{prefijo_serie}{serie_numerica}"
        ot_completa = f"JMO-{serie_numerica[:9]}"

        confirmacion = messagebox.askyesno(
            "⚠️ Confirmar Scrap",
            f"¿El auditor {auditor} autoriza mandar el cable:\n\n{serie_completa}\n\na SCRAP en la línea {linea_seleccionada}?",
            parent=self, icon='warning'
        )
        if not confirmacion: return

        try:
            conn = sqlite3.connect(db_path, timeout=10)
            cursor = conn.cursor()
            
            ilrl_details = json.dumps({'status': 'SCRAP CONFIRMADO', 'details': f'Confirmado/Enviado a scrap por el auditor: {auditor}', 'raw_data': []})
            
            cursor.execute("""
                INSERT INTO cable_verifications (
                    entry_date, serial_number, ot_number, overall_status,
                    ilrl_status, ilrl_details, geo_status, geo_details, digital_seal
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                serie_completa, ot_completa, 'SCRAP CONFIRMADO',
                'N/A (SCRAP CONFIRMADO)', ilrl_details, 'N/A (SCRAP CONFIRMADO)', ilrl_details, serie_completa
            ))
            conn.commit()
            conn.close()

            self.scrap_var.set("")
            self.auditor_var.set("")
            
            try: winsound.Beep(300, 800)
            except: pass
            
            messagebox.showinfo("Scrap Registrado", f"Cable {serie_completa} registrado como SCRAP exitosamente.", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar en BD:\n{e}")

    def sacar_de_scrap(self):
        serie_raw = self.restore_var.get().strip()
        auditor = self.restore_auditor_var.get().strip()

        if not auditor:
            messagebox.showerror("Falta Auditor", "Es obligatorio ingresar el nombre del auditor.", parent=self)
            return

        serie_numerica = re.sub(r'[^0-9]', '', serie_raw)
        if len(serie_numerica) != 13:
            messagebox.showerror("Formato Inválido", "El número de serie debe contener exactamente 13 dígitos.", parent=self)
            return

        # Poka-Yoke de Seguridad: Pedir contraseña de Calidad
        pwd = simpledialog.askstring("Autorización Requerida", "Ingrese la contraseña de Calidad para sacar el cable de Scrap:", show='*', parent=self)
        if pwd != "Calidad2024":
            if pwd is not None: # Si el usuario no le dio a "Cancelar"
                messagebox.showerror("Acceso Denegado", "Contraseña incorrecta.", parent=self)
            return

        linea_seleccionada = self.linea_var.get()
        db_key = f"db_path_{linea_seleccionada.lower().replace('-', '_')}"
        db_path = self.app.config.get(db_key, '')

        if not db_path or not os.path.exists(db_path):
            messagebox.showerror("Error BD", f"Base de datos no encontrada para {linea_seleccionada}.")
            return

        try:
            conn = sqlite3.connect(db_path, timeout=10)
            cursor = conn.cursor()

            # 1. Verificar si el cable realmente se encuentra en Scrap actualmente
            cursor.execute("""
                SELECT overall_status FROM cable_verifications 
                WHERE serial_number LIKE ? OR digital_seal LIKE ?
                ORDER BY id DESC LIMIT 1
            """, (f"%{serie_numerica}%", f"%{serie_numerica}%"))
            
            registro = cursor.fetchone()

            if not registro or "SCRAP" not in registro[0]:
                messagebox.showinfo("Aviso", "Este cable NO se encuentra en estado de SCRAP actualmente.", parent=self)
                conn.close()
                return

            # 2. Eliminar de la base de datos los registros de Scrap de ese cable
            # Esto hará que el último registro válido (APROBADO o RECHAZADO) vuelva a ser el estado oficial
            cursor.execute("""
                DELETE FROM cable_verifications 
                WHERE (serial_number LIKE ? OR digital_seal LIKE ?) AND overall_status LIKE '%SCRAP%'
            """, (f"%{serie_numerica}%", f"%{serie_numerica}%"))

            conn.commit()
            conn.close()

            self.restore_var.set("")
            self.restore_auditor_var.set("")

            messagebox.showinfo("Restauración Exitosa", f"El cable ha sido retirado de SCRAP exitosamente por el auditor {auditor}.\n\nEl cable ha recuperado su estado anterior en la línea {linea_seleccionada}.", parent=self)

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo restaurar en BD:\n{e}")

    def buscar_sello(self):
        busqueda_raw = self.search_var.get().strip()
        busqueda_numerica = re.sub(r'[^0-9]', '', busqueda_raw)

        if not busqueda_numerica:
            messagebox.showwarning("Búsqueda Vacía", "Por favor ingresa un número de serie válido.")
            return

        linea_seleccionada = self.linea_var.get()
        db_key = f"db_path_{linea_seleccionada.lower().replace('-', '_')}"
        db_path = self.app.config.get(db_key, '')

        if not db_path or not os.path.exists(db_path):
            messagebox.showerror("Base de Datos no encontrada", f"No se encuentra la base de datos para {linea_seleccionada}.")
            return

        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)

        try:
            conn = sqlite3.connect(db_path, timeout=10)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT entry_date, digital_seal, serial_number, ot_number, overall_status, 
                       ilrl_status, ilrl_details, geo_status, geo_details
                FROM cable_verifications 
                WHERE digital_seal LIKE ? OR serial_number LIKE ?
                ORDER BY id DESC LIMIT 1
            """, (f"%{busqueda_numerica}%", f"%{busqueda_numerica}%"))
            
            registro = cursor.fetchone()
            conn.close()

            if not registro:
                self.result_text.insert(tk.END, f"\nNo se encontró historial para el sello: {busqueda_numerica} en {linea_seleccionada}\n", "RECHAZADO")
                self.result_text.config(state=tk.DISABLED)
                return

            # --- VALIDACIÓN SI ES SCRAP ---
            estado_global = registro['overall_status']
            if "SCRAP" in estado_global:
                self.result_text.insert(tk.END, f"⚠️ CERTIFICADO ANULADO - {estado_global} ({linea_seleccionada}) ⚠️\n", "titulo")
                self.result_text.insert(tk.END, "="*60 + "\n\n")
                self.result_text.insert(tk.END, f"El cable {registro['serial_number']} fue reportado como DAÑADO/IRREPARABLE.\n\n", "SCRAP")
                
                try: 
                    ilrl_dict = json.loads(registro['ilrl_details'])
                    self.result_text.insert(tk.END, f"Motivo: {ilrl_dict.get('details', 'N/A')}\n", "normal")
                except: pass
                
                self.result_text.insert(tk.END, f"Fecha de registro: {registro['entry_date']}\n\n", "normal")
                self.result_text.config(state=tk.DISABLED)
                return
            # ------------------------------

            sello_db = registro['digital_seal']
            sello = registro['serial_number'] if (not sello_db or sello_db == "N/A") else sello_db
            fecha = registro['entry_date']
            ot = registro['ot_number']
            
            try: ilrl_dict = json.loads(registro['ilrl_details']) if registro['ilrl_details'] else {}
            except: ilrl_dict = {}
            try: geo_dict = json.loads(registro['geo_details']) if registro['geo_details'] else {}
            except: geo_dict = {}

            self.result_text.insert(tk.END, f"CERTIFICADO DE PRODUCTO ({linea_seleccionada})\n", "titulo")
            self.result_text.insert(tk.END, "="*60 + "\n\n")
            self.result_text.insert(tk.END, "Sello Digital: ", "bold")
            self.result_text.insert(tk.END, f"{sello}\n", "normal")
            self.result_text.insert(tk.END, "Orden (O.T.):  ", "bold")
            self.result_text.insert(tk.END, f"{ot}\n", "normal")
            self.result_text.insert(tk.END, "Fecha Verif.:  ", "bold")
            self.result_text.insert(tk.END, f"{fecha}\n\n", "normal")
            self.result_text.insert(tk.END, "ESTADO FINAL:  ", "bold")
            self.result_text.insert(tk.END, f"{estado_global}\n\n", estado_global)
            self.result_text.insert(tk.END, "="*60 + "\n\n")

            self.result_text.insert(tk.END, "Resultados IL/RL:\n", "subtitulo")
            self.result_text.insert(tk.END, f"  Estado:  {registro['ilrl_status']}\n", "normal")
            self.result_text.insert(tk.END, f"  Detalle: {ilrl_dict.get('details', 'N/A').replace('Archivo:', '\\n  Archivo:')}\n\n", "normal")

            self.result_text.insert(tk.END, "Resultados Geometría:\n", "subtitulo")
            self.result_text.insert(tk.END, f"  Estado:  {registro['geo_status']}\n", "normal")
            self.result_text.insert(tk.END, f"  Detalle: {geo_dict.get('details', 'N/A').replace('Archivo:', '\\n  Archivo:')}\n\n", "normal")

        except Exception as e:
            self.result_text.insert(tk.END, f"Error al consultar la base de datos:\n{str(e)}", "RECHAZADO")

        self.result_text.config(state=tk.DISABLED)
    def borrar_registro(self):
        serie_raw = self.delete_var.get().strip()
        auditor = self.delete_auditor_var.get().strip()

        if not auditor:
            messagebox.showerror("Falta Auditor", "Es obligatorio ingresar el nombre del auditor.", parent=self)
            return

        serie_numerica = re.sub(r'[^0-9]', '', serie_raw)
        if len(serie_numerica) != 13:
            messagebox.showerror("Formato Inválido", "El número de serie debe contener exactamente 13 dígitos.", parent=self)
            return

        # Poka-Yoke de Seguridad extrema
        pwd = simpledialog.askstring("Autorización Requerida", "Ingrese la contraseña de Calidad para ELIMINAR PERMANENTEMENTE este registro:", show='*', parent=self)
        if pwd != "Calidad2024":
            if pwd is not None:
                messagebox.showerror("Acceso Denegado", "Contraseña incorrecta.", parent=self)
            return

        linea_seleccionada = self.linea_var.get()
        db_key = f"db_path_{linea_seleccionada.lower().replace('-', '_')}"
        db_path = self.app.config.get(db_key, '')

        if not db_path or not os.path.exists(db_path):
            messagebox.showerror("Error BD", f"Base de datos no encontrada para {linea_seleccionada}.")
            return

        try:
            conn = sqlite3.connect(db_path, timeout=10)
            cursor = conn.cursor()

            cursor.execute("SELECT id FROM cable_verifications WHERE serial_number LIKE ? OR digital_seal LIKE ?", (f"%{serie_numerica}%", f"%{serie_numerica}%"))
            registro = cursor.fetchone()

            if not registro:
                messagebox.showinfo("Aviso", "No se encontró ningún registro para este cable.", parent=self)
                conn.close()
                return

            # Eliminar de raíz
            cursor.execute("DELETE FROM cable_verifications WHERE serial_number LIKE ? OR digital_seal LIKE ?", (f"%{serie_numerica}%", f"%{serie_numerica}%"))
            conn.commit()
            conn.close()

            self.delete_var.set("")
            self.delete_auditor_var.set("")

            messagebox.showinfo("Eliminación Exitosa", f"El registro erróneo {serie_numerica} ha sido borrado permanentemente por {auditor}.", parent=self)

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar de la BD:\n{e}")

if __name__ == "__main__":
    app = App()
    app.mainloop()
